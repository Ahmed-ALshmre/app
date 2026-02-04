# -*- coding: utf-8 -*-
"""
Flask web port of your Tkinter Orders Manager.
Single-file app: run with `python app.py` then open http://127.0.0.1:5000

Key features kept:
- Passcode gate (1977)
- XLSX datastore with identical columns and logic
- PDF import (page-by-page) with the same extract rules
- Invoice PDF match -> auto mark Delivered
- Search, barcode (mark Returned), edit/delete, dedupe, move to Shipping with group product name
- Pending list with date filters
- Detailed stats (summary / by price / daily trend)

Folders auto-created under a per-user data dir (like the desktop version).
"""


from __future__ import annotations
import os
import re
import time
import random
import io
import sys
import traceback
from pathlib import Path
from datetime import datetime, date, timedelta
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.utils import secure_filename
from flask import (
    Flask, render_template_string, request, redirect, url_for,
    session, flash, send_from_directory, abort
)

import requests  # تأكد pip install requests
from io import BytesIO
import pandas as pd
import pdfplumber




try:
    import openpyxl  # noqa: F401
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    get_column_letter = None

# ----------------------------- CONFIG ---------------------------------
PASSCODE = "1977"
SECRET_KEY = os.environ.get("SECRET_KEY", "dev-secret-change-me")
# Telegram config (ضع القيم في متغيرات البيئة أو مباشرة للتجربة)
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "8311293130:AAF5ALNUB9DZkJQ6KWoEYSiBedZxZneu6S8")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "-5043262753")  # ID الكروب


# ------------------------- SAFE PATH HELPERS ---------------------------
def is_frozen():
    return getattr(sys, "frozen", False)


def app_dir():
    if is_frozen():
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent


def user_data_dir():
    if os.name == "nt":
        base = os.environ.get("APPDATA") or str(Path.home() / "AppData" / "Roaming")
        p = Path(base) / "OrdersManagerWeb"
    else:
        p = Path.home() / ".local" / "share" / "OrdersManagerWeb"
    p.mkdir(parents=True, exist_ok=True)
    (p / "uploads").mkdir(exist_ok=True)
    return p


def resource_path(*parts):
    return str((app_dir() / Path(*parts)).resolve())

# ------------------------------ STORAGE -------------------------------
STATUS_PROCESSING = "قيد المعالجة"
STATUS_READY = "قيد التجهيز"
STATUS_SHIPPING = "قيد التوصيل"
STATUS_DELIVERED = "تم التوصيل"
STATUS_RETURNED = "راجع"

BASE_COLUMNS = [
    "Product Name",
    "Page Name",
    "Transaction ID",
    "Time and Date",
    "Contact Numbers",
    "Address",
    "Order Price",
    "Status",
    "Status Updated At",
    "Shipping At",
    "Delivered At",
    "Returned At",
    "Return Reason",
    "Notes",
    "Client Orders Count",
    "Items",
]

EXCEL_FILE = str((user_data_dir() / "orders_data.xlsx").resolve())
ERROR_LOG = str((user_data_dir() / "error.log").resolve())
UPLOAD_DIR = str((user_data_dir() / "uploads").resolve())

# ------------------------------ UTILS ---------------------------------

def now_str():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def send_telegram(msg: str):
    """
    إرسال رسالة بسيطة إلى تلغرام.
    يعتمد على TELEGRAM_BOT_TOKEN و TELEGRAM_CHAT_ID من متغيرات البيئة.
    """
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        return  # لو مو متهيئة، نطنش بصمت حتى ما يوقع البرنامج

    if requests is None:
        return  # لو مكتبة requests مو منصبة

    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        requests.post(url, data={
            "chat_id": TELEGRAM_CHAT_ID,
            "text": msg
        }, timeout=5)
    except Exception as e:
        # نسجل الخطأ في اللوج بدون ما نوقف التطبيق
        try:
            _fatal_box("Telegram send failed", e)
        except Exception:
            pass

def send_telegram_document(file_bytes: bytes, filename: str, caption: str = ""):
    """Send a document (Excel/ZIP/etc.) to Telegram."""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        return
    if requests is None:
        return
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
        files = {"document": (filename, file_bytes)}
        data = {"chat_id": TELEGRAM_CHAT_ID}
        if caption:
            data["caption"] = caption
        requests.post(url, data=data, files=files, timeout=20)
    except Exception:
        pass


def build_inventory_daily_excel_bytes(d: str):
    """Return (bytes, filename) for daily inventory movements report."""
    d = (d or date.today().isoformat()).strip()
    try:
        inventory.movements.reload()
        mv = inventory.movements.filter_by_date(d).copy()
    except Exception:
        mv = pd.DataFrame(columns=InventoryMovementStore.COLS)

    if mv is None or mv.empty:
        return None, None

    mv['Delta'] = pd.to_numeric(mv.get('Delta'), errors='coerce').fillna(0).astype(int)

    summary = (mv.groupby(['Product Code','Product Name','Movement Type'])['Delta']
                 .sum().reset_index())
    piv = summary.pivot_table(index=['Product Code','Product Name'],
                              columns='Movement Type',
                              values='Delta',
                              aggfunc='sum',
                              fill_value=0).reset_index()

    totals = {}
    try:
        for c in piv.columns:
            if c not in ['Product Code','Product Name']:
                totals[c] = int(pd.to_numeric(piv[c], errors='coerce').sum() or 0)
    except Exception:
        totals = {}

    out = BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        pd.DataFrame([{
            "Date": d,
            **{str(k): v for k, v in totals.items()},
            "Total Movements Rows": len(mv)
        }]).to_excel(writer, index=False, sheet_name='Summary')
        mv.to_excel(writer, index=False, sheet_name='Movements')
        piv.to_excel(writer, index=False, sheet_name='By Product & Type')
    out.seek(0)
    return out.getvalue(), f"inventory_daily_{d}.xlsx"


def build_withdrawn_daily_summary(d: str):
    """Build a short text summary for WITHDRAW movements for a given day."""
    d = (d or date.today().isoformat()).strip()
    try:
        inventory.movements.reload()
        mv = inventory.movements.filter_by_date(d).copy()
    except Exception:
        mv = pd.DataFrame(columns=InventoryMovementStore.COLS)

    if mv is None or mv.empty:
        return f"📦 ملخص السحب من المخزن\nالتاريخ: {d}\nلا توجد حركات اليوم."

    mv = mv.copy()
    mv['Delta'] = pd.to_numeric(mv.get('Delta'), errors='coerce').fillna(0).astype(int)
    mv['Movement Type'] = mv.get('Movement Type','').astype(str)

    wd = mv[mv['Movement Type'].str.upper().eq('WITHDRAW')].copy()
    if wd.empty:
        return f"📦 ملخص السحب من المخزن\nالتاريخ: {d}\nلا توجد عمليات سحب اليوم."

    wd['Withdraw Qty'] = wd['Delta'].apply(lambda x: abs(int(x)))

    total_withdraw = int(wd['Withdraw Qty'].sum() or 0)
    top = (wd.groupby(['Product Code','Product Name'])['Withdraw Qty']
             .sum().sort_values(ascending=False).head(10).reset_index())

    lines = []
    lines.append("📦 ملخص السحب من المخزن")
    lines.append(f"📅 التاريخ: {d}")
    lines.append(f"✅ إجمالي المسحوب: {total_withdraw} قطعة")
    lines.append("")
    lines.append("🔥 أعلى 10 منتجات مسحوبة:")
    for i, r in top.iterrows():
        lines.append(f"{i+1}) {r['Product Name']} ({r['Product Code']}): {int(r['Withdraw Qty'])} قطعة")
    return "\n".join(lines)



def normalize_digits(s: str) -> str:
    if s is None:
        return ""
    trans = {
        ord('٠'): '0', ord('١'): '1', ord('٢'): '2', ord('٣'): '3', ord('٤'): '4',
        ord('٥'): '5', ord('٦'): '6', ord('٧'): '7', ord('٨'): '8', ord('٩'): '9',
        ord('۰'): '0', ord('۱'): '1', ord('۲'): '2', ord('۳'): '3', ord('۴'): '4',
        ord('۵'): '5', ord('۶'): '6', ord('۷'): '7', ord('۸'): '8', ord('۹'): '9',
        ord('\u066C'): ',',  # ARABIC THOUSANDS SEPARATOR -> ,
        ord('\u200f'): None, ord('\u200e'): None,  # RLM/LRM
    }
    return str(s).translate(trans)


def to_int(num_str: str):
    if not num_str:
        return None
    s = normalize_digits(num_str).replace(",", "").replace(" ", "")
    if not re.search(r'\d', s):
        return None
    try:
        return int(re.search(r'(\d+)', s).group(1))
    except Exception:
        return None


class DataStore:
    def __init__(self, path):
        self.path = path
        self._last_mtime = None
        self.df = self._load_or_create()
        self._ensure_index()
        self._touch_mtime()

    def _load_or_create(self):
        path = Path(self.path)
        if not path.exists():
            df = pd.DataFrame(columns=BASE_COLUMNS)
            try:
                with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Sheet1")
                    if get_column_letter is not None:
                        ws = writer.sheets["Sheet1"]
                        tid_idx = BASE_COLUMNS.index("Transaction ID") + 1
                        for cell in ws[get_column_letter(tid_idx)]:
                            cell.number_format = "@"
            except Exception:
                df.to_excel(self.path, index=False)
            return df

        try:
            df = pd.read_excel(self.path, dtype=str)
        except Exception:
            df = pd.read_excel(self.path)
            if "Transaction ID" in df.columns:
                df["Transaction ID"] = df["Transaction ID"].astype(str)

        if "Product Name" not in df.columns:
            if "Title" in df.columns:
                df.rename(columns={"Title": "Product Name"}, inplace=True)
            else:
                df["Product Name"] = pd.NA
        if "Order Price" not in df.columns:
            df["Order Price"] = pd.NA
        for old_col in ["Delivery Type", "Delivery Cost", "Pieces Count", "Page Number"]:
            if old_col in df.columns:
                df.drop(columns=[old_col], inplace=True)
        for c in BASE_COLUMNS:
            if c not in df.columns:
                df[c] = pd.NA
        df["Transaction ID"] = df["Transaction ID"].astype(str).str.strip()
        df["Order Price"] = pd.to_numeric(df["Order Price"], errors="coerce")
        df["Status"] = df["Status"].fillna(STATUS_READY)
        # ensure new columns ordering
        df = df[BASE_COLUMNS]
        return df

    def _ensure_index(self):
        if "Transaction ID" not in self.df.columns:
            self.df["Transaction ID"] = ""
        try:
            self.df.set_index("Transaction ID", drop=False, inplace=True)
        except Exception:
            pass

    def _touch_mtime(self):
        """Track file mtime for lightweight reload checks."""
        try:
            self._last_mtime = os.path.getmtime(self.path)
        except Exception:
            self._last_mtime = None

    def reload_if_changed(self):
        """Reload excel file only if it changed on disk (prevents stale counts after reload)."""
        try:
            current = os.path.getmtime(self.path)
        except Exception:
            return
        if self._last_mtime is None:
            self._last_mtime = current
            return
        if current != self._last_mtime:
            try:
                self.df = self._load_or_create()
                self._ensure_index()
            finally:
                self._touch_mtime()


    def save(self):
        to_save = self.df.reset_index(drop=True).copy()
        to_save["Transaction ID"] = to_save["Transaction ID"].astype(str)
        try:
            with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
                to_save.to_excel(writer, index=False, sheet_name="Sheet1")
                if get_column_letter is not None:
                    ws = writer.sheets["Sheet1"]
                    tid_idx = BASE_COLUMNS.index("Transaction ID") + 1
                    for cell in ws[get_column_letter(tid_idx)]:
                        cell.number_format = "@"
        except Exception:
            to_save.to_excel(self.path, index=False)
        try:
            self._touch_mtime()
        except Exception:
            pass


    def exists(self, txn):
        return str(txn).strip() in self.df.index

    def get_row(self, txn):
        txn = str(txn).strip()
        if self.exists(txn):
            return self.df.loc[txn]
        return None

    def upsert_row(self, row_dict: dict):
        txn = str(row_dict.get("Transaction ID", "")).strip()
        if not txn or not re.fullmatch(r'\d{6,}', txn):
            return False, "Transaction ID غير صالح (أرقام فقط وبحد أدنى 6 خانات)."
        row_dict = row_dict.copy()
        if not row_dict.get("Status"):
            row_dict["Status"] = STATUS_READY
        for c in BASE_COLUMNS:
            if c not in row_dict:
                row_dict[c] = pd.NA
        if self.exists(txn):
            for k, v in row_dict.items():
                self.df.at[txn, k] = v
            return True, "تم التحديث"
        else:
            new_df = pd.DataFrame([row_dict], columns=BASE_COLUMNS)
            new_df["Transaction ID"] = new_df["Transaction ID"].astype(str).str.strip()
            new_df.set_index("Transaction ID", drop=False, inplace=True)
            self.df = pd.concat([self.df, new_df], axis=0, ignore_index=False)
            return True, "تمت الإضافة"

    def add_bulk(self, rows_list):
        """دالة لإضافة مجموعة صفوف دفعة واحدة لتسريع العملية"""
        if not rows_list:
            return

        # تحويل القائمة إلى DataFrame مرة واحدة
        new_df = pd.DataFrame(rows_list, columns=BASE_COLUMNS)

        # ضبط العمود المفتاحي
        new_df["Transaction ID"] = new_df["Transaction ID"].astype(str).str.strip()
        new_df.set_index("Transaction ID", drop=False, inplace=True)

        # استبعاد الأعمدة الفارغة تماماً لتجنب التحذير (FutureWarning)
        new_df = new_df.dropna(axis=1, how='all')

        # دمج البيانات مرة واحدة فقط
        self.df = pd.concat([self.df, new_df], axis=0, ignore_index=False)
        self.save()
    def update_status(self, txn, new_status, return_reason=None):
        # نتعامل مع رقم الشحنة كـ Transaction ID (نصي)
        txn = str(txn).strip()

        # تأكد أن الشحنة موجودة
        if not self.exists(txn):
            return False, "الشحنة غير موجودة"

        # الحالة القديمة قبل التغيير
        old_status = self.df.at[txn, "Status"] if "Status" in self.df.columns else None

        # حدّث الحالة الجديدة
        self.df.at[txn, "Status"] = new_status


        # ⏱️ توثيق تاريخ/وقت آخر تحديث للحالة (مهم للإحصائيات)
        ts = now_str()
        if "Status Updated At" in self.df.columns:
            self.df.at[txn, "Status Updated At"] = ts
        if new_status == STATUS_SHIPPING and "Shipping At" in self.df.columns:
            self.df.at[txn, "Shipping At"] = ts
        if new_status == STATUS_DELIVERED and "Delivered At" in self.df.columns:
            self.df.at[txn, "Delivered At"] = ts
        if new_status == STATUS_RETURNED and "Returned At" in self.df.columns:
            self.df.at[txn, "Returned At"] = ts
        # لو راجع ومعاه سبب
        if new_status == STATUS_RETURNED and return_reason:
            self.df.at[txn, "Return Reason"] = return_reason

        # صف الشحنة بعد التحديث
        row = self.df.loc[txn]

        # احفظ في الإكسل
        self.save()

        ret = {"msg": "تم تحديث الحالة", "old": old_status, "new": new_status, "row": row}

        # الهُوك الخاص بالمخزن (ينقص/يزيد الكمية)
        try:
            adjust_inventory_on_transition(row, old_status, new_status)
        except Exception:
            pass

        return True, ret

    def drop_by_txn(self, txn):
        txn = str(txn).strip()
        if not self.exists(txn):
            return 0
        self.df = self.df.drop(index=txn)
        return 1

    def drop_duplicates_keep_last(self):
        before = len(self.df)
        self.df = (
            self.df.reset_index(drop=True)
                   .drop_duplicates(subset=["Transaction ID"], keep="last")
        )
        self._ensure_index()
        after = len(self.df)
        return before - after

    def stats_global(self, df=None):
        d = self.df if df is None else df
        d = d.copy()
        d["Order Price"] = pd.to_numeric(d.get("Order Price"), errors="coerce")
        total_orders = len(d)
        total_amount = float(d["Order Price"].sum() or 0)

        def _count(status):
            return int((d["Status"] == status).sum()) if "Status" in d.columns else 0

        def _amount(status):
            if d.empty or "Status" not in d.columns:
                return 0.0
            return float(d.loc[d["Status"] == status, "Order Price"].sum() or 0)

        delivered = _count(STATUS_DELIVERED)
        returned  = _count(STATUS_RETURNED)
        shipping  = _count(STATUS_SHIPPING)
        ready     = _count(STATUS_READY)

        delivered_amt = _amount(STATUS_DELIVERED)
        returned_amt  = _amount(STATUS_RETURNED)
        shipping_amt  = _amount(STATUS_SHIPPING)
        ready_amt     = _amount(STATUS_READY)

        pct = lambda x: (x / total_orders * 100) if total_orders else 0.0
        return {
            "العدد الكلي للطلبات": total_orders,
            "المجموع المالي (Order Price)": total_amount,

            f"عدد {STATUS_DELIVERED}": delivered,
            f"مبلغ {STATUS_DELIVERED}": delivered_amt,

            f"عدد {STATUS_RETURNED}": returned,
            f"مبلغ {STATUS_RETURNED}": returned_amt,

            f"عدد {STATUS_SHIPPING}": shipping,
            f"مبلغ {STATUS_SHIPPING}": shipping_amt,

            f"عدد {STATUS_READY}": ready,
            f"مبلغ {STATUS_READY}": ready_amt,

            f"نسبة {STATUS_DELIVERED} %": round(pct(delivered), 2),
            f"نسبة {STATUS_RETURNED} %": round(pct(returned), 2),
            f"نسبة {STATUS_SHIPPING} %": round(pct(shipping), 2),
            f"نسبة {STATUS_READY} %": round(pct(ready), 2),
        }


    def stats_by_product_price(self, df=None):
        d = self.df if df is None else df
        d = d.copy()
        d["Order Price"] = pd.to_numeric(d["Order Price"], errors="coerce")
        cols = [
            "السعر", "عدد الطلبات",
            STATUS_DELIVERED, STATUS_RETURNED, STATUS_SHIPPING, STATUS_READY,
            "المبلغ المُسلَّم", "نسبة الراجع %"
        ]
        if d.empty or d["Order Price"].isna().all():
            return pd.DataFrame(columns=cols)
        rows = []
        for price, g in d.groupby("Order Price", dropna=False):
            total = len(g)
            delivered = (g["Status"] == STATUS_DELIVERED).sum()
            returned = (g["Status"] == STATUS_RETURNED).sum()
            shipping = (g["Status"] == STATUS_SHIPPING).sum()
            ready = (g["Status"] == STATUS_READY).sum()
            delivered_amount = pd.to_numeric(
                g.loc[g["Status"] == STATUS_DELIVERED, "Order Price"], errors="coerce"
            ).sum()
            return_rate = (returned / total * 100) if total else 0.0
            rows.append({
                "السعر": price,
                "عدد الطلبات": total,
                STATUS_DELIVERED: delivered,
                STATUS_RETURNED: returned,
                STATUS_SHIPPING: shipping,
                STATUS_READY: ready,
                "المبلغ المُسلَّم": float(delivered_amount or 0),
                "نسبة الراجع %": round(return_rate, 2),
            })
        out_df = pd.DataFrame(rows, columns=cols)
        if not out_df.empty:
            out_df = out_df.sort_values(
                by=["المبلغ المُسلَّم", "عدد الطلبات"],
                ascending=[False, False],
                na_position="last"
            )
        return out_df

    def daily_trend(self, df=None):
        d = self.df if df is None else df
        d = d.copy()
        d["Time and Date"] = pd.to_datetime(d["Time and Date"], errors="coerce")
        d = d.dropna(subset=["Time and Date"])
        d["Date"] = d["Time and Date"].dt.date
        daily = d.groupby("Date").size().reset_index(name="Order Count").sort_values("Date")
        daily["Trend"] = daily["Order Count"].diff().apply(
            lambda x: "ارتفاع" if x and x > 0 else ("انخفاض" if x and x < 0 else "ثابت")
        )
        return daily

_data_root = Path(EXCEL_FILE).parent
# ------------------------------ APP -----------------------------------
app = Flask(__name__)
app.secret_key = SECRET_KEY
store = DataStore(EXCEL_FILE)

# --------------------------- TEMPLATES --------------------------------
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["200 per hour"]
)

INVENTORY_HTML = r"""
{% extends 'base.html' %}
{% block content %}

<style>
  .inv-card {
    border-radius: 1rem;
    box-shadow: 0 2px 10px rgba(0,0,0,.06);
  }
  .inv-badge-pill {
    border-radius: 999px;
    font-size: .75rem;
    padding: .15rem .6rem;
  }
  .table-inventory thead th {
    white-space: nowrap;
    font-size: .85rem;
  }
</style>

<div class="row g-3">

  <!-- Dashboard: Top selling + Delivery/Return rates -->
  <div class="col-12">
    <div class="inv-card p-3 bg-white">
      <div class="d-flex flex-wrap justify-content-between align-items-center gap-2">
        <div>
          <h5 class="mb-1">لوحة المخزن</h5>
          <div class="text-muted small">أفضل المنتجات مبيعاً + نسبة الواصل والراجع (حسب القطع)</div>
        </div>
        <div class="d-flex flex-wrap gap-2">
          <a class="btn btn-outline-success btn-sm" href="{{ url_for('telegram_send_inventory_daily') }}">ارسال تقرير المخزن اليومي للتلكرام</a>
          <a class="btn btn-outline-primary btn-sm" href="{{ url_for('telegram_send_withdrawn_daily') }}">ارسال ملخص السحب اليومي للتلكرام</a>
        </div>
      </div>

      <div class="row g-2 mt-2">
        <div class="col-md-3">
          <div class="p-3 border rounded-4">
            <div class="text-muted small mb-1">الواصل (Delivered)</div>
            <div class="fs-5 fw-bold">{{ overall_rates.delivered_pieces }}</div>
            <div class="text-muted small">{{ overall_rates.delivered_pct }}%</div>
          </div>
        </div>
        <div class="col-md-3">
          <div class="p-3 border rounded-4">
            <div class="text-muted small mb-1">الراجع (Returned)</div>
            <div class="fs-5 fw-bold">{{ overall_rates.returned_pieces }}</div>
            <div class="text-muted small">{{ overall_rates.returned_pct }}%</div>
          </div>
        </div>
        <div class="col-md-6">
          <div class="p-3 border rounded-4 h-100">
            <div class="text-muted small">أفضل 10 منتجات مبيعاً (حسب Delivered Pieces)</div>
            <div class="table-responsive mt-2">
              <table class="table table-sm mb-0">
                <thead>
                  <tr>
                    <th>#</th>
                    <th>المنتج</th>
                    <th>Code</th>
                    <th>وصل</th>
                    <th>راجع</th>
                    <th>نسبة الواصل</th>
                    <th>نسبة الراجع</th>
                  </tr>
                </thead>
                <tbody>
                  {% for p in top_selling %}
                  <tr>
                    <td>{{ loop.index }}</td>
                    <td>{{ p['Product Name'] }}</td>
                    <td>{{ p['Product Code'] }}</td>
                    <td class="fw-bold">{{ p['Delivered Pieces'] }}</td>
                    <td class="fw-bold">{{ p['Returned Pieces'] }}</td>
                    <td>{{ p['Delivered %'] }}%</td>
                    <td>{{ p['Returned %'] }}%</td>
                  </tr>
                  {% endfor %}
                  {% if not top_selling %}
                  <tr><td colspan="7" class="text-muted text-center py-3">لا توجد بيانات كافية بعد</td></tr>
                  {% endif %}
                </tbody>
              </table>
            </div>
          </div>
        </div>
      </div>

    </div>
  </div>



  {# حساب بسيط لإجمالي الموديلات والكمية #}
  {% set total_models = rows|length %}
  {% set ns = namespace(total_qty=0) %}
  {% for r in rows %}
    {% set ns.total_qty = ns.total_qty + (r['Quantity'] or 0)|int %}
  {% endfor %}

  <!-- العمود الرئيسي: جدول المخزن -->
  <div class="col-xl-8">
    <div class="card inv-card p-3 h-100">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <div>
          <h5 class="mb-0">المخزن</h5>
          <small class="text-muted">
            عرض كل الموديلات المسجّلة مع حالة الكميات.
          </small>
          <div class="mt-2 d-flex flex-wrap gap-2">
            <a class="btn btn-sm btn-outline-dark" href="{{ url_for('inventory_daily_report') }}">تحميل تقرير اليوم</a>
            <a class="btn btn-sm btn-outline-danger" href="{{ url_for('inventory_stagnant') }}">البضاعة الراكدة</a>
          </div>
        </div>
        <div class="d-flex flex-column align-items-end">
          <span class="badge bg-secondary mb-1">
            عدد الموديلات: {{ total_models }}
          </span>
          <span class="badge bg-dark">
            إجمالي الكمية: {{ ns.total_qty }}
          </span>
        </div>
      </div>

      <!-- بحث داخل المخزن -->
      <form method="get" action="{{ url_for('inventory_home') }}" class="row g-2 mt-2">
        <div class="col-8">
          <input name="q" value="{{ q or '' }}" class="form-control" placeholder="بحث بالاسم أو الكود أو النوع">
        </div>
        <div class="col-4 d-grid">
          <button class="btn btn-outline-secondary">بحث</button>
        </div>
      </form>

      <div class="table-responsive mt-2">
        <table class="table table-hover table-striped align-middle table-inventory mb-0">
          <thead class="table-light">
            <tr>
              <th>الكود</th>
              <th>الاسم</th>
              <th>النوع</th>
              <th>الكمية</th>
              <th>متر/قطعة</th>
              <th>سعر متر الخام</th>
              <th>الخياطة</th>
              <th>الإكسسوارات</th>
              <th>تكاليف إضافية</th>
              <th>سعر البيع</th>
              <th class="text-center">إجراءات</th>
            </tr>
          </thead>
          <tbody>
            {% for r in rows %}
              {% set qty = (r['Quantity'] or 0)|int %}
              {% if qty <= 0 %}
                {% set row_class = 'table-danger' %}
              {% elif qty <= 5 %}
                {% set row_class = 'table-warning' %}
              {% else %}
                {% set row_class = '' %}
              {% endif %}

              <tr class="{{ row_class }}">
                <td class="fw-bold">{{ r['Product Code'] }}</td>
                <td>{{ r['Product Name'] }}</td>
                <td>{{ r['Type'] }}</td>
                <td>
                  <div class="d-flex flex-column">
                    <span>{{ qty }}</span>
                    {% if qty <= 0 %}
                      <span class="inv-badge-pill bg-danger text-white mt-1">
                        منتهية
                      </span>
                    {% elif qty <= 5 %}
                      <span class="inv-badge-pill bg-warning text-dark mt-1">
                        كمية قليلة
                      </span>
                    {% endif %}
                  </div>
                </td>
                <td>{{ r['Meters per Unit'] }}</td>
                <td>{{ r['Fabric Meter Price'] }}</td>
                <td>{{ r['Sewing Cost'] }}</td>
                <td>{{ r['Accessories Cost'] }}</td>
                <td>{{ r['Extra Costs'] }}</td>
                <td>{{ r['Sale Price'] }}</td>
                <td class="text-center text-nowrap">
                  <button class="btn btn-sm btn-success"
                          data-bs-toggle="modal"
                          data-bs-target="#addQtyModal"
                          data-code="{{ r['Product Code'] }}" data-pname="{{ r['Product Name'] }}">
                    + كمية
                  </button>
                  <button class="btn btn-sm btn-outline-danger ms-1"
                          data-bs-toggle="modal"
                          data-bs-target="#takeQtyModal"
                          data-code="{{ r['Product Code'] }}" data-pname="{{ r['Product Name'] }}">
                    - كمية
                  </button>
                  <a href="{{ url_for('inventory_product', code=r['Product Code']) }}" class="btn btn-sm btn-outline-secondary ms-1">تفاصيل</a>
                  <a href="{{ url_for('inventory_edit', code=r['Product Code']) }}"
                     class="btn btn-sm btn-outline-primary ms-1">
                    تعديل
                  </a>
                  <a href="{{ url_for('inventory_delete', code=r['Product Code']) }}"
                     class="btn btn-sm btn-outline-dark ms-1"
                     onclick="return confirm('هل أنت متأكد من حذف هذا المنتج من المخزن؟');">
                    حذف
                  </a>
                </td>
              </tr>
            {% endfor %}

            {% if not rows %}
            <tr>
              <td colspan="11" class="text-center text-muted py-3">
                لا توجد منتجات في المخزن حاليًا.
              </td>
            </tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- العمود الجانبي: إضافة كمية/منتج للمخزن -->
  <div class="col-xl-4">
    <div class="card inv-card p-3">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <div>
          <h6 class="mb-0">إضافة للمخزن</h6>
          <small class="text-muted">
            اختيار منتج جاهز من إدارة المنتجات ثم تحديد الكمية فقط.
          </small>
        </div>
        <a href="{{ url_for('products_home') }}" class="small text-decoration-none">
          إدارة المنتجات المفصّلة
        </a>
      </div>

      <form method="post" action="{{ url_for('inventory_add') }}" class="row g-2 mt-1">
        <div class="col-12">
          <label class="form-label">المنتج / الموديل</label>
          <select required name="code" id="invProduct" class="form-select">
            <option value="">— اختر المنتج من القائمة —</option>
            {% for p in all_rows %}
              <option value="{{ p['Product Name'] }}"
                      data-type="{{ p['Type'] }}">
                {{ p['Product Name'] }}
                {% if p['Type'] %} — {{ p['Type'] }}{% endif %}
                {% if p['Product Code'] %} ({{ p['Product Code'] }}){% endif %}
              </option>
            {% endfor %}
          </select>
        </div>

        {# نوع البضاعة يُملأ تلقائياً (مخفي) #}
        <input type="hidden" name="type" id="invType">
        <input type="hidden" name="pname" id="invPName">

        <div class="col-6">
          <label class="form-label">الكمية</label>
          <input name="qty" type="number" class="form-control" value="0"
                 min="0" inputmode="numeric" pattern="[0-9]*">
        </div>

        {# باقي الحقول غير مطلوبة من المستخدم، نرسلها صفر حتى لا نكسر الكود الخلفي #}
        <input type="hidden" name="mpu"           value="0">
        <input type="hidden" name="fabric_price"  value="0">
        <input type="hidden" name="sew"           value="0">
        <input type="hidden" name="access"        value="0">
        <input type="hidden" name="extra"         value="0">
        <input type="hidden" name="price"         value="0">
        <input type="hidden" name="fabric"        value="0">

        <div class="col-12 mt-1">
          <button class="btn btn-dark w-100">إضافة للمخزن</button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Modal: إضافة كمية -->
<div class="modal fade" id="addQtyModal" tabindex="-1">
  <div class="modal-dialog">
    <form method="post" action="{{ url_for('inventory_adjust_bulk') }}" class="modal-content">
      <div class="modal-header">
        <h6 class="modal-title">إضافة كمية للمخزن</h6>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body">
        <input type="hidden" name="code" id="addQtyCode">
        <div class="mb-2">
          <label class="form-label">الكمية التي ستُضاف</label>
          <input required name="qty" type="number" class="form-control"
                 value="1" min="1" inputmode="numeric" pattern="[0-9]*" autofocus>
        </div>
      </div>
      <div class="modal-footer">
        <button class="btn btn-success">إضافة</button>
      </div>
    </form>
  </div>
</div>

<!-- Modal: سحب كمية -->
<div class="modal fade" id="takeQtyModal" tabindex="-1">
  <div class="modal-dialog">
    <form method="post" action="{{ url_for('inventory_adjust_bulk') }}" class="modal-content">
      <div class="modal-header">
        <h6 class="modal-title">سحب كمية من المخزن</h6>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body">
        <input type="hidden" name="code" id="takeQtyCode">
        <div class="mb-2">
          <label class="form-label">الكمية التي ستُسحب</label>
          <input required name="qty" type="number" class="form-control"
                 value="-1" step="1" inputmode="numeric" pattern="-?[0-9]*" autofocus>
          <div class="form-text">استخدم قيمة سالبة للسحب (مثال: -5)</div>
        </div>
      </div>
      <div class="modal-footer">
        <button class="btn btn-danger">سحب</button>
      </div>
    </form>
  </div>
</div>

<!-- Feedback Modal (after redirect) -->
<div class="modal fade" id="feedbackModal" tabindex="-1">
  <div class="modal-dialog">
    <div class="modal-content">
      <div class="modal-header">
        <h6 class="modal-title">تحديث المخزن</h6>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body">
        {% if added %}
          تم إضافة <b>{{ added }}</b> قطعة إلى المنتج <b>{{ name }}</b>.
        {% elif taken %}
          تم سحب <b>{{ taken }}</b> قطعة من المنتج <b>{{ name }}</b>.
        {% endif %}
      </div>
      <div class="modal-footer">
        <button class="btn btn-secondary" data-bs-dismiss="modal">إغلاق</button>
      </div>
    </div>
  </div>
</div>

<script>
  // ربط اسم المنتج في مودال + كمية
  const addQtyModal = document.getElementById('addQtyModal');
  addQtyModal?.addEventListener('show.bs.modal', event => {
    const btn = event.relatedTarget;
    const code = btn.getAttribute('data-code');
    document.getElementById('addQtyCode').value = code;
  });

  // ربط اسم المنتج في مودال - كمية
  const takeQtyModal = document.getElementById('takeQtyModal');
  takeQtyModal?.addEventListener('show.bs.modal', event => {
    const btn = event.relatedTarget;
    const code = btn.getAttribute('data-code');
    document.getElementById('takeQtyCode').value = code;
  });

  // تعبئة نوع المنتج تلقائياً عند اختيار المنتج من القائمة
  const invProduct = document.getElementById('invProduct');
  const invType    = document.getElementById('invType');
  const invPName  = document.getElementById('invPName');
  invProduct?.addEventListener('change', () => {
    const opt = invProduct.selectedOptions[0];
    if (opt) {
      invType.value = opt.getAttribute('data-type') || '';
      invPName.value = opt.getAttribute('data-name') || '';
    }
  });

  // auto show feedback if present
  {% if added or taken %}
  const fb = new bootstrap.Modal(document.getElementById('feedbackModal'));
  fb.show();
  {% endif %}
</script>

{% endblock %}
"""


INVENTORY_PRODUCT_HTML = r"""
{% extends 'base.html' %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <div>
    <h5 class="mb-0">تفاصيل المنتج</h5>
    <div class="text-muted small">{{ item['Product Name'] }} - {{ item['Product Code'] }}</div>
  </div>
  <div class="d-flex gap-2">
    <a class="btn btn-outline-secondary" href="{{ url_for('inventory_home') }}">رجوع</a>
    <a class="btn btn-dark" href="{{ url_for('inventory_daily_report') }}">تقرير اليوم</a>
  </div>
</div>

<div class="row g-3">
  <div class="col-lg-5">
    <div class="card p-3">
      <h6 class="mb-2">إحصائيات المنتج</h6>
      <div class="table-responsive">
        <table class="table table-sm align-middle mb-0">
          <tbody>
            <tr><td>الكمية الحالية</td><td class="fw-bold">{{ stats['Current Quantity'] }}</td></tr>
            <tr><td>إجمالي الإنتاج</td><td class="fw-bold">{{ stats['Total Produced'] }}</td></tr>
            <tr><td>إجمالي المسحوب</td><td class="fw-bold">{{ stats['Total Withdrawn'] }}</td></tr>
            <tr><td>إجمالي المُرجع للمخزن</td><td class="fw-bold">{{ stats['Total Returned (to stock)'] }}</td></tr>
            <tr><td>قطع تم التوصيل</td><td class="fw-bold">{{ stats['Delivered Pieces'] }}</td></tr>
            <tr><td>قطع راجع</td><td class="fw-bold">{{ stats['Returned Pieces'] }}</td></tr>
            <tr><td>نسبة المُسلّم</td><td class="fw-bold">{{ stats['Delivered %'] }}%</td></tr>
            <tr><td>نسبة الراجع</td><td class="fw-bold">{{ stats['Returned %'] }}%</td></tr>
            <tr><td>آخر سحب</td><td class="fw-bold">{{ stats['Last Withdraw DateTime'] or '—' }}</td></tr>
          </tbody>
        </table>
      </div>
    </div>

    <div class="card p-3 mt-3">
      <h6 class="mb-2">تعديل سريع</h6>
      <form method="post" action="{{ url_for('inventory_adjust_bulk') }}" class="row g-2">
        <input type="hidden" name="code" value="{{ item['Product Code'] }}">
        <div class="col-12">
          <label class="form-label">+ إضافة / - سحب</label>
          <input name="qty" type="number" class="form-control" value="1" step="1" required>
          <div class="form-text">سيتم تسجيل الحركة تلقائياً.</div>
        </div>
        <div class="col-12 d-grid">
          <button class="btn btn-primary">تنفيذ</button>
        </div>
      </form>
    </div>
  </div>

  <div class="col-lg-7">
    <div class="card p-3">
      <h6 class="mb-2">سجل الحركات (آخر 50)</h6>
      <div class="table-responsive">
        <table class="table table-sm table-striped align-middle mb-0">
          <thead>
            <tr>
              <th>الوقت</th>
              <th>النوع</th>
              <th>التغير</th>
              <th>مرجع</th>
              <th>ملاحظات</th>
            </tr>
          </thead>
          <tbody>
            {% for m in moves %}
            <tr>
              <td>{{ m['DateTime'] }}</td>
              <td>{{ m['Movement Type'] }}</td>
              <td class="fw-bold {% if (m['Delta']|int) < 0 %}text-danger{% else %}text-success{% endif %}">{{ m['Delta'] }}</td>
              <td>{{ m['Ref'] }}</td>
              <td>{{ m['Notes'] }}</td>
            </tr>
            {% endfor %}
            {% if not moves %}
            <tr><td colspan="5" class="text-center text-muted py-3">لا توجد حركات لهذا المنتج.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>
{% endblock %}
"""

# (Bootstrap from CDN; RTL-friendly)

PRODUCTS_HTML = r"""
{% extends 'base.html' %}
{% block content %}
<div class="row g-3">
  <div class="col-xl-7">
    <div class="card p-3">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <h5 class="mb-0">إدارة المنتجات</h5>
        <span class="badge bg-secondary">عدد المنتجات: {{ rows|length }}</span>
      </div>
      <div class="table-responsive">
        <table class="table table-striped align-middle">
          <thead>
            <tr>
              <th>الكود</th>
              <th>المنتج</th>
              <th>النوع</th>
              <th>متر/قطعة</th>
              <th>سعر متر الخام</th>
              <th>الخياطة</th>
              <th>الإكسسوارات</th>
              <th>تكاليف إضافية</th>
              <th>سعر البيع</th>
              <th>الكمية بالمخزن</th>
            </tr>
          </thead>
          <tbody>
            {% for r in rows %}
            <tr>
              <td>{{ r['Product Code'] }}</td>
              <td>{{ r['Product Name'] }}</td>
              <td>{{ r['Type'] }}</td>
              <td>{{ r['Meters per Unit'] }}</td>
              <td>{{ r['Fabric Meter Price'] }}</td>
              <td>{{ r['Sewing Cost'] }}</td>
              <td>{{ r['Accessories Cost'] }}</td>
              <td>{{ r['Extra Costs'] }}</td>
              <td>{{ r['Sale Price'] }}</td>
              <td>{{ r['Quantity'] }}</td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <div class="col-xl-5">
    <div class="card p-3">
      <h6 class="mb-3">إضافة منتج جديد (تفاصيل الموديل)</h6>
      <form method="post" action="{{ url_for('products_add') }}" class="row g-2">
        <div class="col-12">
          <label class="form-label">اسم المنتج / الموديل</label>
          <input required name="name" class="form-control" autocomplete="off">
        </div>
        <div class="col-12">
          <label class="form-label">نوع البضاعة</label>
          <select name="type" class="form-select">
            <option value="">—</option>
            <option>عباية</option>
            <option>ملابس أطفال</option>
            <option>نساء</option>
            <option>سوت</option>
          </select>
        </div>
        <div class="col-6">
          <label class="form-label">متر/قطعة (عدد الأمتار المطلوبة)</label>
          <input name="mpu" type="number" step="0.01" class="form-control" value="0">
        </div>
        <div class="col-6">
          <label class="form-label">سعر متر الخام</label>
          <input name="fabric_price" type="number" step="0.01" class="form-control" value="0">
        </div>
        <div class="col-6">
          <label class="form-label">سعر الخياطة/قطعة</label>
          <input name="sew" type="number" step="0.01" class="form-control" value="0">
        </div>
        <div class="col-6">
          <label class="form-label">تكلفة الإكسسوارات</label>
          <input name="access" type="number" step="0.01" class="form-control" value="0">
        </div>
        <div class="col-6">
          <label class="form-label">تكاليف إضافية</label>
          <input name="extra" type="number" step="0.01" class="form-control" value="0">
        </div>
        <div class="col-6">
          <label class="form-label">سعر البيع</label>
          <input name="price" type="number" step="0.01" class="form-control" value="0">
        </div>
        <div class="col-6">
          <label class="form-label">الكمية الابتدائية بالمخزن</label>
          <input name="qty" type="number" class="form-control" value="0">
        </div>
        <div class="col-6">
          <label class="form-label">إجمالي أمتار القماش (اختياري)</label>
          <input name="fabric" type="number" step="0.01" class="form-control" value="0">
        </div>
        <div class="col-12">
          <button class="btn btn-dark w-100">حفظ المنتج</button>
        </div>
      </form>
    </div>
  </div>
</div>
{% endblock %}
"""

# ----------------------------- ISSUES TEMPLATE --------------------------

ISSUES_HTML = r"""
{% extends 'base.html' %}
{% block content %}

<style>
  .issue-card {
    border-radius: 1rem;
    box-shadow: 0 2px 10px rgba(0,0,0,.06);
  }
  .issue-image {
    border-radius: .75rem;
    object-fit: cover;
  }
  .status-pill {
    font-size: .8rem;
    padding: .15rem .7rem;
    border-radius: 999px;
  }
</style>

<div class="row g-3">

  <!-- إضافة مشكلة جديدة -->
  <div class="col-xl-4">
    <div class="card p-3 h-100">
      <h6 class="mb-3">إضافة مشكلة جديدة</h6>
      <form method="post" action="{{ url_for('issues_add') }}" enctype="multipart/form-data" class="row g-2">
        <div class="col-12">
          <label class="form-label">عنوان المشكلة</label>
          <input required name="title" class="form-control" placeholder="مثال: مشكلة في طلب رقم 123" autocomplete="off">
        </div>
        <div class="col-12">
          <label class="form-label">وصف مختصر</label>
          <input name="desc" class="form-control" placeholder="شرح بسيط للمشكلة">
        </div>
        <div class="col-12">
          <label class="form-label">صورة (اختياري)</label>
          <input type="file" name="image" accept="image/*" class="form-control">
        </div>
        <div class="col-12 mt-2">
          <button class="btn btn-dark w-100">رفع المشكلة</button>
        </div>
      </form>
    </div>
  </div>

  <!-- عرض المشاكل -->
  <div class="col-xl-8">
    <div class="card p-3 mb-3">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <div>
          <h6 class="mb-0">المشاكل</h6>
          <small class="text-muted">عرض على شكل بوكسات ملونة حسب حالة المشكلة.</small>
        </div>
        <span class="badge bg-secondary">عدد السجلات: {{ rows|length }}</span>
      </div>

      <!-- ملخص سريع للحالات -->
      <div class="row g-2 small">
        <div class="col-6 col-md-4">
          <div class="border rounded-3 p-2 d-flex justify-content-between align-items-center bg-warning-subtle">
            <span class="small">قيد المعالجة</span>
            <span class="badge bg-warning text-dark">
              {{ rows|rejectattr('Status','equalto','Solved')|list|length }}
            </span>
          </div>
        </div>
        <div class="col-6 col-md-4">
          <div class="border rounded-3 p-2 d-flex justify-content-between align-items-center bg-success-subtle">
            <span class="small">تم الحل</span>
            <span class="badge bg-success">
              {{ rows|selectattr('Status','equalto','Solved')|list|length }}
            </span>
          </div>
        </div>
      </div>
    </div>

    <!-- البوكسات -->
    <div class="row g-3">
      {% for r in rows %}
        {% if r['Status'] == 'Solved' %}
          {% set card_border = 'border-success' %}
          {% set card_bg = 'bg-success-subtle' %}
          {% set pill_class = 'bg-success text-white' %}
        {% else %}
          {% set card_border = 'border-warning' %}
          {% set card_bg = 'bg-warning-subtle' %}
          {% set pill_class = 'bg-warning text-dark' %}
        {% endif %}

        <div class="col-md-6">
          <div class="issue-card card {{ card_border }} {{ card_bg }} h-100">
            <div class="card-body">
              <div class="d-flex justify-content-between align-items-start mb-2">
                <div>
                  <div class="small text-muted">#{{ r['ID'] }}</div>
                  <h6 class="mb-1">{{ r['Title'] }}</h6>
                  {% if r['Description'] %}
                  <p class="small text-muted mb-1">{{ r['Description'] }}</p>
                  {% endif %}
                </div>
                <span class="status-pill {{ pill_class }}">
                  {% if r['Status'] == 'Solved' %}
                    تم الحل
                  {% else %}
                    قيد المعالجة
                  {% endif %}
                </span>
              </div>

              <div class="row g-2 align-items-center">
                <div class="col-4">
                  {% if r['ImagePath'] %}
                    <img src="/static-proxy?f={{ r['ImagePath'] }}"
                         class="img-fluid issue-image"
                         style="height:80px;width:100%;">
                  {% else %}
                    <div class="border rounded-3 d-flex align-items-center justify-content-center text-muted"
                         style="height:80px;">
                      لا توجد صورة
                    </div>
                  {% endif %}
                </div>
                <div class="col-8 small">
                  <div class="d-flex justify-content-between">
                    <span class="text-muted">تاريخ الإنشاء:</span>
                    <span class="fw-semibold">{{ r['CreatedAt'] }}</span>
                  </div>

                  {% if r['Status'] == 'Solved' and r['Solver'] %}
                    <div class="mt-1 text-success fw-semibold">
                      تم حل المشكلة بواسطة: <span>{{ r['Solver'] }}</span>
                    </div>
                  {% endif %}
                </div>
              </div>
            </div>

            <div class="card-footer bg-transparent border-0 pt-0 pb-3 px-3">
              <div class="d-flex flex-wrap gap-1 align-items-center">
                {% if r['Status'] != 'Solved' %}
                  <form method="post" action="{{ url_for('issues_solve') }}" class="d-flex flex-wrap gap-1 align-items-center">
                    <input type="hidden" name="id" value="{{ r['ID'] }}">
                    <input name="solver"
                           class="form-control form-control-sm"
                           style="width:140px"
                           placeholder="تم الحل بواسطة..."
                           required>
                    <button class="btn btn-sm btn-success">تم الحل</button>
                  </form>
                {% else %}
                  <span class="small text-muted">
                    تم الحل بواسطة: <b>{{ r['Solver'] or 'غير مذكور' }}</b>
                  </span>
                {% endif %}

                <a class="btn btn-sm btn-outline-danger ms-auto"
                   href="{{ url_for('issues_delete', iid=r['ID']) }}"
                   onclick="return confirm('حذف المشكلة؟');">
                  حذف
                </a>
              </div>
            </div>
          </div>
        </div>
      {% endfor %}

      {% if not rows %}
        <div class="col-12">
          <div class="alert alert-light text-center border rounded-3">
            لا توجد مشاكل مسجّلة حالياً.
          </div>
        </div>
      {% endif %}
    </div>
  </div>
</div>

{% endblock %}
"""
SEAMSTRESS_HTML = r"""
{% extends 'base.html' %}
{% block content %}

<style>
  .seam-row-active    { background-color: #e8f5e9 !important; }  /* خياطة فعّالة - أخضر فاتح */
  .seam-row-inactive  { background-color: #f5f5f5 !important; }  /* غير فعّالة - رمادي فاتح */

  .card-title-small {
    font-size: 0.85rem;
    color: #6c757d;
    margin-bottom: 0.15rem;
  }
  .card-number-big {
    font-size: 1.2rem;
    font-weight: 700;
  }
</style>

{# حساب بعض الإحصائيات السريعة #}
{% set st = namespace(total_seams=0, active_seams=0, total_logs=0, unpaid_amount=0, total_pieces=0) %}
{% set st.total_seams = seamstresses|length %}
{% set st.total_logs = logs|length %}
{% for s in seamstresses %}
  {% if s['Active'] %}
    {% set st.active_seams = st.active_seams + 1 %}
  {% endif %}
{% endfor %}
{% for r in logs %}
  {% set st.total_pieces = st.total_pieces + (r['Pieces']|int) %}
  {% if not r['Paid'] %}
    {% set st.unpaid_amount = st.unpaid_amount + (r['Total']|float) %}
  {% endif %}
{% endfor %}

<div class="row g-3">

  <!-- كروت إحصائيات سريعة -->
  <div class="col-12">
    <div class="row g-3">
      <div class="col-md-3 col-6">
        <div class="card shadow-sm border-0 h-100">
          <div class="card-body py-2">
            <div class="card-title-small">عدد الخياطات الكلي</div>
            <div class="card-number-big">{{ st.total_seams }}</div>
          </div>
        </div>
      </div>
      <div class="col-md-3 col-6">
        <div class="card shadow-sm border-0 h-100">
          <div class="card-body py-2">
            <div class="card-title-small">الخياطات الفعّالات</div>
            <div class="card-number-big">{{ st.active_seams }}</div>
          </div>
        </div>
      </div>
      <div class="col-md-3 col-6">
        <div class="card shadow-sm border-0 h-100">
          <div class="card-body py-2">
            <div class="card-title-small">عدد سجلات الإنجاز</div>
            <div class="card-number-big">{{ st.total_logs }}</div>
          </div>
        </div>
      </div>
      <div class="col-md-3 col-6">
        <div class="card shadow-sm border-0 h-100">
          <div class="card-body py-2">
            <div class="card-title-small">إجمالي المبلغ غير المصفّى</div>
            <div class="card-number-big">{{ "{:,.0f}".format(((st.unpaid_amount|default(0))|float)) }}</div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- العمود الأيسر: الخياطات + الفلاتر + السجل -->
  <div class="col-xl-7">
    <!-- جدول الخياطات -->
    <div class="card p-3 shadow-sm mb-3">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <h6 class="mb-0">الخياطات</h6>
        <span class="badge bg-secondary">عدد السجلات: {{ seamstresses|length }}</span>
      </div>
      <div class="table-responsive">
        <table class="table table-striped align-middle mb-0">
          <thead>
            <tr>
              <th>#</th>
              <th>الاسم</th>
              <th>الهاتف</th>
              <th>ملاحظات</th>
              <th>الحالة</th>
              <th class="text-end">إجراءات</th>
            </tr>
          </thead>
          <tbody>
            {% for r in seamstresses %}
            <tr class="{{ 'seam-row-active' if r['Active'] else 'seam-row-inactive' }}">
              <td>{{ r['ID'] }}</td>
              <td>{{ r['Name'] }}</td>
              <td>{{ r['Phone'] }}</td>
              <td>{{ r['Notes'] }}</td>
              <td>
                {% if r['Active'] %}
                  <span class="badge bg-success">فعّالة</span>
                {% else %}
                  <span class="badge bg-secondary">موقوفة</span>
                {% endif %}
              </td>
              <td class="text-nowrap text-end">
                <button
                  class="btn btn-sm btn-outline-primary"
                  data-bs-toggle="modal"
                  data-bs-target="#editSeam"
                  data-id="{{r['ID']}}"
                  data-name="{{r['Name']}}"
                  data-phone="{{r['Phone']}}"
                  data-notes="{{r['Notes']}}"
                  data-active="{{r['Active']}}"
                >
                  تعديل
                </button>
                <a
                  class="btn btn-sm btn-outline-danger ms-1"
                  href="{{ url_for('seam_delete', sid=r['ID']) }}"
                  onclick="return confirm('حذف {{r['Name']}}؟');"
                >
                  حذف
                </a>
              </td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>

    <!-- فلاتر السجل -->
    <div class="card p-3 shadow-sm mb-3">
      <h6 class="mb-2">تصفية سجل الإنجاز</h6>
      <form method="get" class="row g-2">
        <div class="col-md-3">
          <label class="form-label">من تاريخ</label>
          <input type="date" name="from" class="form-control" value="{{ dfrom or '' }}">
        </div>
        <div class="col-md-3">
          <label class="form-label">إلى تاريخ</label>
          <input type="date" name="to" class="form-control" value="{{ dto or '' }}">
        </div>
        <div class="col-md-3">
          <label class="form-label">الخياطة</label>
          <select name="sid" class="form-select">
            <option value="">الكل</option>
            {% for r in seamstresses %}
              <option value="{{ r['ID'] }}" {{ 'selected' if sel_sid and sel_sid|int == r['ID'] else '' }}>
                {{ r['Name'] }}
              </option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-3">
          <label class="form-label">الحالة</label>
          <select name="paid" class="form-select">
            <option value="">الكل</option>
            <option value="paid"   {{ 'selected' if sel_paid=='paid' else '' }}>مدفوع</option>
            <option value="unpaid" {{ 'selected' if sel_paid=='unpaid' else '' }}>غير مدفوع</option>
          </select>
        </div>
        <div class="col-12 text-end mt-1">
          <button class="btn btn-secondary btn-sm">تطبيق</button>
          <a href="{{ url_for('home') }}" class="btn btn-outline-secondary btn-sm">إلغاء</a>
        </div>
      </form>
    </div>

    <!-- جدول سجل الإنجاز اليومي (ملخص لكل خياطة + زر تفاصيل) -->
    <div class="card p-3 shadow-sm">
      {% set grouped = logs|groupby('SeamstressID')|list %}
      <div class="d-flex justify-content-between align-items-center mb-2">
        <h6 class="mb-0">سجل الإنجاز اليومي</h6>
        <div class="d-flex align-items-center gap-2">
          <span class="badge bg-secondary">
            عدد الخياطات في السجل: {{ grouped|length }}
          </span>
          <span class="badge bg-light text-muted">
            إجمالي القطع: {{ st.total_pieces }}
          </span>
        </div>
      </div>

      <div class="table-responsive">
        <table class="table align-middle mb-0">
          <thead>
            <tr>
              <th>#</th>
              <th>اسم الخياطة</th>
              <th>إجمالي القطع</th>
              <th>إجمالي المبلغ</th>
              <th class="text-end">إجراءات</th>
            </tr>
          </thead>
          <tbody>
            {% for g in grouped %}
              {% set sid = g.grouper %}
              {% set name = seam_name_map.get(sid, sid) %}
              {% set ns = namespace(total_pieces=0, total_amount=0) %}
              {% for r in g.list %}
                {% set ns.total_pieces = ns.total_pieces + (r['Pieces']|int) %}
                {% set ns.total_amount = ns.total_amount + (r['Total']|float) %}
              {% endfor %}

              <!-- صف الملخص -->
              <tr>
                <td>{{ sid }}</td>
                <td>{{ name }}</td>
                <td>{{ ns.total_pieces }}</td>
                <td>{{ "{:,.0f}".format(((ns.total_amount|default(0))|float)) }}</td>
                <td class="text-end">
                  <button class="btn btn-sm btn-outline-primary"
                          type="button"
                          data-bs-toggle="collapse"
                          data-bs-target="#logs-{{ sid }}"
                          aria-expanded="false"
                          aria-controls="logs-{{ sid }}">
                    عرض الإنجازات
                  </button>
                </td>
              </tr>

              <!-- صف التفاصيل (يظهر عند الضغط على الزر) -->
              <tr class="collapse-row">
                <td colspan="5" class="p-0 border-0">
                  <div class="collapse" id="logs-{{ sid }}">
                    <div class="p-2 border-top">
                      <div class="small text-muted mb-1">
                        إنجازات الخياطة: <b>{{ name }}</b>
                      </div>
                      <div class="table-responsive">
                        <table class="table table-sm align-middle mb-0">
                          <thead class="table-light">
                            <tr>
                              <th>#</th>
                              <th>التاريخ</th>
                              <th>الموديل</th>
                              <th>القطع</th>
                              <th>سعر/قطعة</th>
                              <th>الإجمالي</th>
                              <th>الحالة</th>
                              <th class="text-end">إجراءات</th>
                            </tr>
                          </thead>
                          <tbody>
                            {% for r in g.list %}
                            <tr class="{% if r['Paid'] %}table-success{% else %}table-warning{% endif %}">
                              <td>{{ r['LogID'] }}</td>
                              <td>{{ r['Date'] }}</td>
                              <td>{{ r['Model'] }}</td>
                              <td>{{ r['Pieces'] }}</td>
                              <td>{{ r['UnitCost'] }}</td>
                              <td>{{ r['Total'] }}</td>
                              <td>
                                {% if r['Paid'] %}
                                  <span class="badge bg-success">مدفوع</span>
                                {% else %}
                                  <span class="badge bg-warning text-dark">غير مدفوع</span>
                                {% endif %}
                              </td>
                              <td class="text-end">
                                {% if not r['Paid'] %}
                                  <a class="btn btn-sm btn-success"
                                     href="{{ url_for('sew_mark_paid', log_id=r['LogID']) }}">
                                    تصفية
                                  </a>
                                {% else %}
                                  <a class="btn btn-sm btn-outline-secondary"
                                     href="{{ url_for('sew_mark_unpaid', log_id=r['LogID']) }}">
                                    إلغاء التصفية
                                  </a>
                                {% endif %}
                              </td>
                            </tr>
                            {% endfor %}
                          </tbody>
                        </table>
                      </div>
                    </div>
                  </div>
                </td>
              </tr>
            {% endfor %}

            {% if not grouped %}
              <tr>
                <td colspan="5" class="text-center text-muted py-3">
                  لا توجد إنجازات مسجّلة حاليًا.
                </td>
              </tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- العمود الأيمن: إضافة خياطة + تسجيل إنجاز -->
  <div class="col-xl-5">
    <!-- إضافة خياطة جديدة -->
    <div class="card p-3 shadow-sm mb-3">
      <h6 class="mb-3">إضافة خياطة</h6>
      <form method="post" action="{{ url_for('seam_add') }}" class="row g-2">
        <div class="col-6">
          <label class="form-label">الاسم</label>
          <input required name="name" class="form-control" autocomplete="off">
        </div>
        <div class="col-6">
          <label class="form-label">الهاتف</label>
          <input name="phone" class="form-control" inputmode="numeric" pattern="[0-9]*">
        </div>
        <div class="col-12">
          <label class="form-label">ملاحظات</label>
          <input name="notes" class="form-control">
        </div>
        <div class="col-12">
          <button class="btn btn-dark w-100">إضافة</button>
        </div>
      </form>
    </div>

    <!-- تسجيل إنجاز اليوم -->
    <div class="card p-3 shadow-sm">
      <h6 class="mb-3">تسجيل إنجاز اليوم</h6>
      <form method="post" action="{{ url_for('sew_add_log') }}" class="row g-2">
        <div class="col-6">
          <label class="form-label">الخياطة</label>
          <select name="sid" class="form-select" required>
            <option value="">—</option>
            {% for r in seamstresses %}
              <option value="{{ r['ID'] }}">{{ r['Name'] }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-6">
          <label class="form-label">المنتج / الموديل</label>
          <select name="model" class="form-select" required>
            <option value="">— اختر المنتج —</option>
            {% for p in products %}
              <option value="{{ p['Product Name'] }}">
                {{ p['Product Name'] }}
                {% if p['Type'] %} — {{ p['Type'] }}{% endif %}
                {% if p['Product Code'] %} ({{ p['Product Code'] }}){% endif %}
              </option>
            {% endfor %}
          </select>
        </div>
        <div class="col-6">
          <label class="form-label">عدد القطع</label>
          <input required type="number" name="pieces" class="form-control" min="1" value="1" inputmode="numeric" pattern="[0-9]*">
        </div>
        <div class="col-6">
          <label class="form-label">سعر الخياطة/قطعة</label>
          <input required type="number" step="0.01" name="unit" class="form-control" value="0" inputmode="decimal">
        </div>
        <div class="col-12">
          <button class="btn btn-success w-100">تسجيل & زيادة المخزون</button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Modal تعديل خياطة -->
<div class="modal fade" id="editSeam" tabindex="-1">
  <div class="modal-dialog">
    <form method="post" action="{{ url_for('seam_edit') }}" class="modal-content">
      <div class="modal-header">
        <h6 class="modal-title">تعديل خياطة</h6>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body">
        <input type="hidden" name="id" id="seamID">
        <div class="mb-2">
          <label class="form-label">الاسم</label>
          <input name="name" id="seamName" class="form-control">
        </div>
        <div class="mb-2">
          <label class="form-label">الهاتف</label>
          <input name="phone" id="seamPhone" class="form-control" inputmode="numeric" pattern="[0-9]*">
        </div>
        <div class="mb-2">
          <label class="form-label">ملاحظات</label>
          <input name="notes" id="seamNotes" class="form-control">
        </div>
        <div class="form-check">
          <input class="form-check-input" type="checkbox" name="active" id="seamActive">
          <label class="form-check-label" for="seamActive">فعّالة</label>
        </div>
      </div>
      <div class="modal-footer">
        <button class="btn btn-primary">حفظ</button>
      </div>
    </form>
  </div>
</div>

<script>
  const editSeam = document.getElementById('editSeam');
  editSeam?.addEventListener('show.bs.modal', e => {
    const b = e.relatedTarget;
    document.getElementById('seamID').value    = b.getAttribute('data-id');
    document.getElementById('seamName').value  = b.getAttribute('data-name');
    document.getElementById('seamPhone').value = b.getAttribute('data-phone');
    document.getElementById('seamNotes').value = b.getAttribute('data-notes');
    document.getElementById('seamActive').checked = (b.getAttribute('data-active') === 'True');
  });
</script>

{% endblock %}
"""


# ---------------------------- CUTTING TEMPLATE --------------------------
CUTTING_HTML = r"""
{% extends 'base.html' %}
{% block content %}

<style>
  .cutting-card {
    border-radius: 1rem;
    box-shadow: 0 2px 10px rgba(0,0,0,.06);
  }
  .cutting-card img {
    border-radius: .75rem;
    object-fit: cover;
  }
  .status-pill {
    font-size: .8rem;
    padding: .15rem .55rem;
    border-radius: 999px;
  }
</style>

<div class="row g-3">

  <!-- عمود إنشاء فصل جديد -->
  <div class="col-xl-4">
    <div class="card p-3 h-100">
      <h6 class="mb-3">إنشاء فصل جديد</h6>
      <form method="post" action="{{ url_for('cutting_add') }}" enctype="multipart/form-data" class="row g-2">
        <div class="col-12">
          <label class="form-label">اسم الموديل</label>
          <input required name="model" class="form-control" autocomplete="off">
        </div>
        <div class="col-6">
          <label class="form-label">موعد الفصال</label>
          <input required type="date" name="due" class="form-control">
        </div>
        <div class="col-6">
          <label class="form-label">عدد القطع المطلوبة</label>
          <input required type="number" name="qty" class="form-control" min="1" value="1" inputmode="numeric" pattern="[0-9]*">
        </div>
        <div class="col-12">
          <label class="form-label">ملاحظات</label>
          <input name="notes" class="form-control">
        </div>
        <div class="col-12">
          <label class="form-label">صورة الموديل</label>
          <input type="file" name="image" accept="image/*" class="form-control">
        </div>
        <div class="col-12 mt-2">
          <button class="btn btn-dark w-100">إنشاء</button>
        </div>
      </form>
    </div>
  </div>

  <!-- عمود طلبات الفصال (بوكسات ملونة) -->
  <div class="col-xl-8">
    <div class="card p-3 mb-3">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <div>
          <h6 class="mb-0">طلبات الفصال</h6>
          <small class="text-muted">عرض على شكل بوكسات ملونة حسب الحالة.</small>
        </div>
        <span class="badge bg-secondary">عدد السجلات: {{ rows|length }}</span>
      </div>

      <!-- ملخص سريع للحالات -->
      <div class="row g-2 small">
        <div class="col-6 col-md-3">
          <div class="border rounded-3 p-2 d-flex justify-content-between align-items-center bg-secondary-subtle">
            <span class="small">قيد الانتظار</span>
            <span class="badge bg-secondary">
              {{ rows|selectattr('Status','equalto','قيد الانتظار')|list|length }}
            </span>
          </div>
        </div>
        <div class="col-6 col-md-3">
          <div class="border rounded-3 p-2 d-flex justify-content-between align-items-center bg-warning-subtle">
            <span class="small">قيد العمل</span>
            <span class="badge bg-warning text-dark">
              {{ rows|selectattr('Status','equalto','قيد العمل')|list|length }}
            </span>
          </div>
        </div>
        <div class="col-6 col-md-3 mt-2 mt-md-0">
          <div class="border rounded-3 p-2 d-flex justify-content-between align-items-center bg-success-subtle">
            <span class="small">مكتمل</span>
            <span class="badge bg-success">
              {{ rows|selectattr('Status','equalto','مكتمل')|list|length }}
            </span>
          </div>
        </div>
        <div class="col-6 col-md-3 mt-2 mt-md-0">
          <div class="border rounded-3 p-2 d-flex justify-content-between align-items-center bg-danger-subtle">
            <span class="small">مرفوض</span>
            <span class="badge bg-danger">
              {{ rows|selectattr('Status','equalto','مرفوض')|list|length }}
            </span>
          </div>
        </div>
      </div>
    </div>

    <!-- البوكسات نفسها -->
    <div class="row g-3">
      {% for r in rows %}
        {% set st = r['Status'] %}
        {% if st == 'قيد العمل' %}
          {% set card_border = 'border-warning' %}
          {% set card_bg = 'bg-warning-subtle' %}
          {% set pill_class = 'bg-warning text-dark' %}
        {% elif st == 'مكتمل' %}
          {% set card_border = 'border-success' %}
          {% set card_bg = 'bg-success-subtle' %}
          {% set pill_class = 'bg-success text-white' %}
        {% elif st == 'مرفوض' %}
          {% set card_border = 'border-danger' %}
          {% set card_bg = 'bg-danger-subtle' %}
          {% set pill_class = 'bg-danger text-white' %}
        {% elif st == 'قيد الانتظار' %}
          {% set card_border = 'border-secondary' %}
          {% set card_bg = 'bg-secondary-subtle' %}
          {% set pill_class = 'bg-secondary text-white' %}
        {% else %}
          {% set card_border = 'border-light' %}
          {% set card_bg = 'bg-light' %}
          {% set pill_class = 'bg-light text-dark' %}
        {% endif %}

        <div class="col-md-6">
          <div class="cutting-card card {{ card_border }} {{ card_bg }} h-100">
            <div class="card-body">
              <div class="d-flex justify-content-between align-items-start mb-2">
                <div>
                  <div class="small text-muted">#{{ r['ID'] }}</div>
                  <h6 class="mb-0">{{ r['Model'] }}</h6>
                </div>
                <span class="status-pill {{ pill_class }}">
                  {{ st }}
                </span>
              </div>

              <div class="row g-2 align-items-center">
                <div class="col-4">
                  {% if r['ImagePath'] %}
                    <img src="/static-proxy?f={{ r['ImagePath'] }}" class="img-fluid" style="height:90px;width:100%;">
                  {% else %}
                    <div class="border rounded-3 d-flex align-items-center justify-content-center text-muted"
                         style="height:90px;">
                      لا توجد صورة
                    </div>
                  {% endif %}
                </div>
                <div class="col-8">
                  <div class="small">
                    <div class="d-flex justify-content-between">
                      <span class="text-muted">الموعد:</span>
                      <span class="fw-semibold">{{ r['DueDate'] }}</span>
                    </div>
                    <div class="d-flex justify-content-between">
                      <span class="text-muted">المطلوب:</span>
                      <span class="fw-semibold">{{ r['RequiredQty'] }} قطعة</span>
                    </div>
                    {% if r['Notes'] %}
                    <div class="mt-1">
                      <span class="text-muted">ملاحظات:</span>
                      <span>{{ r['Notes'] }}</span>
                    </div>
                    {% endif %}
                    {% if r['RejectionReason'] %}
                    <div class="mt-1 text-danger">
                      <span class="text-muted">سبب الرفض:</span>
                      <span>{{ r['RejectionReason'] }}</span>
                    </div>
                    {% endif %}
                  </div>
                </div>
              </div>
            </div>

            <div class="card-footer bg-transparent border-0 pt-0 pb-3 px-3">
              <div class="d-flex flex-wrap gap-1">
                <a class="btn btn-sm btn-outline-secondary"
                   href="{{ url_for('cutting_status', cid=r['ID'], s='قيد الانتظار') }}">
                   انتظار
                </a>
                <a class="btn btn-sm btn-primary"
                   href="{{ url_for('cutting_status', cid=r['ID'], s='قيد العمل') }}">
                   عمل
                </a>
                <a class="btn btn-sm btn-success"
                   href="{{ url_for('cutting_status', cid=r['ID'], s='مكتمل') }}">
                   مكتمل
                </a>
                <button class="btn btn-sm btn-outline-danger"
                        data-bs-toggle="modal"
                        data-bs-target="#rejectModal"
                        data-id="{{ r['ID'] }}">
                  رفض
                </button>
                <a class="btn btn-sm btn-outline-danger ms-auto"
                   href="{{ url_for('cutting_delete', cid=r['ID']) }}"
                   onclick="return confirm('حذف الفصال؟');">
                  حذف
                </a>
              </div>
            </div>
          </div>
        </div>
      {% endfor %}

      {% if not rows %}
        <div class="col-12">
          <div class="alert alert-light text-center border rounded-3">
            لا توجد طلبات فصال حالياً.
          </div>
        </div>
      {% endif %}
    </div>
  </div>
</div>

<!-- مودال الرفض -->
<div class="modal fade" id="rejectModal" tabindex="-1">
  <div class="modal-dialog">
    <form method="post" action="{{ url_for('cutting_reject') }}" class="modal-content">
      <div class="modal-header">
        <h6 class="modal-title">رفض طلب فصال</h6>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body">
        <input type="hidden" name="id" id="rejID">
        <label class="form-label">سبب الرفض</label>
        <textarea required name="reason" class="form-control" rows="3"></textarea>
      </div>
      <div class="modal-footer">
        <button class="btn btn-danger">رفض</button>
      </div>
    </form>
  </div>
</div>

<script>
  const rej = document.getElementById('rejectModal');
  rej?.addEventListener('show.bs.modal', e => {
    document.getElementById('rejID').value = e.relatedTarget.getAttribute('data-id');
  });
</script>

{% endblock %}
"""


BASE_HTML = r"""
<!doctype html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{{ title or 'نظام إدارة الطلبات (ويب)' }}</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    body{background:#f8f9fb}
    .table thead th{white-space:nowrap}
    .card{box-shadow:0 2px 10px rgba(0,0,0,.06)}
    .form-control, .btn{border-radius:0.75rem}
  </style>
</head>
<body>
<nav class="navbar navbar-expand-lg bg-white border-bottom">
  <div class="container-fluid">
    <a class="navbar-brand fw-bold" href="{{ url_for('home') }}">🗂️ نظام الطلبات</a>
    <div class="d-flex">
      {% if session.get('auth') %}
      <a class="btn btn-sm btn-outline-secondary me-2" href="{{ url_for('download_excel') }}">تنزيل ملف Excel</a>
      <a class="btn btn-sm btn-outline-dark me-2" href="{{ url_for('system_export') }}">تنزيل تقرير شامل</a>
      <a class="btn btn-sm btn-outline-primary me-2" href="{{ url_for('report_orders_status', status_key='shipping') }}">تقرير قيد التوصيل</a>
      <a class="btn btn-sm btn-outline-primary me-2" href="{{ url_for('report_orders_status', status_key='ready') }}">تقرير قيد التجهيز</a>
      <a class="btn btn-sm btn-outline-success me-2" href="{{ url_for('report_inventory_withdrawn') }}">تقرير المسحوب من المخزن</a>
      <a class="btn btn-sm btn-danger" href="{{ url_for('logout') }}">تسجيل خروج</a>
      {% endif %}
    </div>
  </div>
</nav>

<div class="container py-4">
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
      {% for cat, msg in messages %}
        <div class="alert alert-{{ 'success' if cat=='ok' else ('danger' if cat=='err' else 'info') }}">{{ msg }}</div>
      {% endfor %}
    {% endif %}
  {% endwith %}
  {% block content %}{% endblock %}
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""

LOGIN_HTML = r"""
{% extends 'base.html' %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-md-5">
    <div class="card p-4">
      <h5 class="mb-3">أدخل رمز الدخول</h5>
      <form method="post">
        <div class="mb-3">
          <input required name="code" type="password" class="form-control form-control-lg" placeholder="••••">
        </div>
        <button class="btn btn-primary w-100">دخول</button>
      </form>
    </div>
  </div>
</div>
{% endblock %}
"""

HOME_HTML = r"""

{% extends 'base.html' %}
{% block content %}

<div class="d-flex flex-wrap gap-2 mb-3">
  <a class="btn btn-outline-warning" href="{{ url_for('orders_processing') }}">📌 الطلبات قيد المعالجة</a>
  <a class="btn btn-outline-success" href="{{ url_for('orders_import_text') }}">📝 إدخال طلبات من نص</a>
</div>
<div class="row g-3 mb-3">
  <!-- كروت الإحصاءات السريعة -->
  <div class="col-md-3">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="small text-muted">العدد الكلي للطلبات</div>
            <div class="fs-4 fw-bold mt-1">{{ summary["العدد الكلي للطلبات"] or 0 }}</div>
            <div class="small text-muted mt-1">كل الطلبات حسب الفلاتر الحالية.</div>
          </div>
          <span class="fs-2">📦</span>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-3">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="small text-muted">الطلبات {{ status_delivered }}</div>
            <div class="fs-4 fw-bold mt-1">
              {{ summary["عدد " ~ status_delivered] }}
            </div>
            <div class="small text-muted mt-1">
              نسبة من الكل: {{ summary["نسبة " ~ status_delivered ~ " %"] }}%
            </div>
          </div>
          <span class="fs-2">✅</span>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-3">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="small text-muted">الطلبات {{ status_returned }}</div>
            <div class="fs-4 fw-bold mt-1">
              {{ summary["عدد " ~ status_returned] }}
            </div>
            <div class="small text-muted mt-1">
              نسبة من الكل: {{ summary["نسبة " ~ status_returned ~ " %"] }}%
            </div>
          </div>
          <span class="fs-2">↩️</span>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-3">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="small text-muted">إجمالي قيمة الطلبات</div>
            <div class="fs-5 fw-bold mt-1">
              {{ "%.0f"|format(summary["المجموع المالي (Order Price)"] or 0) }} د.ع
            </div>
            <div class="small text-muted mt-1">إجمالي المبالغ حسب الفلاتر الحالية.</div>
          </div>
          <span class="fs-2">💰</span>
        </div>
      </div>
    </div>
  </div>
</div>

<div class="row g-3">
  <div class="col-xl-8">
    <div class="card p-3">
      <form class="row g-2 align-items-end" method="get" action="{{ url_for('home') }}">
        <div class="col-md-3">
          <label class="form-label">بحث</label>
          <input name="q" value="{{ q or '' }}" class="form-control" placeholder="كلمة مفتاحية" autofocus>
        </div>
        <div class="col-md-3">
          <label class="form-label">اسم المنتج</label>
          <select name="product" class="form-select">
            <option value="">الكل</option>
            {% for p in all_products %}
              <option value="{{ p }}" {{ 'selected' if sel_product==p else '' }}>{{ p }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-3">
          <label class="form-label">اسم البيج</label>
          <select name="page" class="form-select">
            <option value="">الكل</option>
            {% for p in all_pages %}
              <option value="{{ p }}" {{ 'selected' if sel_page==p else '' }}>{{ p }}</option>
            {% endfor %}
          </select>
        </div>

        <div class="col-md-3">
          <label class="form-label">الحالة</label>
          <select name="status" class="form-select">
            <option value="">الكل</option>
            {% for s in all_statuses %}
              <option value="{{ s }}" {{ 'selected' if sel_status==s else '' }}>{{ s }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-3">
          <label class="form-label">من تاريخ</label>
          <input name="from" type="date" class="form-control" value="{{ dfrom or '' }}">
        </div>
        <div class="col-md-3">
          <label class="form-label">إلى تاريخ</label>
          <input name="to" type="date" class="form-control" value="{{ dto or '' }}">
        </div>
        <div class="col-md-3 text-end align-self-end">
          <button class="btn btn-secondary mt-2">تطبيق</button>
          <a href="{{ url_for('home') }}" class="btn btn-outline-secondary mt-2">إلغاء</a>
        </div>
      </form>
    </div>

   <div class="card p-3 mt-3">
  <div class="d-flex justify-content-between align-items-center mb-2">
    <div class="d-flex align-items-center gap-2">
      <h6 class="mb-0">الطلبات</h6>
      <button class="btn btn-sm btn-outline-secondary"
              type="button"
              data-bs-toggle="collapse"
              data-bs-target="#ordersCollapse"
              aria-expanded="false"
              aria-controls="ordersCollapse">
        فتح / إخفاء
      </button>
    </div>
    <span class="badge bg-secondary">عدد السجلات: {{ rows|length }}</span>
  </div>

  <div class="collapse" id="ordersCollapse">
    <div class="table-responsive">
    <table class="table table-hover table-striped align-middle mb-0">
      <thead class="table-light">
        <tr>
          <th>رقم الشحنة</th>
          <th>التاريخ</th>
          <th>اسم المنتج</th>
          <th>اسم البيج</th>
          <th>السعر</th>
          <th>الحالة</th>
          <th>ملاحظات</th>
          <th class="text-center">إجراءات</th>
        </tr>
      </thead>
      <tbody>
        {% for r in rows %}
        {% set st = (r.get('Status') or '') %}
        <tr class="
          {% if   st == status_delivered %}table-success
          {% elif st == status_returned %}table-danger
          {% elif st == 'قيد التوصيل' %}table-warning
          {% elif st == 'قيد التجهيز' %}table-info
          {% endif %}
        ">
          <td class="fw-bold">{{ r.get('Transaction ID','') }}</td>
          <td class="text-nowrap">{{ r.get('Time and Date','') }}</td>
          <td>{{ r.get('Product Name','') }}</td>
          <td>{{ r.get('Page Name','') }}</td>
          <td>{{ r.get('Order Price','') }}</td>

          <td>
            {% if st == status_delivered %}
              <span class="badge bg-success">تم التوصيل</span>
            {% elif st == status_returned %}
              <span class="badge bg-danger">راجع</span>
            {% elif st == 'قيد التوصيل' %}
              <span class="badge bg-warning text-dark">قيد التوصيل</span>
            {% elif st == 'قيد التجهيز' %}
              <span class="badge bg-info text-dark">قيد التجهيز</span>
            {% else %}
              <span class="badge bg-secondary">{{ st or 'غير محدد' }}</span>
            {% endif %}
          </td>

          <td class="small text-muted">{{ r.get('Notes','') }}</td>

          <td class="text-center text-nowrap">
            <a href="{{ url_for('edit', txn=r.get('Transaction ID')) }}"
               class="btn btn-sm btn-outline-primary">
              تعديل
            </a>
            <a href="{{ url_for('delete', txn=r.get('Transaction ID')) }}"
               class="btn btn-sm btn-outline-danger ms-1"
               onclick="return confirm('هل أنت متأكد من حذف الطلب رقم {{ r.get('Transaction ID') }} ؟');">
              حذف
            </a>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
    </div>
  </div>
</div>

  </div>

  <div class="col-xl-4">
    <div class="card p-3">
      <h6 class="mb-3">قارئ باركود (تحديث إلى راجع)</h6>
      <form method="post" action="{{ url_for('mark_returned') }}" class="row g-2">
        <div class="col-8">
          <input required id="barcodeInput" name="txn" class="form-control" placeholder="Transaction ID" autofocus>
        </div>
        <div class="col-4">
          <button class="btn btn-warning w-100">تحديث</button>
        </div>
      </form>
    </div>

    <div class="card p-3 mt-3">
      <h6 class="mb-3">استيراد من PDF</h6>
      <form method="post" action="{{ url_for('upload_pdf') }}" enctype="multipart/form-data" class="row g-2">
        <div class="col-12">
          <input required class="form-control" type="file" name="pdf" accept="application/pdf">
        </div>
        <div class="col-12">
          <button class="btn btn-primary w-100">إضافة ملف PDF</button>
        </div>
      </form>
      <hr>
      <h6 class="mb-3">فاتورة مطابقة (تسليم تلقائي)</h6>
      <form method="post" action="{{ url_for('upload_invoice') }}" enctype="multipart/form-data" class="row g-2">
        <div class="col-12">
          <input required class="form-control" type="file" name="pdf" accept="application/pdf">
        </div>
        <div class="col-12">
          <button class="btn btn-success w-100">رفع فاتورة</button>
        </div>
      </form>
    </div>

    <div class="card p-3 mt-3">
      <h6 class="mb-3">أدوات سريعة ⚙️</h6>
     <div class="d-grid gap-2">
    <!-- الطلبات -->
    <a class="btn btn-outline-secondary" href="{{ url_for('dedupe') }}">حذف مكرر</a>
    
    <a class="btn btn-outline-danger" href="{{ url_for('delete_ready_all') }}" onclick="return confirm('⚠️ هل أنت متأكد من حذف جميع الطلبات قيد التجهيز؟ لا يمكن التراجع!');">حذف كل قيد التجهيز</a>
    <a class="btn btn-outline-secondary" href="{{ url_for('move_to_shipping') }}">تحديث إلى قيد التوصيل</a>
    <a class="btn btn-outline-secondary" href="{{ url_for('returns_bulk') }}">إدارة راجع</a>
    <a class="btn btn-outline-secondary" href="{{ url_for('delivered_bulk') }}">إدارة تم التوصيل</a>
    <a class="btn btn-outline-secondary" href="{{ url_for('pending') }}">الطلبات قيد التوصيل</a>
    <a class="btn btn-outline-primary" href="{{ url_for('stats') }}">الإحصائيات (مفصّل) 🔒</a>

    <!-- المخزن والمنتجات -->
    <a class="btn btn-outline-dark" href="{{ url_for('inventory_home') }}">المخزن</a>
    <a class="btn btn-outline-dark" href="{{ url_for('products_home') }}">إدارة المنتجات</a>

    <!-- الخياطات والفصال والمشاكل -->
    <a class="btn btn-outline-dark" href="{{ url_for('home') }}">الخياطات</a>
    <a class="btn btn-outline-dark" href="{{ url_for('cutting_home') }}">الفصال</a>
    <a class="btn btn-outline-dark" href="{{ url_for('issues_home') }}">المشاكل</a>
  </div>
    </div>
  </div>
</div>

<script>
  // إعادة تركيز المؤشر على خانة الباركود بعد كل تحميل
  const inp = document.getElementById('barcodeInput');
  if (inp) {
    setTimeout(() => { inp.focus(); inp.select(); }, 100);
  }
</script>
{% endblock %}
"""




PROCESSING_ORDERS_HTML = r"""
{% extends 'base.html' %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <div>
    <h5 class="mb-0">الطلبات قيد المعالجة</h5>
    <div class="text-muted small">عدد الطلبات: <b>{{ total }}</b></div>
  </div>
  <div class="d-flex gap-2">
    <a class="btn btn-outline-success btn-sm" href="{{ url_for('orders_import_text') }}">📝 إدخال من نص</a>
    <form method="post" action="{{ url_for('processing_delete_all') }}" onsubmit="return confirm('حذف كل طلبات قيد المعالجة؟')">
      <button class="btn btn-outline-danger btn-sm">🗑️ حذف الكل</button>
    </form>
  </div>
</div>

<form method="get" class="row g-2 mb-3">
  <div class="col-md-5">
    <input name="q" value="{{ q or '' }}" class="form-control" id="q_input" list="q_suggestions"
           placeholder="بحث سريع: رقم الهاتف / العنوان / نص الطلب / اسم المنتج">
    <datalist id="q_suggestions"></datalist>
  </div>
  <div class="col-md-4">
    <select name="product" class="form-select">
      <option value="">— كل المنتجات —</option>
      {% for name, count in product_counts %}
        <option value="{{ name }}" {{ 'selected' if product_filter==name else '' }}>{{ name }} ({{ count }})</option>
      {% endfor %}
    </select>
  </div>
  <div class="col-md-3 d-flex gap-2">
    <button class="btn btn-primary w-100">بحث</button>
    <a class="btn btn-outline-secondary w-100" href="{{ url_for('orders_processing') }}">تفريغ</a>
  </div>
</form>

<div class="table-responsive">
<table class="table table-sm table-bordered align-middle">
  <thead class="table-light">
    <tr>
      <th>التاريخ</th>
      <th>المنتج</th>
      <th>الهاتف</th>
      <th>العنوان</th>
      <th>السعر</th>
      <th style="width:220px">إجراءات</th>
    </tr>
  </thead>
  <tbody>
  {% if rows %}
    {% for r in rows %}
      <tr>
        <td class="small">{{ r.get('Time and Date','') }}</td>
        <td>{{ r.get('Product Name','') }}</td>
        <td>{{ r.get('Contact Numbers','') }}</td>
        <td class="small">{{ r.get('Address','') }}</td>
        <td>{{ r.get('Order Price','') }}</td>
        <td>
          <div class="d-flex flex-wrap gap-1">
            <form method="post" action="{{ url_for('processing_to_shipping') }}">
              <input type="hidden" name="txn" value="{{ r.get('Transaction ID','') }}">
              <button class="btn btn-success btn-sm" type="submit">قيد التوصيل</button>
            </form>

            <a class="btn btn-outline-primary btn-sm" href="{{ url_for('processing_edit', txn=r.get('Transaction ID','')) }}">تحديث</a>

            {% if r.get('Status','') != STATUS_SHIPPING %}
            <form method="post" action="{{ url_for('processing_delete') }}" onsubmit="return confirm('حذف هذا الطلب؟')">
              <input type="hidden" name="txn" value="{{ r.get('Transaction ID','') }}">
              <button class="btn btn-outline-danger btn-sm" type="submit">حذف</button>
            </form>
            {% endif %}
          </div>
        </td>
      </tr>
    {% endfor %}
  {% else %}
    <tr><td colspan="6" class="text-center text-muted">لا توجد طلبات</td></tr>
  {% endif %}
  </tbody>
</table>
</div>

<script>
(function(){
  const inp = document.getElementById('q_input');
  const dl = document.getElementById('q_suggestions');
  if(!inp || !dl) return;
  let t=null;
  inp.addEventListener('input', function(){
    const q = inp.value.trim();
    if(t) clearTimeout(t);
    if(!q){ dl.innerHTML=''; return; }
    t=setTimeout(async ()=>{
      try{
        const res = await fetch(`/orders/processing_suggest?q=${encodeURIComponent(q)}`);
        const data = await res.json();
        dl.innerHTML = (data.items||[]).map(x=>`<option value="${x.replace(/"/g,'&quot;')}"></option>`).join('');
      }catch(e){}
    }, 150);
  });
})();
</script>
{% endblock %}
"""

IMPORT_TEXT_HTML = r"""
{% extends 'base.html' %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <div>
    <h5 class="mb-0">إدخال طلبات من نص</h5>
    <div class="text-muted small">كل طلب 4 أسطر: (1) المنتج (2) رقم أو عنوان (3) عنوان أو رقم (4) السعر. افصل بين الطلبات بسطر فارغ أو ---</div>
  </div>
  <div class="d-flex gap-2">
    <a class="btn btn-outline-secondary btn-sm" href="{{ url_for('orders_processing') }}">قيد المعالجة</a>
    <a class="btn btn-outline-secondary btn-sm" href="{{ url_for('home') }}">الرئيسية</a>
  </div>
</div>

<form method="post" class="card border-0 shadow-sm rounded-4">
  <div class="card-body">
    <label class="form-label">الصق النص هنا</label>
    <textarea name="raw" class="form-control" rows="14" placeholder="مثال:
خفاش
07741002423
بغداد / الجديدة - شارع المسبح
25000

ستن عدد 2
07891203683
صلاح الدين بلد
25000
">{{ raw or "" }}</textarea>

    {% if preview %}
      <div class="mt-3">
        <div class="fw-bold mb-2">معاينة ({{ preview|length }})</div>
        <ul class="small">
          {% for o in preview %}
            <li><b>{{ o.product }}</b> — {{ o.phone }} — {{ o.price }} — {{ o.address }}</li>
          {% endfor %}
        </ul>
      </div>
    {% endif %}

    <div class="d-flex gap-2 mt-3">
      <button name="action" value="preview" class="btn btn-outline-primary">معاينة</button>
      <button name="action" value="save" class="btn btn-primary">حفظ</button>
    </div>
  </div>
</form>
{% endblock %}
"""
EDIT_HTML = r"""
{% extends 'base.html' %}
{% block content %}
<div class="card p-3">
  <h5 class="mb-3">تعديل الطلب {{ txn }}</h5>
  <form method="post" class="row g-3">
    {% for c in columns %}
    <div class="col-md-6">
      <label class="form-label">{{ c }}</label>
      <input class="form-control" name="{{ c }}" value="{{ row.get(c,'') }}">
    </div>
    {% endfor %}
    <div class="col-12 text-end">
      <button class="btn btn-primary">حفظ</button>
      <a class="btn btn-outline-secondary" href="{{ url_for('home') }}">إلغاء</a>
    </div>
  </form>
</div>
{% endblock %}
"""
BULK_HTML = r"""
{% extends 'base.html' %}
{% block content %}
<div class="card p-3">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h5 class="mb-0">{{ title }}</h5>
    {% if action_label %}
      <span class="badge bg-light text-muted small">{{ action_label }}</span>
    {% endif %}
  </div>

  {# كروت إحصائيات اليوم - تظهر فقط في صفحة "تحديث إلى قيد التوصيل" #}
  {% if today_stats %}
  <div class="row g-3 mb-3">
    <div class="col-md-4">
      <div class="card border-0 shadow-sm rounded-4 h-100">
        <div class="card-body py-2">
          <div class="d-flex justify-content-between align-items-center">
            <div>
              <div class="small text-muted">طلبات اليوم (تم إنشاؤها)</div>
              <div class="fs-5 fw-bold mt-1">{{ today_stats.total_today }}</div>
            </div>
            <span class="fs-3">📅</span>
          </div>
        </div>
      </div>
    </div>
    <div class="col-md-4">
      <div class="card border-0 shadow-sm rounded-4 h-100">
        <div class="card-body py-2">
          <div class="d-flex justify-content-between align-items-center">
            <div>
              <div class="small text-muted">اليوم في قيد التوصيل</div>
              <div class="fs-5 fw-bold mt-1">{{ today_stats.shipping_today }}</div>
            </div>
            <span class="fs-3">🚚</span>
          </div>
        </div>
      </div>
    </div>
    <div class="col-md-4">
      <div class="card border-0 shadow-sm rounded-4 h-100">
        <div class="card-body py-2">
          <div class="d-flex justify-content-between align-items-center">
            <div>
              <div class="small text-muted">اليوم قيد التجهيز</div>
              <div class="fs-5 fw-bold mt-1">{{ today_stats.ready_today }}</div>
            </div>
            <span class="fs-3">🧵</span>
          </div>
        </div>
      </div>
    </div>
  </div>
  {% endif %}

  {# اختيار الموديل والبيج للمجموعة (خاص بصفحة قيد التوصيل) #}
  {% if product_name is not none %}
  <form method="post" class="row g-2 mb-3">
    <div class="col-md-5">
      <label class="form-label">بحث المنتج (بالاسم أو الكود) + تحديد الكمية</label>
      <div class="d-flex gap-2">
        <input name="product_key" class="form-control" list="invlist" placeholder="اكتب اسم المنتج أو كوده (مثال: 1001 | عباءة فاتنة)" />
        <datalist id="invlist">
          {% for it in INVENTORY_OPTIONS or [] %}
            <option value="{{ it.code }} | {{ it.name }}"></option>
          {% endfor %}
        </datalist>
        <input name="qty" class="form-control" style="max-width:110px" type="number" min="1" value="1" />
        <button class="btn btn-outline-primary" name="add_product" value="1" type="submit">إضافة</button>
      </div>

      {% if shipping_products %}
        <div class="mt-2 border rounded p-2 bg-light">
          <div class="small text-muted mb-1">المنتجات المختارة للمجموعة:</div>
          {% for sp in shipping_products %}
            <div class="d-flex justify-content-between align-items-center mb-1">
              <div>
                <span class="badge bg-dark">{{ sp.qty }}</span>
                <span class="ms-1">{{ sp.name }}</span>
                {% if sp.code %}<span class="text-muted small">({{ sp.code }})</span>{% endif %}
              </div>
              <button class="btn btn-sm btn-outline-danger" name="remove_product" value="1" type="submit"
                      onclick="this.form.rm_name.value='{{ sp.name }}'">حذف</button>
            </div>
          {% endfor %}
          <input type="hidden" name="rm_name" value="" />
          <button class="btn btn-sm btn-outline-secondary mt-1" name="clear_products" value="1" type="submit">تفريغ القائمة</button>
        </div>
      {% endif %}
    </div></div>
    <div class="col-md-5">
      <label class="form-label">اسم البيج (للمجموعة)</label>
      <select name="page_name" class="form-select">
        <option value="">بدون</option>
        {% for pg in PAGES or [] %}
          <option value="{{ pg }}" {{ 'selected' if page_name==pg else '' }}>{{ pg }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-2 align-self-end">
      <button name="apply_name" value="1" class="btn btn-outline-primary w-100">تطبيق على المجموعة</button>
    </div>
  </form>
  {% endif %}

  <form method="post" class="row g-2 mb-3">
    <div class="col-md-6">
      <label class="form-label">رقم الشحنة</label>
     <input
  type="text"
  id="txn"
  name="txn"
  class="form-control form-control-lg"
  autofocus
  autocomplete="off"
  inputmode="numeric"
  placeholder="امسح الباركود هنا">
    </div>
    <div class="col-md-3 align-self-end">
      <button class="btn btn-secondary w-100">إضافة إلى القائمة</button>
    </div>
    {% if action_label %}
    <div class="col-md-3 align-self-end">
      <button name="apply_all" value="1" class="btn btn-primary w-100">{{ action_label }}</button>
    </div>
    {% endif %}
  </form>

  <div class="table-responsive">
    <table class="table table-sm table-striped align-middle mb-0">
      <thead>
        <tr>
          {% for h in headers %}
            <th>{{ h }}</th>
          {% endfor %}
        </tr>
      </thead>
      <tbody>
        {% if items %}
          {% for r in items %}
          <tr>
            {% for h in headers %}
              <td>{{ r.get(h, '') }}</td>
            {% endfor %}
          </tr>
          {% endfor %}
        {% else %}
          <tr>
            <td colspan="{{ headers|length }}" class="text-center text-muted py-3">
              لا توجد عناصر في القائمة حاليًا.
            </td>
          </tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endblock %}
"""


PENDING_HTML = r"""
{% extends 'base.html' %}
{% block content %}
<div class="card p-3">
  <h5 class="mb-3">الطلبات قيد التوصيل</h5>
  <form method="get" class="row g-2">
    <div class="col-md-3"><label class="form-label">من</label><input name="from" type="date" class="form-control" value="{{ dfrom or '' }}"></div>
    <div class="col-md-3"><label class="form-label">إلى</label><input name="to" type="date" class="form-control" value="{{ dto or '' }}"></div>
    <div class="col-md-3 align-self-end"><button class="btn btn-secondary w-100">تصفية</button></div>
  </form>
  <div class="table-responsive mt-3">
    <table class="table table-striped">
      <thead><tr><th>Transaction ID</th><th>Time and Date</th><th>Order Price</th><th>Status</th></tr></thead>
      <tbody>
        {% for r in rows %}
        <tr><td>{{ r['Transaction ID'] }}</td><td>{{ r['Time and Date'] }}</td><td>{{ r['Order Price'] }}</td><td>{{ r['Status'] }}</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>
{% endblock %}
"""

STATS_HTML = r"""
{% extends 'base.html' %}
{% block content %}

<style>
  .stat-card {
    border-radius: 1.1rem;
    border: none;
    padding: 1.1rem 1.25rem;
  }
  .stat-card .icon-circle {
    width: 40px;
    height: 40px;
    border-radius: 999px;
    display:flex;align-items:center;justify-content:center;
    background: rgba(255,255,255,.2);
    font-size: 1.2rem;
  }
  .bg-soft-primary { background: linear-gradient(135deg,#0d6efd,#4dabf7); color:#fff; }
  .bg-soft-success { background: linear-gradient(135deg,#198754,#4dd4ac); color:#fff; }
  .bg-soft-danger  { background: linear-gradient(135deg,#dc3545,#ff8787); color:#fff; }
  .bg-soft-warning { background: linear-gradient(135deg,#ffc107,#ffe066); color:#000; }
  .small-muted { font-size:.8rem; opacity:.8; }
</style>

<div class="d-flex flex-wrap justify-content-between align-items-center mb-3">
  <div>
    <h4 class="mb-1">لوحة تحكم الطلبات 📊</h4>
    <p class="text-muted mb-0 small">
      نظرة سريعة على الواصل، الراجع، قيد التوصيل، وقيد التجهيز مع تحليل الأرباح التقديرية.
    </p>
  </div>
  <div class="mt-2 mt-md-0">
    <a href="{{ url_for('home') }}" class="btn btn-outline-secondary btn-sm">
      <i class="bi bi-list-ul"></i> صفحة الطلبات
    </a>
    <a href="{{ url_for('daily_analysis') }}" class="btn btn-outline-primary btn-sm ms-1">
      📦 تحليل يومي (فاتورة التوصيل)
    </a>
  </div>
</div>

<!-- فلاتر الداش بورد -->
<div class="card p-3 mb-3">
  <form method="get" class="row g-2 align-items-end">
    <div class="col-md-3">
      <label class="form-label">من تاريخ</label>
      <input type="date" name="from" value="{{ dfrom or '' }}" class="form-control">
    </div>
    <div class="col-md-3">
      <label class="form-label">إلى تاريخ</label>
      <input type="date" name="to" value="{{ dto or '' }}" class="form-control">
    </div>
    <div class="col-md-3">
      <label class="form-label">اسم البيج</label>
      <select name="page" class="form-select">
        <option value="">كل الصفحات</option>
        {% for p in pages %}
          <option value="{{p}}" {{ 'selected' if sel_page==p else '' }}>{{p}}</option>
        {% endfor %}
      </select>
    </div>

    <div class="col-md-2">
      <label class="form-label">أجور الشحن/طلب (د.ع)</label>
      <input type="number" name="shipping_fee" value="{{ shipping_fee or 4000 }}" class="form-control" min="0">
    </div>
    <div class="col-md-2">
      <label class="form-label">تكلفة الإعلانات (للفترة)</label>
      <input type="number" name="ads_cost" value="{{ ads_cost or 0 }}" class="form-control" min="0" step="0.01">
    </div>
    <div class="col-md-2">
      <label class="form-label">مصاريف أخرى</label>
      <input type="number" name="other_cost" value="{{ other_cost or 0 }}" class="form-control" min="0" step="0.01">
    </div>
    <div class="col-md-3 text-end">
      <button class="btn btn-primary mt-2"><i class="bi bi-funnel"></i> تطبيق</button>
      <a href="{{ url_for('stats') }}" class="btn btn-outline-secondary mt-2">إعادة تعيين</a>
    </div>
  </form>
</div>

<!-- كروت الملخص الرئيسية -->
<div class="row g-3 mb-3">
  <div class="col-md-3">
    <div class="stat-card bg-soft-primary h-100">
      <div class="d-flex justify-content-between align-items-center">
        <div>
          <div class="small-muted">العدد الكلي للطلبات</div>
          <div class="fs-4 fw-bold mt-1">{{ summary["العدد الكلي للطلبات"] }}</div>
          <div class="small mt-1">كل الطلبات في الفترة المحددة.</div>
        </div>
        <div class="icon-circle">
          <i class="bi bi-box-seam"></i>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-3">
    <div class="stat-card bg-soft-success h-100">
      <div class="d-flex justify-content-between align-items-center">
        <div>
          <div class="small-muted">الطلبات {{ status_labels[0] }}</div>
          <div class="fs-4 fw-bold mt-1">
            {{ summary["عدد " ~ status_labels[0]] }}
          </div>
          <div class="small mt-1">
            نسبة: {{ summary["نسبة " ~ status_labels[0] ~ " %"] }}%
          </div>
        </div>
        <div class="icon-circle">
          <i class="bi bi-check2-circle"></i>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-3">
    <div class="stat-card bg-soft-danger h-100">
      <div class="d-flex justify-content-between align-items-center">
        <div>
          <div class="small-muted">الطلبات {{ status_labels[1] }}</div>
          <div class="fs-4 fw-bold mt-1">
            {{ summary["عدد " ~ status_labels[1]] }}
          </div>
          <div class="small mt-1">
            نسبة: {{ summary["نسبة " ~ status_labels[1] ~ " %"] }}%
          </div>
        </div>
        <div class="icon-circle">
          <i class="bi bi-arrow-return-left"></i>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-3">
    <div class="stat-card bg-soft-warning h-100">
      <div class="d-flex justify-content-between align-items-center">
        <div>
          <div class="small-muted">إجمالي قيمة الطلبات</div>
          <div class="fs-4 fw-bold mt-1">
            {{ "%.0f"|format(summary["المجموع المالي (Order Price)"] or 0) }}
          </div>
          <div class="small mt-1">المجموع الكلي (د.ع) للفترة المحددة.</div>
        </div>
        <div class="icon-circle">
          <i class="bi bi-cash-stack"></i>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- كروت حالات إضافية -->
<div class="row g-3 mb-4">
  <div class="col-md-4">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="small-muted">قيد التوصيل</div>
            <div class="fs-5 fw-bold mt-1">
              {{ summary["عدد " ~ status_labels[2]] }}
            </div>
            <div class="small text-muted mt-1">
              نسبة: {{ summary["نسبة " ~ status_labels[2] ~ " %"] }}%
            </div>
          </div>
          <span class="fs-2">🚚</span>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-4">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="small-muted">قيد التجهيز</div>
            <div class="fs-5 fw-bold mt-1">
              {{ summary["عدد " ~ status_labels[3]] }}
            </div>
            <div class="small text-muted mt-1">
              نسبة: {{ summary["نسبة " ~ status_labels[3] ~ " %"] }}%
            </div>
          </div>
          <span class="fs-2">🧵</span>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-4">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center mb-2">
          <div>
            <div class="small-muted">إيراد الطلبات المُسلَّمة (تقديري)</div>
            <div class="fs-5 fw-bold mt-1">
              {{ "%.0f"|format(revenue or 0) }}
            </div>
          </div>
          <span class="fs-2">📈</span>
        </div>
        <div class="small text-muted">
          مجموع أسعار الطلبات بحالة "{{ status_labels[0] }}" فقط.
        </div>
      </div>
    </div>
  </div>
</div>

<!-- صافي الربح + تفصيل التكاليف -->
<div class="row g-3 mb-3">
  <div class="col-md-4">
    <div class="card border-0 shadow-sm p-3">
      <div class="d-flex justify-content-between align-items-center">
        <div>
          <div class="text-muted small">صافي الربح (للفترة/الفلتر)</div>
          <div class="fs-4 fw-bold mt-1">{{ "%.0f"|format(net_profit or 0) }}</div>
        </div>
        <span class="fs-2">💰</span>
      </div>
      <div class="small text-muted mt-2">
        صافي الربح = الإيراد (الواصل) - تكلفة الخام/الخياطة - الشحن - الإعلان - مصاريف أخرى
      </div>
    </div>
  </div>

  <div class="col-md-8">
    <div class="card border-0 shadow-sm p-3">
      <div class="fw-bold mb-2">تفصيل التكاليف</div>
      <div class="row g-2">
        <div class="col-md-3">
          <div class="small text-muted">تكلفة الخام + الخياطة</div>
          <div class="fw-bold">{{ "%.0f"|format(cogs_total or 0) }}</div>
        </div>
        <div class="col-md-3">
          <div class="small text-muted">الشحن ({{ shipping_fee or 0 }} × {{ status_counts[0] or 0 }})</div>
          <div class="fw-bold">{{ "%.0f"|format(shipping_total or 0) }}</div>
        </div>
        <div class="col-md-3">
          <div class="small text-muted">الإعلانات</div>
          <div class="fw-bold">{{ "%.0f"|format(ads_cost or 0) }}</div>
        </div>
        <div class="col-md-3">
          <div class="small text-muted">مصاريف أخرى</div>
          <div class="fw-bold">{{ "%.0f"|format(other_cost or 0) }}</div>
        </div>
      </div>
    </div>
  </div>
</div>


<!-- أفضل المنتجات -->
<div class="row g-3 mb-3">
  <div class="col-md-6">
    <div class="card border-0 shadow-sm p-3">
      <div class="fw-bold mb-2">أكثر منتج يتم طلبه (حسب عدد القطع)</div>
      <div class="table-responsive">
        <table class="table table-sm table-striped mb-0">
          <thead><tr><th>#</th><th>المنتج</th><th>عدد القطع</th></tr></thead>
          <tbody>
            {% for nm, q in top_ordered or [] %}
              <tr><td>{{ loop.index }}</td><td>{{ nm }}</td><td>{{ q }}</td></tr>
            {% else %}
              <tr><td colspan="3" class="text-center text-muted">لا توجد بيانات</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      <div class="small text-muted mt-2">يعتمد على اسم المنتج من المخزن/حقل Items (وليس السعر).</div>
    </div>
  </div>

  <div class="col-md-6">
    <div class="card border-0 shadow-sm p-3">
      <div class="fw-bold mb-2">أفضل منتج من ناحية التسليم (راجع قليل)</div>
      <div class="table-responsive">
        <table class="table table-sm table-striped mb-0">
          <thead><tr><th>#</th><th>المنتج</th><th>تسليم</th><th>راجع</th><th>نسبة الراجع</th></tr></thead>
          <tbody>
            {% for r in best_delivered or [] %}
              <tr>
                <td>{{ loop.index }}</td>
                <td>{{ r.name }}</td>
                <td>{{ r.delivered_qty }}</td>
                <td>{{ r.returned_qty }}</td>
                <td>{{ "%.1f"|format(r.return_rate or 0) }}%</td>
              </tr>
            {% else %}
              <tr><td colspan="5" class="text-center text-muted">لا توجد بيانات</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>
<div class="row g-3 mb-3">
  <div class="col-md-4">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="small text-muted">مبلغ الطلبات {{ status_labels[1] }}</div>
            <div class="fs-5 fw-bold mt-1">
              {{ "%.0f"|format(summary["مبلغ " ~ status_labels[1]] or 0) }} د.ع
            </div>
            <div class="small text-muted mt-1">
              عددها: {{ summary["عدد " ~ status_labels[1]] }}
            </div>
          </div>
          <span class="fs-2">↩️</span>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-4">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="small text-muted">مبلغ الطلبات {{ status_labels[2] }}</div>
            <div class="fs-5 fw-bold mt-1">
              {{ "%.0f"|format(summary["مبلغ " ~ status_labels[2]] or 0) }} د.ع
            </div>
            <div class="small text-muted mt-1">
              عددها: {{ summary["عدد " ~ status_labels[2]] }}
            </div>
          </div>
          <span class="fs-2">🚚</span>
        </div>
      </div>
    </div>
  </div>

  <div class="col-md-4">
    <div class="card border-0 shadow-sm rounded-4 h-100">
      <div class="card-body">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="small text-muted">مبلغ الطلبات {{ status_labels[3] }}</div>
            <div class="fs-5 fw-bold mt-1">
              {{ "%.0f"|format(summary["مبلغ " ~ status_labels[3]] or 0) }} د.ع
            </div>
            <div class="small text-muted mt-1">
              عددها: {{ summary["عدد " ~ status_labels[3]] }}
            </div>
          </div>
          <span class="fs-2">⏳</span>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- الجارتات -->
<div class="row g-3">
  <div class="col-xl-6">
    <div class="card p-3 shadow-sm h-100">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <h6 class="mb-0">توزيع الحالات</h6>
        <span class="badge bg-light text-dark small">واصل / راجع / قيد التوصيل / قيد التجهيز</span>
      </div>
      <canvas id="statusChart" height="200"></canvas>
    </div>
  </div>
  <div class="col-xl-6">
    <div class="card p-3 shadow-sm h-100">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <h6 class="mb-0">ترند عدد الطلبات اليومي</h6>
        <span class="badge bg-light text-dark small">آخر الأيام المسجَّلة</span>
      </div>
      <canvas id="dailyChart" height="200"></canvas>
    </div>
  </div>
</div>

<!-- جداول تحليلية: حسب السعر / حسب المنتج -->
<div class="accordion mt-4" id="statsAccordion">
  <div class="accordion-item">
    <h2 class="accordion-header" id="headingPrice">
      <button class="accordion-button" type="button" data-bs-toggle="collapse" data-bs-target="#collapsePrice">
        إحصاءات حسب السعر (السعر / عدد الطلبات / الراجع / الواصل)
      </button>
    </h2>
    <div id="collapsePrice" class="accordion-collapse collapse show" data-bs-parent="#statsAccordion">
      <div class="accordion-body">
        {% if by_price and by_price|length %}
        <div class="table-responsive">
          <table class="table table-sm table-striped align-middle">
            <thead>
              <tr>
                {% for c in price_cols %}<th>{{ c }}</th>{% endfor %}
              </tr>
            </thead>
            <tbody>
              {% for r in by_price %}
              <tr>
                {% for c in price_cols %}
                  <td>{{ r.get(c, '') }}</td>
                {% endfor %}
              </tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
        {% else %}
          <p class="text-muted mb-0 small">لا توجد بيانات كافية لعرض الإحصاءات حسب السعر.</p>
        {% endif %}
      </div>
    </div>
  </div>

  <div class="accordion-item mt-2">
    <h2 class="accordion-header" id="headingProduct">
      <button class="accordion-button collapsed" type="button" data-bs-toggle="collapse" data-bs-target="#collapseProduct">
        تحليل الربح والخسارة حسب المنتج (من بيانات المخزن)
      </button>
    </h2>
    <div id="collapseProduct" class="accordion-collapse collapse" data-bs-parent="#statsAccordion">
      <div class="accordion-body">
        {% if by_product and by_product|length %}
        <div class="table-responsive">
          <table class="table table-sm table-striped align-middle">
            <thead>
              <tr>
                {% for c in product_cols %}<th>{{ c }}</th>{% endfor %}
              </tr>
            </thead>
            <tbody>
              {% for r in by_product %}
              <tr>
                {% for c in product_cols %}
                  <td>{{ r.get(c, '') }}</td>
                {% endfor %}
              </tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
        {% else %}
          <p class="text-muted mb-0 small">
            أرباح المنتجات تُحتسب فقط عندما تكون بيانات الأسعار/التكاليف موجودة في المخزن ومرتبطة بالطلبات.
          </p>
        {% endif %}
      </div>
    </div>
  </div>
</div>

<!-- جدول ترند الأيام (اختياري تحت الجارت) -->
<div class="card p-3 mt-4">
  <h6 class="mb-2">ملخص يومي لعدد الطلبات</h6>
  {% if daily and daily|length %}
  <div class="table-responsive">
    <table class="table table-sm table-striped align-middle">
      <thead>
        <tr>
          <th>التاريخ</th>
          <th>عدد الطلبات</th>
          <th>الترند</th>
        </tr>
      </thead>
      <tbody>
        {% for r in daily %}
        <tr>
          <td>{{ r["Date"] }}</td>
          <td>{{ r["Order Count"] }}</td>
          <td>
            {% if r["Trend"] == "ارتفاع" %}
              <span class="text-success">📈 ارتفاع</span>
            {% elif r["Trend"] == "انخفاض" %}
              <span class="text-danger">📉 انخفاض</span>
            {% else %}
              <span class="text-muted">➖ ثابت</span>
            {% endif %}
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% else %}
    <p class="text-muted mb-0 small">لا توجد بيانات كافية لعرض الترند اليومي.</p>
  {% endif %}
</div>

<!-- Chart.js -->
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script>
  const statusLabels = {{ status_labels|tojson|safe }};
  const statusCounts = {{ status_counts|tojson|safe }};
  const dailyData    = {{ daily|tojson|safe }};

  // Doughnut chart لحالات الطلبات
  const statusCtx = document.getElementById('statusChart');
  if (statusCtx && statusLabels.length) {
    new Chart(statusCtx, {
      type: 'doughnut',
      data: {
        labels: statusLabels,
        datasets: [{
          data: statusCounts,
          backgroundColor: ['#0d6efd','#198754','#ffc107','#dc3545'],
          borderWidth: 1
        }]
      },
      options: {
        plugins: {
          legend: { position: 'bottom' }
        }
      }
    });
  }

  // Line chart للترند اليومي
  const dailyCtx = document.getElementById('dailyChart');
  if (dailyCtx && dailyData && dailyData.length) {
    const labels = dailyData.map(d => d.Date);
    const counts = dailyData.map(d => d["Order Count"]);
    new Chart(dailyCtx, {
      type: 'line',
      data: {
        labels: labels,
        datasets: [{
          label: 'عدد الطلبات',
          data: counts,
          tension: 0.3,
          fill: false,
          borderWidth: 2
        }]
      },
      options: {
        plugins: {
          legend: { display: false }
        },
        scales: {
          x: {
            ticks: { autoSkip: true, maxTicksLimit: 7 }
          }
        }
      }
    });
  }
</script>

{% endblock %}
"""

DAILY_ANALYSIS_HTML = r"""
{% extends 'base.html' %}
{% block content %}

<style>
  .kpi {
    border-radius: 1rem;
    border: 0;
    box-shadow: 0 2px 10px rgba(0,0,0,.06);
  }
  .kpi .label{font-size:.85rem;color:#6c757d}
  .kpi .value{font-size:1.4rem;font-weight:800}
  .pill{border-radius:999px;padding:.2rem .65rem;font-size:.8rem}
  .table thead th{white-space:nowrap}
</style>

<div class="d-flex flex-wrap justify-content-between align-items-center mb-3">
  <div>
    <h4 class="mb-1">تحليل يومي حسب فاتورة شركة التوصيل</h4>
    <div class="text-muted small">
      اختر فترة الفاتورة (تاريخ رفع/تحديث التوصيل) وحدد كم يوم ترجع للخلف لالتقاط الطلبات الأصلية (افتراضي: يوم واحد).
    </div>
  </div>
  <div class="mt-2 mt-md-0 d-flex gap-2">
    <a href="{{ url_for('home') }}" class="btn btn-outline-secondary btn-sm">رجوع للطلبات</a>
    <a href="{{ url_for('stats') }}" class="btn btn-outline-dark btn-sm">الإحصائيات العامة</a>
  </div>
</div>

<form method="get" action="{{ url_for('daily_analysis') }}" class="card p-3 mb-3">
  <div class="row g-2 align-items-end">
    <div class="col-md-3">
      <label class="form-label">من (تاريخ الفاتورة)</label>
      <input type="date" name="dfrom" class="form-control" value="{{ dfrom or '' }}" required>
    </div>
    <div class="col-md-3">
      <label class="form-label">إلى (تاريخ الفاتورة)</label>
      <input type="date" name="dto" class="form-control" value="{{ dto or '' }}" required>
    </div>
    <div class="col-md-2">
      <label class="form-label">الرجوع للخلف (أيام)</label>
      <input type="number" name="offset" class="form-control" min="0" value="{{ offset or 1 }}">
      <div class="form-text">1 = طلبات يوم أمس</div>
    </div>
    <div class="col-md-2">
      <label class="form-label">أجرة التوصيل/طلب</label>
      <input type="number" step="0.01" name="ship_fee" class="form-control" value="{{ ship_fee or 0 }}">
    </div>
    <div class="col-md-2">
      <label class="form-label">إعلانات (للفترة)</label>
      <input type="number" step="0.01" name="ads" class="form-control" value="{{ ads or 0 }}">
    </div>
    <div class="col-md-2">
      <label class="form-label">مصاريف أخرى</label>
      <input type="number" step="0.01" name="other" class="form-control" value="{{ other or 0 }}">
    </div>
    <div class="col-md-2 d-grid">
      <button class="btn btn-primary">عرض التحليل</button>
    </div>
  </div>
</form>

{% if ready is not none %}
<div class="row g-3 mb-3">
  <div class="col-md-3">
    <div class="card kpi p-3">
      <div class="label">طلبات الفترة (حسب إنشاء الطلب)</div>
      <div class="value">{{ base_total }}</div>
      <div class="text-muted small">من {{ base_from }} إلى {{ base_to }}</div>
    </div>
  </div>
  <div class="col-md-3">
    <div class="card kpi p-3">
      <div class="label">تم التوصيل (ضمن الفاتورة)</div>
      <div class="value text-success">{{ delivered }}</div>
      <div class="text-muted small">إيراد: {{ "%.0f"|format(revenue or 0) }}</div>
    </div>
  </div>
  <div class="col-md-3">
    <div class="card kpi p-3">
      <div class="label">راجع (ضمن الفاتورة)</div>
      <div class="value text-danger">{{ returned }}</div>
      <div class="text-muted small">نسبة الراجع: {{ return_rate }}%</div>
    </div>
  </div>
  <div class="col-md-3">
    <div class="card kpi p-3">
      <div class="label">معلّقة</div>
      <div class="value">{{ pending_total }}</div>
      <div class="text-muted small">
        <span class="pill bg-warning text-dark">قيد التجهيز: {{ ready }}</span>
        <span class="pill bg-info text-dark">قيد التوصيل: {{ shipping }}</span>
      </div>
    </div>
  </div>
</div>

<div class="row g-3 mb-3">
  <div class="col-lg-4">
    <div class="card p-3">
      <div class="fw-bold mb-2">صافي الربح</div>
      <div class="fs-3 fw-bold">{{ "%.0f"|format(net_profit or 0) }}</div>
      <div class="text-muted small mt-2">
        صافي الربح = الإيراد (الواصل) - تكلفة الخام/الخياطة/الإكسسوارات - الشحن - الإعلانات - مصاريف أخرى
      </div>
      <hr class="my-2">
      <div class="row g-2 small">
        <div class="col-6"><span class="text-muted">تكلفة البضاعة</span><div class="fw-bold">{{ "%.0f"|format(cogs_total or 0) }}</div></div>
        <div class="col-6"><span class="text-muted">الشحن</span><div class="fw-bold">{{ "%.0f"|format(shipping_total or 0) }}</div></div>
        <div class="col-6"><span class="text-muted">الإعلانات</span><div class="fw-bold">{{ "%.0f"|format(ads or 0) }}</div></div>
        <div class="col-6"><span class="text-muted">مصاريف أخرى</span><div class="fw-bold">{{ "%.0f"|format(other or 0) }}</div></div>
      </div>
    </div>
  </div>

  <div class="col-lg-8">
    <div class="card p-3">
      <div class="fw-bold mb-2">أفضل المنتجات</div>
      <div class="row g-3">
        <div class="col-md-6">
          <div class="small text-muted mb-1">الأكثر توصيلًا (قطع)</div>
          <div class="table-responsive">
            <table class="table table-sm table-striped align-middle mb-0">
              <thead><tr><th>المنتج</th><th>قطع واصل</th><th>قطع راجع</th><th>نسبة الراجع%</th></tr></thead>
              <tbody>
                {% for r in top_delivered %}
                  <tr>
                    <td>{{ r.name }}</td>
                    <td class="fw-bold">{{ r.delivered_qty }}</td>
                    <td class="text-danger">{{ r.returned_qty }}</td>
                    <td>{{ "%.2f"|format(r.return_rate) }}</td>
                  </tr>
                {% endfor %}
                {% if not top_delivered %}
                  <tr><td colspan="4" class="text-center text-muted py-2">لا توجد بيانات.</td></tr>
                {% endif %}
              </tbody>
            </table>
          </div>
        </div>

        <div class="col-md-6">
          <div class="small text-muted mb-1">الأكثر راجعًا (قطع)</div>
          <div class="table-responsive">
            <table class="table table-sm table-striped align-middle mb-0">
              <thead><tr><th>المنتج</th><th>قطع راجع</th><th>قطع واصل</th><th>نسبة الراجع%</th></tr></thead>
              <tbody>
                {% for r in top_returned %}
                  <tr>
                    <td>{{ r.name }}</td>
                    <td class="fw-bold text-danger">{{ r.returned_qty }}</td>
                    <td>{{ r.delivered_qty }}</td>
                    <td>{{ "%.2f"|format(r.return_rate) }}</td>
                  </tr>
                {% endfor %}
                {% if not top_returned %}
                  <tr><td colspan="4" class="text-center text-muted py-2">لا توجد بيانات.</td></tr>
                {% endif %}
              </tbody>
            </table>
          </div>
        </div>

      </div>
    </div>
  </div>
</div>

<div class="card p-3 mb-3">
  <div class="d-flex justify-content-between align-items-center mb-2">
    <div class="fw-bold">تحليل المتاجر (Page Name)</div>
    <div class="text-muted small">مرتب حسب الربح الصافي (مع توزيع الإعلانات حسب الإيراد)</div>
  </div>
  {% if best_store_by_delivered or best_store_by_delivery_rate %}
  <div class="row g-2 mb-2">
    <div class="col-md-6">
      <div class="alert alert-success py-2 mb-0">
        <div class="small text-muted">أفضل متجر (أكثر واصل)</div>
        <div class="fw-bold">
          {{ best_store_by_delivered.page if best_store_by_delivered else '—' }}
          — {{ best_store_by_delivered.delivered_orders if best_store_by_delivered else 0 }} واصل
        </div>
      </div>
    </div>
    <div class="col-md-6">
      <div class="alert alert-primary py-2 mb-0">
        <div class="small text-muted">أفضل متجر (أفضل نسبة وصول)</div>
        <div class="fw-bold">
          {{ best_store_by_delivery_rate.page if best_store_by_delivery_rate else '—' }}
          — {{ "%.2f"|format(best_store_by_delivery_rate.delivery_rate) if best_store_by_delivery_rate else "0.00" }}%
        </div>
      </div>
    </div>
  </div>
  {% endif %}
  <div class="table-responsive">
    <table class="table table-sm table-hover align-middle mb-0">
      <thead class="table-light">
        <tr>
          <th>المتجر</th>
          <th>طلبات واصل</th>
          <th>طلبات راجع</th>
          <th>طلبات قيد التوصيل</th>
          <th>نسبة الوصول%</th>
          <th>إيراد</th>
          <th>مبلغ قيد التوصيل</th>
          <th>تكلفة بضاعة</th>
          <th>شحن</th>
          <th>حصة إعلانات</th>
          <th>صافي ربح</th>
          <th>نسبة الراجع%</th>
        </tr>
      </thead>
      <tbody>
        {% for r in page_rows %}
          <tr>
            <td class="fw-bold">{{ r.page }}</td>
            <td>{{ r.delivered_orders }}</td>
            <td class="text-danger">{{ r.returned_orders }}</td>
            <td>{{ r.shipping_orders }}</td>
            <td>{{ "%.2f"|format(r.delivery_rate) }}</td>
            <td>{{ "%.0f"|format(r.revenue) }}</td>
            <td>{{ "%.0f"|format(r.shipping_amount) }}</td>
            <td>{{ "%.0f"|format(r.cogs) }}</td>
            <td>{{ "%.0f"|format(r.shipping) }}</td>
            <td>{{ "%.0f"|format(r.ads_share) }}</td>
            <td class="{% if r.net_profit >= 0 %}text-success{% else %}text-danger{% endif %} fw-bold">
              {{ "%.0f"|format(r.net_profit) }}
            </td>
            <td>{{ "%.2f"|format(r.return_rate) }}</td>
          </tr>
        {% endfor %}
        {% if not page_rows %}
          <tr><td colspan="12" class="text-center text-muted py-3">لا توجد بيانات للمتاجر ضمن الفلتر.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>

{% endif %}

{% endblock %}
"""




# Register templates in-memory (use DictLoader so `{% extends 'base.html' %}` works)
from jinja2 import DictLoader
app.jinja_loader = DictLoader({
    'base.html': BASE_HTML,
    'login.html': LOGIN_HTML,
    'home.html': HOME_HTML,
    'edit.html': EDIT_HTML,
    'bulk.html': BULK_HTML,
    'pending.html': PENDING_HTML,
    'stats.html': STATS_HTML,
    'inventory.html': INVENTORY_HTML,
    'products.html': PRODUCTS_HTML,
    'seamstress.html': SEAMSTRESS_HTML,
    'issues.html': ISSUES_HTML,
    'cutting.html': CUTTING_HTML,
})

# --------------------------- AUTH DECORATOR ----------------------------
from functools import wraps

def login_required(fn):
    @wraps(fn)
    def _wrap(*args, **kwargs):
        if not session.get('auth'):
            return redirect(url_for('login'))
        return fn(*args, **kwargs)
    return _wrap

# ---------------------------- EXTRACTORS -------------------------------

def extract_from_text(text: str):
    text = normalize_digits(text)
    lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
    full = "\n".join(lines)

    txn = None
    m = re.search(r'رقم\s*(?:الشحنة|الوصل|الطلب)\s*[:：]?\s*(\d{6,})', full)
    if m:
        txn = m.group(1)
    else:
        m2 = re.search(r'(?<!\d)(?!07)\d{8,14}(?!\d)', full)
        if m2:
            txn = m2.group(0)

    phones = re.findall(r'(07\d{9})', full)
    seen = set(); uniq = []
    for p in phones:
        if p not in seen:
            seen.add(p); uniq.append(p)
    phone_str = ", ".join(uniq) if uniq else None

    def parse_price_from_lines(ls):
        label = r'(المبلغ(?:\s*الكلي)?(?:\s*ل?لفاتورة)?|السعر|قيمة\s*الطلب|Price|Total|IQD|دينار|د\.ع)'
        num   = r'(\d{1,3}(?:,\d{3})+|\d{4,9})'
        for ln in ls:
            cand = ln
            m1 = re.search(fr'{label}[^\d]{{0,40}}{num}', cand)
            if m1:
                v = int(m1.group(2 if m1.lastindex and m1.lastindex >= 2 else 1).replace(",", ""))
                if str(v).endswith("000"):
                    return v
            m2 = re.search(fr'{num}\s*{label}', cand)
            if m2:
                v = int(m2.group(1).replace(",", ""))
                if str(v).endswith("000"):
                    return v
        all_nums = [int(n.replace(",", "")) for n in re.findall(r'(\d{1,3}(?:,\d{3})+|\d{4,9})', full)]
        candidates = [n for n in all_nums if str(n).endswith("000")]
        return max(candidates) if candidates else None

    order_price = parse_price_from_lines(lines)

    def parse_address(ls):
        for i, ln in enumerate(ls):
            m = re.search(r'(?:العنوان|عنوان\s*الزبون|Address)\s*[:：]?\s*(.+)$', ln)
            if m and m.group(1).strip():
                return m.group(1).strip(" ,:؛-")
            if any(lbl in ln for lbl in ("العنوان", "عنوان الزبون", "Address")):
                if i+1 < len(ls) and ls[i+1].strip():
                    return ls[i+1].strip(" ,:؛-")
                if i > 0 and ls[i-1].strip():
                    return ls[i-1].strip(" ,:؛-")
        return None

    address = parse_address(lines)
    return txn, phone_str, order_price, address
# ---------------------------- SEAM / SEW STORE --------------------------
# ------------------------------ ISSUES STORE ----------------------------
class IssuesStore:
    COLS = ['ID', 'Title', 'Description', 'ImagePath', 'Status', 'Solver', 'CreatedAt']

    def __init__(self, root_dir: Path):
        self.path = root_dir / 'issues.xlsx'
        self.df = self._load()

    def _load(self):
        if not self.path.exists():
            df = pd.DataFrame(columns=self.COLS)
            df.to_excel(self.path, index=False)
            return df
        df = pd.read_excel(self.path)
        for c in self.COLS:
            if c not in df.columns:
                df[c] = pd.NA
        return df[self.COLS]

    def _touch_mtime(self):
        try:
            self._last_mtime = os.path.getmtime(self.path)
        except Exception:
            self._last_mtime = None

    def reload_if_changed(self):
        try:
            current = os.path.getmtime(self.path)
        except Exception:
            return
        if self._last_mtime is None:
            self._last_mtime = current
            return
        if current != self._last_mtime:
            try:
                self.df = self._load()
            finally:
                self._touch_mtime()


    def _save(self):
        self.df.to_excel(self.path, index=False)

    def _next_id(self):
        if self.df.empty:
            return 1
        vals = pd.to_numeric(self.df['ID'], errors='coerce').dropna()
        return int(vals.max() + 1) if len(vals) else 1

    def add_issue(self, title, desc='', img_path=''):
        new_id = self._next_id()
        row = {
            'ID': new_id,
            'Title': title,
            'Description': desc,
            'ImagePath': img_path,
            'Status': 'Open',
            'Solver': '',
            'CreatedAt': now_str(),
        }
        self.df = pd.concat([self.df, pd.DataFrame([row])], ignore_index=True)
        self._save()

    def solve(self, iid, solver):
        idx = self.df[self.df['ID'] == iid].index
        if not len(idx):
            return
        i = idx[0]
        self.df.at[i, 'Status'] = 'Solved'
        self.df.at[i, 'Solver'] = solver
        self._save()

    def delete(self, iid):
        self.df = self.df[self.df['ID'] != iid]
        self._save()


issues = IssuesStore(_data_root)

class SeamStore:
    MAST_COLS = ['ID', 'Name', 'Phone', 'Notes', 'Active']
    LOG_COLS = ['LogID', 'Date', 'SeamstressID', 'Model', 'Pieces', 'UnitCost', 'Total', 'Paid']

    def __init__(self, root_dir: Path):
        self.mast_path = root_dir / 'seamstresses.xlsx'
        self.log_path = root_dir / 'sewing_logs.xlsx'
        self.mast = self._load_mast()
        self.log = self._load_log()

    def _load_mast(self):
        if not self.mast_path.exists():
            df = pd.DataFrame(columns=self.MAST_COLS)
            df.to_excel(self.mast_path, index=False)
            return df
        df = pd.read_excel(self.mast_path)
        for c in self.MAST_COLS:
            if c not in df.columns:
                df[c] = pd.NA
        return df[self.MAST_COLS]

    def _load_log(self):
        if not self.log_path.exists():
            df = pd.DataFrame(columns=self.LOG_COLS)
            df.to_excel(self.log_path, index=False)
            return df
        df = pd.read_excel(self.log_path)
        for c in self.LOG_COLS:
            if c not in df.columns:
                df[c] = pd.NA
        return df[self.LOG_COLS]

    def _save_mast(self):
        self.mast.to_excel(self.mast_path, index=False)

    def _save_log(self):
        self.log.to_excel(self.log_path, index=False)

    def _next_id(self, col_name, df):
        if df.empty or col_name not in df.columns:
            return 1
        vals = pd.to_numeric(df[col_name], errors='coerce')
        vals = vals.dropna()
        return int(vals.max() + 1) if len(vals) else 1

    def add_seamstress(self, name, phone='', notes=''):
        new_id = self._next_id('ID', self.mast)
        row = {
            'ID': new_id,
            'Name': name,
            'Phone': phone,
            'Notes': notes,
            'Active': True,
        }
        self.mast = pd.concat([self.mast, pd.DataFrame([row])], ignore_index=True)
        self._save_mast()

    def update_seamstress(self, sid, **kwargs):
        idx = self.mast[self.mast['ID'] == sid].index
        if not len(idx):
            return
        i = idx[0]
        for k, v in kwargs.items():
            if k in self.mast.columns:
                self.mast.at[i, k] = v
        self._save_mast()

    def delete_seamstress(self, sid):
        self.mast = self.mast[self.mast['ID'] != sid]
        # حذف السجلات المرتبطة من سجل الإنجاز
        self.log = self.log[self.log['SeamstressID'] != sid]
        self._save_mast()
        self._save_log()

    def add_log(self, sid, model, pieces, unit_cost):
        log_id = self._next_id('LogID', self.log)
        total = float(pieces) * float(unit_cost)
        row = {
            'LogID': log_id,
            'Date': date.today().isoformat(),
            'SeamstressID': sid,
            'Model': model,
            'Pieces': pieces,
            'UnitCost': unit_cost,
            'Total': total,
            'Paid': False,
        }
        self.log = pd.concat([self.log, pd.DataFrame([row])], ignore_index=True)
        self._save_log()
        # زيادة المخزون تلقائيًا بالموديل وعدد القطع
        try:
            inventory.adjust_quantity(model, pieces, movement_type='Production', ref=f'SEAM:{log_id}', notes=f'SeamstressID={sid}')
        except Exception:
            pass

    def set_paid(self, log_id, paid: bool):
        idx = self.log[self.log['LogID'] == log_id].index
        if not len(idx):
            return
        self.log.at[idx[0], 'Paid'] = bool(paid)
        self._save_log()


# إنشاء كائن seams

seams = SeamStore(_data_root)


# ------------------------------- INVENTORY ------------------------------
# ------------------------------- INVENTORY ------------------------------
class InventoryMovementStore:
    """Log every inventory movement (production/withdraw/return/manual)."""

    COLS = [
        'MoveID',
        'DateTime',
        'Date',
        'Product Code',
        'Product Name',
        'Delta',
        'Movement Type',
        'Ref',
        'Notes',
    ]

    def __init__(self, path: str):
        self.path = str(path)
        self._last_mtime = None
        self.df = self._load()
        self._touch_mtime()

    def _touch_mtime(self):
        try:
            self._last_mtime = os.path.getmtime(self.path)
        except Exception:
            self._last_mtime = None

    def reload_if_changed(self):
        """Reload only if file changed on disk (faster than reloading every request)."""
        try:
            current = os.path.getmtime(self.path)
        except Exception:
            return
        if self._last_mtime is None or current != self._last_mtime:
            self.reload()
            self._touch_mtime()

    def _load(self):
        p = Path(self.path)
        if not p.exists():
            df = pd.DataFrame(columns=self.COLS)
            df.to_excel(self.path, index=False)
            return df
        try:
            df = pd.read_excel(self.path)
        except Exception:
            df = pd.DataFrame(columns=self.COLS)
            df.to_excel(self.path, index=False)
            return df
        for c in self.COLS:
            if c not in df.columns:
                df[c] = pd.NA
        df = df[self.COLS].copy()
        # Ensure types
        df['Product Code'] = df['Product Code'].astype(str).str.strip()
        df['Product Name'] = df['Product Name'].astype(str).fillna('').str.strip()
        df['Movement Type'] = df['Movement Type'].astype(str).fillna('').str.strip()
        df['Delta'] = pd.to_numeric(df['Delta'], errors='coerce').fillna(0).astype(int)
        return df

    def save(self):
        self.df.to_excel(self.path, index=False)
        self._touch_mtime()

    def reload(self):
        self.df = self._load()
        return self.df

    def _next_id(self) -> int:
        try:
            s = pd.to_numeric(self.df['MoveID'], errors='coerce')
            m = int(s.max()) if (s is not None and not s.empty and pd.notna(s.max())) else 0
            return m + 1
        except Exception:
            return len(self.df) + 1

    def add(self, product_code: str, product_name: str, delta: int,
            movement_type: str = "Manual", ref: str = "", notes: str = ""):
        now = datetime.now()
        row = {
            'MoveID': self._next_id(),
            'DateTime': now.strftime('%Y-%m-%d %H:%M:%S'),
            'Date': now.date().isoformat(),
            'Product Code': str(product_code or '').strip(),
            'Product Name': str(product_name or '').strip(),
            'Delta': int(delta),
            'Movement Type': str(movement_type or '').strip(),
            'Ref': str(ref or '').strip(),
            'Notes': str(notes or '').strip(),
        }
        self.df = pd.concat([self.df, pd.DataFrame([row], columns=self.COLS)], ignore_index=True)
        self.save()

    def filter_by_product_code(self, code: str):
        code = str(code or '').strip()
        if self.df is None or self.df.empty:
            return pd.DataFrame(columns=self.COLS)
        x = self.df.copy()
        x['Product Code'] = x['Product Code'].astype(str).str.strip()
        return x[x['Product Code'] == code]

    def filter_by_date(self, d: str):
        d = str(d or '').strip()
        if self.df is None or self.df.empty:
            return pd.DataFrame(columns=self.COLS)
        return self.df[self.df['Date'].astype(str) == d]

class InventoryStore:
    COLS = [
        'Product Code',       # كود المنتج
        'Product Name',       # اسم المنتج / الموديل
        'Type',               # نوع البضاعة (عباية، أطفال..)
        'Quantity',           # الكمية بالمخزن (عدد القطع)
        'Fabric Meters',      # إجمالي أمتار القماش في المخزن (اختياري)
        'Meters per Unit',    # عدد الأمتار المطلوبة للقطعة الواحدة
        'Fabric Meter Price', # سعر متر الخام
        'Sewing Cost',        # سعر الخياطة للقطعة
        'Accessories Cost',   # تكلفة الإكسسوارات
        'Extra Costs',        # تكاليف إضافية أخرى
        'Sale Price',         # سعر البيع للقطعة
    ]

    def __init__(self, path):
        # نخلي ملف المخزن باسم ثابت بجانب ملف الطلبات
        self.path = str(Path(path).with_name('inventory.xlsx'))
        self._last_mtime = None
        self.movements = InventoryMovementStore(Path(self.path).with_name('inventory_movements.xlsx'))
        self.df = self._load()
        self._touch_mtime()

    def _touch_mtime(self):
        try:
            self._last_mtime = os.path.getmtime(self.path)
        except Exception:
            self._last_mtime = None

    def reload_if_changed(self):
        """Reload inventory only if the backing file changed on disk."""
        try:
            current = os.path.getmtime(self.path)
        except Exception:
            return
        if self._last_mtime is None or current != self._last_mtime:
            self.reload()
            self._touch_mtime()

    def _load(self):
        p = Path(self.path)
        if not p.exists():
            df = pd.DataFrame(columns=self.COLS)
            df.to_excel(self.path, index=False)
            return df

        # لو الملف تالف أو ما ينقري نعيد إنشاءه حتى ما يوقع البرنامج
        try:
            df = pd.read_excel(self.path)
        except Exception:
            df = pd.DataFrame(columns=self.COLS)
            df.to_excel(self.path, index=False)
            return df

        for c in self.COLS:
            if c not in df.columns:
                df[c] = pd.NA

        df = df[self.COLS].copy()
        df['Product Code'] = df['Product Code'].astype(str).fillna('').str.strip()
        df['Product Name'] = df['Product Name'].astype(str).fillna('').str.strip()
        df['Type'] = df['Type'].astype(str).fillna('').str.strip()

        # Safe numeric conversions
        df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0).astype(int)
        for c in ['Fabric Meters','Meters per Unit','Fabric Meter Price','Sewing Cost','Accessories Cost','Extra Costs','Sale Price']:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)

        return df

    def save(self):
        self.df.to_excel(self.path, index=False)
        self._touch_mtime()

    def reload(self):
        """Reload inventory from disk to avoid stale values."""
        self.df = self._load()
        return self.df

    def next_code(self):
        prefix = 'INV'
        nums = []
        for x in self.df.get('Product Code', pd.Series([], dtype=str)).dropna().astype(str):
            x = x.strip()
            if x.startswith(prefix):
                try:
                    nums.append(int(x.replace(prefix, '') or 0))
                except Exception:
                    pass
        n = (max(nums) if nums else 0) + 1
        return f'{prefix}{n:04d}'

    def add_item(self, row):
        row = {**{c: pd.NA for c in self.COLS}, **row}
        # Normalize
        row['Product Code'] = str(row.get('Product Code') or '').strip()
        row['Product Name'] = str(row.get('Product Name') or '').strip()
        row['Type'] = str(row.get('Type') or '').strip()
        row['Quantity'] = int(float(row.get('Quantity') or 0) or 0)
        for c in ['Fabric Meters','Meters per Unit','Fabric Meter Price','Sewing Cost','Accessories Cost','Extra Costs','Sale Price']:
            try:
                row[c] = float(row.get(c) or 0)
            except Exception:
                row[c] = 0.0

        self.df = pd.concat([self.df, pd.DataFrame([row], columns=self.COLS)], ignore_index=True)
        self.save()

    def get_by_code(self, code: str):
        code = str(code).strip()
        d = self.df[self.df['Product Code'].astype(str) == code]
        if d.empty:
            return None
        return d.iloc[0].to_dict()

    def find_index_by_code(self, code: str):
        code = str(code or '').strip()
        idx = self.df[self.df['Product Code'].astype(str) == code].index
        return idx[0] if len(idx) else None

    def find_index_by_name(self, name: str):
        name = str(name or '').strip()
        idx = self.df[self.df['Product Name'].astype(str) == name].index
        return idx[0] if len(idx) else None

    def update_item(self, code: str, **kwargs):
        code = str(code).strip()
        i = self.find_index_by_code(code)
        if i is None:
            return 0
        for k, v in kwargs.items():
            if k not in self.df.columns:
                continue
            if k == 'Quantity':
                try:
                    self.df.at[i, k] = int(float(v or 0) or 0)
                except Exception:
                    self.df.at[i, k] = 0
            elif k in ['Fabric Meters','Meters per Unit','Fabric Meter Price','Sewing Cost','Accessories Cost','Extra Costs','Sale Price']:
                try:
                    self.df.at[i, k] = float(v or 0)
                except Exception:
                    self.df.at[i, k] = 0.0
            else:
                self.df.at[i, k] = v
        self.save()
        return 1

    def delete_item(self, code):
        code = str(code).strip()
        idx = self.df[self.df['Product Code'].astype(str) == code].index
        if not len(idx):
            return 0
        self.df = self.df.drop(index=idx)
        self.save()
        return 1

    def resolve_index(self, code_or_name: str):
        """Prefer code. If name is provided and exists multiple times, return None."""
        v = str(code_or_name or '').strip()
        if not v:
            return None
        i = self.find_index_by_code(v)
        if i is not None:
            return i

        # Name path (guard duplicates)
        matches = self.df[self.df['Product Name'].astype(str) == v]
        if len(matches.index) == 1:
            return matches.index[0]
        return None
    
  
# --------------------------- ORDER ITEMS HELPERS ------------------------
import json as _json

def parse_items_from_row(row):
    """Return list of items: [{'code':..., 'name':..., 'qty':...}, ...]"""
    try:
        raw = row.get('Items') if isinstance(row, dict) else row.get('Items', None)
    except Exception:
        raw = None
    items = []
    if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
        try:
            if isinstance(raw, str) and raw.strip():
                items = _json.loads(raw)
        except Exception:
            items = []
    if isinstance(items, list) and items:
        norm=[]
        for it in items:
            if not isinstance(it, dict):
                continue
            code=str(it.get('code','') or '').strip()
            name=str(it.get('name','') or '').strip()
            try:
                qty=int(float(it.get('qty',1) or 1))
            except Exception:
                qty=1
            if qty<=0: qty=1
            if not name and code:
                try:
                    inv = inventory.get_by_code(code)
                    if inv:
                        name=str(inv.get('Product Name','') or '').strip()
                except Exception:
                    pass
            if name:
                norm.append({'code':code,'name':name,'qty':qty})
        return norm
    # fallback: single product name
    try:
        name = row.get('Product Name')
    except Exception:
        name = None
    if name is not None and not (isinstance(name, float) and pd.isna(name)):
        name=str(name).strip()
        if name:
            return [{'code':'','name':name,'qty':1}]
    return []


inventory = InventoryStore(EXCEL_FILE)


@app.before_request
def _reload_data_if_changed():
    try:
        store.reload_if_changed()
    except Exception:
        pass
    try:
        inventory.reload_if_changed()
    except Exception:
        pass

def inventory_product_stats(product_code: str):
    """Compute per-product stats from movements + orders."""
    product_code = str(product_code).strip()
    inv_item = inventory.get_by_code(product_code)
    if not inv_item:
        return None

    try:
        inventory.movements.reload()
        mv = inventory.movements.filter_by_product_code(product_code).copy()
    except Exception:
        mv = pd.DataFrame(columns=InventoryMovementStore.COLS)

    mv['Delta'] = pd.to_numeric(mv.get('Delta'), errors='coerce').fillna(0).astype(int)
    mv['Movement Type'] = mv.get('Movement Type', '').astype(str).str.strip()

    total_produced = int(mv.loc[mv['Movement Type'].str.lower() == 'production', 'Delta'].clip(lower=0).sum() or 0)
    total_withdrawn = int((-mv.loc[mv['Movement Type'].str.lower() == 'withdraw', 'Delta'].clip(upper=0)).sum() or 0)
    total_returned_mov = int(mv.loc[mv['Movement Type'].str.lower() == 'return', 'Delta'].clip(lower=0).sum() or 0)

    last_withdraw_dt = ""
    try:
        wd = mv[(mv['Movement Type'].str.lower() == 'withdraw') & (mv['Delta'] < 0)]
        if not wd.empty:
            last_withdraw_dt = str(wd['DateTime'].iloc[-1])
    except Exception:
        pass

    delivered_pieces = 0
    returned_pieces = 0
    try:
        d = store.df.copy()
        d = d[d['Status'].isin([STATUS_DELIVERED, STATUS_RETURNED])]
        prod_name = str(inv_item.get('Product Name','') or '').strip()
        for _, r in d.iterrows():
            items = parse_items_from_row(r.to_dict())
            for it in items:
                it_code = str(it.get('code','') or '').strip()
                it_name = str(it.get('name','') or '').strip()
                if (it_code and it_code == product_code) or (not it_code and it_name == prod_name):
                    qty = int(it.get('qty', 1) or 1)
                    if r.get('Status') == STATUS_DELIVERED:
                        delivered_pieces += qty
                    elif r.get('Status') == STATUS_RETURNED:
                        returned_pieces += qty
    except Exception:
        pass

    total_done = delivered_pieces + returned_pieces
    delivered_pct = (delivered_pieces / total_done * 100) if total_done else 0.0
    returned_pct  = (returned_pieces  / total_done * 100) if total_done else 0.0

    return {
        "Product Code": product_code,
        "Product Name": inv_item.get("Product Name", ""),
        "Type": inv_item.get("Type", ""),
        "Current Quantity": int(pd.to_numeric(inv_item.get("Quantity"), errors="coerce") or 0),
        "Total Produced": total_produced,
        "Total Withdrawn": total_withdrawn,
        "Total Returned (to stock)": total_returned_mov,
        "Delivered Pieces": delivered_pieces,
        "Returned Pieces": returned_pieces,
        "Delivered %": round(delivered_pct, 2),
        "Returned %": round(returned_pct, 2),
        "Last Withdraw DateTime": last_withdraw_dt,
    }


# hook: adjust inventory when status transitions

def adjust_inventory_on_transition(row, old_status, new_status):
    try:
        items = parse_items_from_row(row)
        if not items:
            return

        # READY -> SHIPPING: decrement
        if old_status == STATUS_READY and new_status == STATUS_SHIPPING:
            for it in items:
                inventory.adjust_quantity(it['code'] or it['name'], -int(it.get('qty', 1) or 1), movement_type='Withdraw', ref=str(row.get('Transaction ID','') or ''), notes='READY->SHIPPING')

        # SHIPPING -> RETURNED: add back
        if old_status == STATUS_SHIPPING and new_status == STATUS_RETURNED:
            for it in items:
                inventory.adjust_quantity(it['code'] or it['name'], +int(it.get('qty', 1) or 1), movement_type='Return', ref=str(row.get('Transaction ID','') or ''), notes='SHIPPING->RETURNED')
    except Exception:
        pass


# --------------------------- INVENTORY TEMPLATES ------------------------

# --------------------------- EXTRA UPLOAD HELPERS ----------------------
UPLOAD_DIR = user_data_dir() / 'uploads'
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_IMG_EXT = {'.png', '.jpg', '.jpeg', '.webp'}

def _is_allowed_image(filename: str) -> bool:
    ext = Path(filename).suffix.lower()
    return ext in ALLOWED_IMG_EXT

def _save_image(file_storage):
    if not file_storage or not file_storage.filename:
        return ''
    if not _is_allowed_image(file_storage.filename):
        return ''
    fname = secure_filename(file_storage.filename)
    dst = UPLOAD_DIR / (datetime.now().strftime('%Y%m%d%H%M%S_') + fname)
    file_storage.save(dst)
    return str(dst)

# --------------------------- SEAMSTRESS TEMPLATE ------------------------



# ------------------------------- ROUTES --------------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        code = (request.form.get('code') or '').strip()
        if code == PASSCODE:
            session['auth'] = True
            # ✅ افتح الداش بورد (صفحة الإحصائيات) أولاً
            return redirect(url_for('home'))
        flash('رمز غير صحيح', 'err')
    return render_template_string(LOGIN_HTML)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# -------------------- Text Import / Processing Orders --------------------
_AR_NUMS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")

def _norm_digits(s: str) -> str:
    return (s or "").translate(_AR_NUMS).strip()

def _extract_phone_any(text: str) -> str:
    t = _norm_digits(text)
    cand = re.findall(r"(?:\+?964|964)?7\d{9}|07\d{9}", t)
    if not cand:
        return ""
    last = cand[-1].replace("+", "")
    if last.startswith("9647"):
        last = "0" + last[3:]
    if last.startswith("7") and len(last) == 10:
        last = "0" + last
    return last

def _norm_ar_numbers(text: str) -> str:
    """
    تحويل الأرقام العربية إلى إنكليزية
    """
    if not text:
        return text

    trans = str.maketrans(
        "٠١٢٣٤٥٦٧٨٩",
        "0123456789"
    )
    return text.translate(trans)


def _extract_price_any(text: str):
    t = _norm_digits(text)
    m = re.search(r"(?:حساب|الحساب|السعر)\s*[:：]?\s*([0-9]{1,6})", t)
    if m:
        p = int(m.group(1))
        if p < 1000: p *= 1000
        return p
    nums = re.findall(r"\b\d{2,6}\b", t)
    if nums:
        p = int(nums[-1])
        if p < 1000: p *= 1000
        return p
    return ""

def _new_txn() -> str:
    # 12 digits
    base = int(time.time() * 1000)
    for _ in range(50):
        txn = str(base + random.randint(0, 999)).rjust(12, "0")
        if not store.exists(txn):
            return txn
    return str(base).rjust(12, "0")


def _parse_orders_from_text(raw: str):
    """Parse pasted multi-order text (Telegram export OR simple blocks).

    ✅ Improvements:
    - If Telegram header lines exist (e.g. 'Name, [1/19/2026 10:56 PM]') we split ONLY by these headers
      (because inside a single message there are blank lines).
    - Ignore noise lines (stars/emoji-only) and sender headers.
    - More flexible extraction:
        * phone: last valid phone in the block (supports Arabic digits and 964/+964)
        * price: 'حساب..' OR '<number> الف' OR '<number> مع التوصيل'
        * address: prefers lines after (العنوان/عنواني/الموقع) or location-like lines
        * product: first 1-2 descriptive lines before address/phone/price
    - If extraction fails, keep FULL raw text as Notes, do NOT split into lines.
    """
    raw = (raw or "").replace("\r\n", "\n").strip()
    if not raw:
        return []

    lines0 = [ln.rstrip() for ln in raw.split("\n")]

    # Telegram header pattern: "Name, [1/19/2026 10:56 PM]" (any name)
    tg_hdr_re = re.compile(r"^.*?,\s*\[\s*\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}\s+(?:AM|PM)\s*\]\s*$")
    # Sometimes user pastes only timestamp part inside line
    tg_any_ts_re = re.compile(r"\[\s*\d{1,2}/\d{1,2}/\d{4}.*?\]")

    def _is_noise_line(s: str) -> bool:
        s = (s or "").strip()
        if not s:
            return False
        # stars / separators / emoji-only / repeated symbols
        if re.fullmatch(r"[⭐🌟✨⚡️🔥💥❗️‼️\-_=~•·\s]+", s):
            return True
        # single-letter or very short markers (ا/و/م)
        if s in {"ا", "و", "م", "ن", "ه"}:
            return True
        return False

    # detect telegram mode
    has_tg_headers = any(tg_hdr_re.match(ln.strip()) for ln in lines0)

    blocks = []
    current = []

    def flush():
        nonlocal current
        b = [x for x in current if x is not None]
        b = [x.strip() for x in b if x.strip() and not _is_noise_line(x)]
        if b:
            blocks.append(b)
        current = []

    for ln in lines0:
        s = ln.strip()
        if not s:
            # In Telegram mode: keep blank lines (do not split)
            if not has_tg_headers:
                # in simple mode, blank line separates orders
                flush()
            else:
                # keep blank as internal separator (later ignored)
                current.append("")
            continue

        if s == "---":
            flush()
            continue

        if tg_hdr_re.match(s):
            # Start new message block
            flush()
            continue

        # Sometimes Telegram header is not isolated in line; treat as separator
        if has_tg_headers and tg_any_ts_re.search(s) and "," in s:
            flush()
            continue

        # ignore pure noise lines
        if _is_noise_line(s):
            continue

        current.append(s)

    flush()

    parsed = []

    # helper: normalize Arabic digits
    def _norm_digits(s: str) -> str:
        try:
            return _norm_ar_numbers(s)
        except Exception:
            return s

    # patterns for address keywords
    addr_kw_re = re.compile(r"^(?:العنوان|عنواني|الموقع)\s*[:：]?\s*(.*)$")

    for b in blocks:
        if not b:
            continue

        # Build raw text
        raw_block = "\n".join([x for x in b if x is not None]).strip()

        # Extract phone and price from the whole block text
        joined = "\n".join(b)
        phone = _extract_phone_any(joined)  # already normalizes Arabic digits
        price = _extract_price_any(joined)

        # Clean lines for structure
        clean = []
        for ln in b:
            ln2 = ln.strip()
            if not ln2:
                continue
            # remove 'الرقم ..' etc prefixes
            ln2 = re.sub(r"^(?:الرقم|رقم الهاتف|رقمي)\s*[:：.]?\s*", "", ln2).strip()
            clean.append(ln2)

        # Identify explicit address lines
        addr_lines = []
        product_lines = []
        other_lines = []

        in_addr_mode = False
        for ln in clean:
            # mark address start if keyword present
            maddr = addr_kw_re.match(ln)
            if maddr:
                in_addr_mode = True
                rest = (maddr.group(1) or "").strip()
                if rest:
                    addr_lines.append(rest)
                continue

            # classify line types
            ln_norm = _norm_digits(ln)
            is_phone = bool(_extract_phone_any(ln_norm))
            is_priceish = bool(_extract_price_any(ln_norm)) or bool(re.search(r"(?:مع\s*التوصيل|الف)\b", ln_norm))
            is_only_number = bool(re.fullmatch(r"[0-9٠-٩]+", ln.strip()))
            looks_like_location = (len(ln) >= 6 and ("/" in ln or "-" in ln or "بغداد" in ln or "كربلاء" in ln or "ديالى" in ln or "بابل" in ln or "الموصل" in ln or "نينوى" in ln or "النجف" in ln or "ذي قار" in ln or "واسط" in ln or "الرمادي" in ln))

            if is_phone or is_priceish or is_only_number:
                in_addr_mode = False
                other_lines.append(ln)
                continue

            if in_addr_mode:
                addr_lines.append(ln)
                continue

            # if line looks like location and we already have product lines, treat as address
            if looks_like_location and product_lines:
                addr_lines.append(ln)
            else:
                # default: product/description first
                if not product_lines and looks_like_location:
                    # first line is location -> treat it as address
                    addr_lines.append(ln)
                else:
                    # treat as product/detail
                    product_lines.append(ln)

        # Decide product
        product = ""
        if product_lines:
            # take first 1-2 lines as product description
            product = " | ".join(product_lines[:2]).strip()

        # Decide address
        address = ""
        if addr_lines:
            address = " ".join(addr_lines).strip()
        else:
            # fallback: take any remaining non-phone non-price lines not used as product
            # (but avoid duplicating product)
            candidates = []
            for ln in clean:
                if ln in product_lines:
                    continue
                ln_norm = _norm_digits(ln)
                if _extract_phone_any(ln_norm):
                    continue
                if _extract_price_any(ln_norm) or re.search(r"(?:مع\s*التوصيل|الف)\b", ln_norm):
                    continue
                if len(ln) < 5:
                    continue
                candidates.append(ln)
            address = " ".join(candidates).strip()

        # If failed to extract, keep block but do not split
        parsed.append({
            "product": product.strip(),
            "phone": phone or "",
            "address": address.strip(),
            "price": price or "",
            "raw": raw_block
        })

    return parsed

@app.route("/orders/import_text", methods=["GET","POST"])
def orders_import_text():
    raw = (request.form.get("raw") if request.method=="POST" else "") or ""
    action = (request.form.get("action") if request.method=="POST" else "") or ""
    preview=[]
    if request.method=="POST":
        preview=_parse_orders_from_text(raw)
        if action=="save":
            saved=0
            for o in preview:
                txn=_new_txn()
                row={
                    "Transaction ID": txn,
                    "Time and Date": now_str(),
                    "Product Name": o["product"],
                    "Contact Numbers": o["phone"],
                    "Address": o["address"],
                    "Order Price": o["price"],
                    "Notes": o["raw"],
                    "Status": STATUS_PROCESSING,
                    "Status Updated At": now_str(),
                }
                ok, _msg = store.upsert_row(row)
                if ok:
                    saved += 1
            if saved:
                store.save()
            flash(f"تم حفظ {saved} طلب", "ok")
            return redirect(url_for("orders_processing"))
    return render_template_string(IMPORT_TEXT_HTML, raw=raw, preview=preview)

@app.route("/orders/processing")
@login_required
def orders_processing():
    try:
        store.reload_if_changed()
    except Exception:
        pass
    df = store.df.copy()
    if "Status" in df.columns:
        df = df[df["Status"].astype(str) == STATUS_PROCESSING]

    q = (request.args.get("q") or "").strip()
    product_filter = (request.args.get("product") or "").strip()

    if product_filter:
        df = df[df["Product Name"].astype(str) == product_filter]

    if q:
        ql = q.lower()
        mask = pd.Series(False, index=df.index)
        for c in ["Product Name", "Contact Numbers", "Address", "Notes", "Transaction ID"]:
            if c in df.columns:
                mask = mask | df[c].astype(str).str.lower().str.contains(ql, na=False)
        df = df[mask]

    # product counts
    product_counts=[]
    if "Product Name" in df.columns:
        vc = df["Product Name"].astype(str).replace("nan","").replace("None","")
        vc = vc[vc!=""].value_counts()
        product_counts = [(k, int(v)) for k,v in vc.items()]

    total = len(df)
    # sort by time if exists
    if "Time and Date" in df.columns:
        try:
            dts = pd.to_datetime(df["Time and Date"], errors="coerce")
            df = df.assign(_dt=dts).sort_values("_dt", ascending=False).drop(columns=["_dt"])
        except Exception:
            pass
    rows = df.fillna("").to_dict(orient="records")
    return render_template_string(PROCESSING_ORDERS_HTML, rows=rows, total=total, q=q, product_filter=product_filter, product_counts=product_counts, STATUS_SHIPPING=STATUS_SHIPPING)

@app.route("/orders/processing_suggest")
@login_required
def orders_processing_suggest():
    q = (request.args.get("q") or "").strip()
    if not q:
        return {"items": []}
    try:
        store.reload_if_changed()
    except Exception:
        pass
    df = store.df.copy()
    if "Status" in df.columns:
        df = df[df["Status"].astype(str) == STATUS_PROCESSING]
    ql=q.lower()
    items=[]
    for _, r in df.fillna("").iterrows():
        blob=" ".join([str(r.get(c,"")) for c in ["Product Name","Contact Numbers","Address","Notes"]]).lower()
        if ql in blob:
            s=" — ".join([x for x in [str(r.get("Contact Numbers","")).strip(), str(r.get("Product Name","")).strip(), str(r.get("Address","")).strip()] if x])
            if s and s not in items:
                items.append(s)
        if len(items)>=12:
            break
    return {"items": items}

@app.route("/orders/processing/to_shipping", methods=["POST"])
@login_required
def processing_to_shipping():
    txn=(request.form.get("txn") or "").strip()
    ok, msg = store.update_status(txn, STATUS_SHIPPING)
    if ok:
        store.save()
        flash("تم تحويل الطلب إلى قيد التوصيل", "ok")
    else:
        flash(msg, "err")
    return redirect(url_for("orders_processing"))

@app.route("/orders/processing/delete", methods=["POST"])
@login_required
def processing_delete():
    txn=(request.form.get("txn") or "").strip()
    try:
        st = str(store.df.at[txn, "Status"]) if "Status" in store.df.columns and txn in store.df.index else ""
    except Exception:
        st=""
    if st == STATUS_SHIPPING:
        flash("لا يمكن حذف طلب تم تحويله إلى قيد التوصيل", "err")
        return redirect(url_for("orders_processing"))
    if txn in store.df.index:
        store.df = store.df.drop(index=txn)
        store.save()
        flash("تم حذف الطلب", "ok")
    else:
        flash("الطلب غير موجود", "err")
    return redirect(url_for("orders_processing"))

@app.route("/orders/processing/delete_all", methods=["POST"])
@login_required
def processing_delete_all():
    try:
        df = store.df.copy()
        if "Status" in df.columns:
            store.df = df[df["Status"].astype(str) != STATUS_PROCESSING]
            store.save()
            flash("تم حذف كل طلبات قيد المعالجة", "ok")
    except Exception as e:
        flash(f"تعذر الحذف: {e}", "err")
    return redirect(url_for("orders_processing"))

@app.route("/orders/processing/edit/<txn>", methods=["GET","POST"])
@login_required
def processing_edit(txn):
    txn=str(txn).strip()
    if request.method=="POST":
        prod=(request.form.get("product") or "").strip()
        phone=(request.form.get("phone") or "").strip()
        addr=(request.form.get("address") or "").strip()
        price=(request.form.get("price") or "").strip()
        notes=(request.form.get("notes") or "").strip()
        if txn in store.df.index:
            if "Product Name" in store.df.columns: store.df.at[txn,"Product Name"]=prod
            if "Contact Numbers" in store.df.columns: store.df.at[txn,"Contact Numbers"]=phone
            if "Address" in store.df.columns: store.df.at[txn,"Address"]=addr
            if "Order Price" in store.df.columns: store.df.at[txn,"Order Price"]=price
            if "Notes" in store.df.columns: store.df.at[txn,"Notes"]=notes
            store.save()
            flash("تم تحديث الطلب", "ok")
        return redirect(url_for("orders_processing"))
    if txn not in store.df.index:
        flash("الطلب غير موجود", "err")
        return redirect(url_for("orders_processing"))
    r = store.df.loc[txn].fillna("")
    return render_template_string(r"""
    {% extends 'base.html' %}
    {% block content %}
    <h5 class="mb-3">تحديث الطلب</h5>
    <form method="post" class="card border-0 shadow-sm rounded-4">
      <div class="card-body">
        <div class="row g-2">
          <div class="col-md-6">
            <label class="form-label">المنتج</label>
            <input name="product" class="form-control" value="{{ r.get('Product Name','') }}">
          </div>
          <div class="col-md-6">
            <label class="form-label">الموبايل</label>
            <input name="phone" class="form-control" value="{{ r.get('Contact Numbers','') }}">
          </div>
          <div class="col-md-8">
            <label class="form-label">العنوان</label>
            <input name="address" class="form-control" value="{{ r.get('Address','') }}">
          </div>
          <div class="col-md-4">
            <label class="form-label">السعر</label>
            <input name="price" class="form-control" value="{{ r.get('Order Price','') }}">
          </div>
          <div class="col-12">
            <label class="form-label">النص/ملاحظات</label>
            <textarea name="notes" rows="4" class="form-control">{{ r.get('Notes','') }}</textarea>
          </div>
        </div>
        <div class="d-flex gap-2 mt-3">
          <button class="btn btn-primary">حفظ</button>
          <a class="btn btn-outline-secondary" href="{{ url_for('orders_processing') }}">إلغاء</a>
        </div>
      </div>
    </form>
    {% endblock %}
    """, r=r)


@app.route('/')
@login_required
def home():
    q = (request.args.get('q') or '').strip()
    prod = (request.args.get('product') or '').strip()
    page = (request.args.get('page') or '').strip()
    status = (request.args.get('status') or '').strip()
    dfrom = request.args.get('from')
    dto = request.args.get('to')

    d = store.df.copy()
    # بحث نصي
    if q:
        mask = pd.Series(False, index=d.index)
        for c in ["Transaction ID", "Product Name", "Page Name",
                  "Contact Numbers", "Address", "Notes", "Return Reason"]:
            if c in d.columns:
                mask = mask | d[c].astype(str).str.contains(q, na=False)
        d = d[mask]

    # فلتر المنتج
    if prod and "Product Name" in d.columns:
        d = d[d["Product Name"].astype(str) == prod]

    # فلتر الصفحة
    if page and "Page Name" in d.columns:
        d = d[d["Page Name"].astype(str) == page]

    # فلتر الحالة
    if status and "Status" in d.columns:
        d = d[d["Status"].astype(str) == status]

    # فلتر التاريخ + ترتيب
    if "Time and Date" in d.columns:
        d["Time and Date"] = pd.to_datetime(d["Time and Date"], errors="coerce")
        if dfrom:
            try:
                start = datetime.strptime(dfrom, "%Y-%m-%d")
                d = d[d["Time and Date"] >= start]
            except Exception:
                pass
        if dto:
            try:
                end = datetime.strptime(dto, "%Y-%m-%d")
                d = d[d["Time and Date"] <= end]
            except Exception:
                pass
        d = d.sort_values("Time and Date", ascending=False, na_position="last")
        d["Time and Date"] = d["Time and Date"].dt.strftime("%Y-%m-%d %H:%M:%S")

    rows = d.fillna("").to_dict(orient="records")

    # قوائم الفلاتر
    all_products = []
    all_pages = []
    if "Product Name" in store.df.columns:
        all_products = sorted(list({str(x) for x in store.df["Product Name"].dropna().unique()}))
    if "Page Name" in store.df.columns:
        all_pages = sorted(list({str(x) for x in store.df["Page Name"].dropna().unique()}))

    # ملخص الإحصائيات (للكروت العلوية) مع النسب
    summary = store.stats_global(d)

    all_statuses = [STATUS_READY, STATUS_SHIPPING, STATUS_DELIVERED, STATUS_RETURNED]

    return render_template_string(
        HOME_HTML,
        columns=BASE_COLUMNS,
        rows=rows,
        q=q,
        all_products=all_products,
        all_pages=all_pages,
        sel_product=prod,
        sel_page=page,
        sel_status=status,
        all_statuses=all_statuses,
        dfrom=dfrom,
        dto=dto,
        summary=summary,
        status_delivered=STATUS_DELIVERED,
        status_shipping=STATUS_SHIPPING,
        status_ready=STATUS_READY,
        status_returned=STATUS_RETURNED,
    )


@app.route('/mark_returned', methods=['POST'])
@login_required
def mark_returned():
    txn = (request.form.get('txn') or '').strip()
    ok, msg = store.update_status(txn, STATUS_RETURNED)
    if ok:
        store.save(); flash('تم تحديث الحالة إلى راجع', 'ok')
    else:
        flash(msg, 'err')
    return redirect(url_for('home'))



@app.route('/upload_pdf', methods=['POST'])
@login_required
def upload_pdf():
    file = request.files.get('pdf')
    if not file:
        flash('يرجى اختيار ملف PDF', 'err')
        return redirect(url_for('home'))

    path = Path(UPLOAD_DIR) / f"import_{int(datetime.now().timestamp())}.pdf"
    file.save(path)

    client_count = {}
    added, updated = 0, 0
    page_errors = []

    try:
        with pdfplumber.open(str(path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                try:
                    text = page.extract_text() or ""
                    txn, phone_str, order_price, address = extract_from_text(text)
                    if not txn:
                        continue

                    main_phone = None
                    if phone_str:
                        main_phone = phone_str.split(',')[0].strip()

                    client_count[main_phone] = client_count.get(main_phone, 0) + 1

                    page_data = {
                        "Product Name": pd.NA,
                        "Transaction ID": str(txn),
                        "Time and Date": now_str(),
                        "Contact Numbers": phone_str,
                        "Address": address,
                        "Order Price": order_price,
                        "Status": STATUS_READY,
                        "Return Reason": "لا يوجد",
                        "Notes": None,
                        "Client Orders Count": client_count.get(main_phone, 1) if main_phone else pd.NA,
                    }

                    ok, msg = store.upsert_row(page_data)
                    if ok and msg == "تمت الإضافة":
                        added += 1
                    elif ok and msg == "تم التحديث":
                        updated += 1

                except Exception as pe:
                    page_errors.append((page_num, f"{type(pe).__name__}: {pe}"))

        store.save()

        info = f"تمت معالجة PDF. المضاف: {added} | المحدّث: {updated}"
        if page_errors:
            info += f" | تعذّر قراءة {len(page_errors)} صفحة"
        flash(info, 'ok')

    except Exception as e:
        _fatal_box('فشل استيراد PDF', e)
        flash('فشل استيراد PDF', 'err')

    return redirect(url_for('home'))

@app.route('/upload_invoice', methods=['POST'])
@login_required
def upload_invoice():
    file = request.files.get('pdf')
    if not file:
        flash('يرجى اختيار ملف PDF', 'err'); return redirect(url_for('home'))
    path = Path(UPLOAD_DIR) / f"invoice_{int(datetime.now().timestamp())}.pdf"
    file.save(path)

    updated_rows, skipped_rows = [], []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                text = normalize_digits(page.extract_text() or "")
                for ln in text.split("\n"):
                    ln = ln.strip()
                    if not ln:
                        continue

                    txn = None
                    price_val = None

                    # المحاولة القديمة (لو كان رقم الشحنة يجي أول السطر متبوعاً بالسعر)
                    m = re.search(r'(\d{6,})\s+((?:\d{1,3}(?:,\d{3})+|\d{4,9}))', ln)
                    if m:
                        txn = m.group(1).strip()
                        price_val = to_int(m.group(2))

                    # لو ما زبطت، نستخدم الطريقة الجديدة المناسبة لملف الشركة
                    if not txn or price_val is None:
                        # كل الأسعار اللي بشكل 25,000 / 42,000 / 6,000 ...
                        price_matches = re.findall(r'(\d{1,3}(?:,\d{3})+)', ln)
                        # كل الأرقام الطويلة (6 أرقام أو أكثر) = غالباً هاتف + رقم شحنة
                        long_nums = re.findall(r'(\d{6,})', ln)

                        if price_matches and long_nums:
                            # رقم الشحنة = آخر رقم طويل في السطر (الأول يكون موبايل)
                            txn = long_nums[-1].strip()

                            # نأخذ أكبر مبلغ كقيمة الشحنة (أكبر من أجرة التوصيل والصافي)
                            max_price = None
                            for p in price_matches:
                                v = to_int(p)
                                if v is not None and (max_price is None or v > max_price):
                                    max_price = v
                            price_val = max_price

                    # لو ما طلع معنا شيء مفيد نكمل للسطر التالي
                    if not txn or price_val is None:
                        continue

                    # تأكيد أن المبلغ ينتهي بـ 000 مثل 25,000 ... الخ
                    if not str(price_val).endswith("000"):
                        continue

                    # نحدّث حالة الطلب لو موجود والـ Order Price مطابق
                    if store.exists(txn):
                        exist = store.get_row(txn)
                        exist_price = pd.to_numeric(exist.get("Order Price"), errors="coerce")
                        if pd.notna(exist_price) and int(exist_price) == int(price_val):
                            store.update_status(txn, STATUS_DELIVERED)
                            updated_rows.append((txn, price_val, "OK"))
                        else:
                            skipped_rows.append((txn, price_val, f"سعر مختلف (المسجل: {exist_price})"))
                    else:
                        skipped_rows.append((txn, price_val, "الشحنة غير موجودة"))

        store.save()
        flash(f"تم التحديث: {len(updated_rows)} | لم يتم: {len(skipped_rows)}", 'ok')
    except Exception as e:
        _fatal_box('فشل رفع الفاتورة', e)
        flash('فشل رفع الفاتورة', 'err')
    return redirect(url_for('home'))


@app.route('/dedupe')
@login_required
def dedupe():
    removed = store.drop_duplicates_keep_last()
    store.save()
    flash(f"تم حذف {removed} صف مكرر.", 'ok')
    return redirect(url_for('home'))




@app.route('/delete-ready-all')
@login_required
def delete_ready_all():
    """حذف جميع الطلبات التي هي قيد التجهيز (STATUS_READY)"""
    try:
        before = len(store.df)
        if "Status" in store.df.columns:
            store.df = store.df[store.df["Status"].astype(str) != str(STATUS_READY)]
        after = len(store.df)
        deleted = before - after
        store.save()
        flash(f"تم حذف {deleted} طلب ({STATUS_READY})", 'ok')
    except Exception as e:
        _fatal_box('فشل حذف طلبات قيد التجهيز', e)
        flash('فشل حذف طلبات قيد التجهيز', 'err')
    return redirect(url_for('home'))

@app.route('/delete/<txn>')
@login_required
def delete(txn):
    deleted = store.drop_by_txn(txn)
    if deleted:
        store.save(); flash('تم الحذف', 'ok')
    else:
        flash('الشحنة غير موجودة', 'err')
    return redirect(url_for('home'))


@app.route('/edit/<txn>', methods=['GET', 'POST'])
@login_required
def edit(txn):
    if not store.exists(txn):
        abort(404)
    if request.method == 'POST':
        new_vals = {c: request.form.get(c) for c in BASE_COLUMNS}
        if 'Order Price' in new_vals:
            new_vals['Order Price'] = pd.to_numeric(new_vals['Order Price'], errors='coerce')
        ok, msg = store.upsert_row(new_vals)
        if ok:
            store.save(); flash('تم التعديل', 'ok'); return redirect(url_for('home'))
        flash(msg, 'err')
    row = store.get_row(txn).fillna("").to_dict()
    return render_template_string(EDIT_HTML, txn=txn, columns=BASE_COLUMNS, row=row)


@app.route('/move-to-shipping', methods=['GET', 'POST'])
@login_required
def move_to_shipping():
    session.setdefault('shipping_items', [])
    session.setdefault('shipping_products', [])  # [{'code','name','qty'}]
    # إزالة التكرار من القائمة
    session['shipping_items'] = list(dict.fromkeys(session['shipping_items']))
    headers = ['Transaction ID', 'Page', 'Product', 'Status']
    title = 'تحديث الحالة إلى قيد التوصيل'
    shipping_products = session.get('shipping_products', [])
    product_name = ''  # deprecated single-name
    page_name = session.get('page_name', '')

    PAGES = ['فاتنة', 'لمسة حرير', 'براعم', 'أنيقا', 'خيوط']
    # جميع المنتجات من المخزن (إدارة المنتجات)
    INVENTORY_OPTIONS = []
    try:
        if not inventory.df.empty:
            for _, r in inventory.df[['Product Code','Product Name']].dropna(subset=['Product Name']).iterrows():
                INVENTORY_OPTIONS.append({'code': str(r['Product Code']), 'name': str(r['Product Name'])})
    except Exception:
        pass

    if request.method == 'POST':
        # إضافة منتج إلى قائمة المنتجات للمجموعة (بحث بالاسم أو الكود)
        if request.form.get('add_product'):
            key = (request.form.get('product_key') or '').strip()
            qty = (request.form.get('qty') or '1').strip()
            try:
                qty_i = int(float(qty))
            except Exception:
                qty_i = 1
            if qty_i <= 0:
                qty_i = 1

            code_key = ''
            name_key = ''
            # الصيغة المتوقعة: "CODE | NAME"
            if '|' in key:
                parts = [p.strip() for p in key.split('|', 1)]
                code_key = parts[0]
                name_key = parts[1] if len(parts) > 1 else ''
            else:
                # قد يكون كود فقط أو اسم فقط
                code_key = key if key.isdigit() else ''
                name_key = key if not key.isdigit() else ''

            # حاول تجيب من المخزن
            try:
                if code_key:
                    inv = inventory.get_by_code(code_key)
                    if inv:
                        name_key = str(inv.get('Product Name','') or '').strip()
                if not name_key and name_key:
                    pass
            except Exception:
                pass

            if not name_key:
                flash('يرجى اختيار منتج صحيح من المخزن', 'err')
                return redirect(url_for('move_to_shipping'))

            # دمج إذا موجود مسبقاً
            sp = session.get('shipping_products', [])
            found = False
            for it in sp:
                if str(it.get('name','')) == name_key or (code_key and str(it.get('code','')) == code_key):
                    it['qty'] = int(it.get('qty', 1) or 1) + qty_i
                    found = True
                    break
            if not found:
                sp.append({'code': code_key, 'name': name_key, 'qty': qty_i})
            session['shipping_products'] = sp
            flash('تمت إضافة المنتج للمجموعة', 'ok')
            return redirect(url_for('move_to_shipping'))

        if request.form.get('remove_product'):
            rm = (request.form.get('rm_name') or '').strip()
            sp = [it for it in session.get('shipping_products', []) if str(it.get('name','')) != rm]
            session['shipping_products'] = sp
            flash('تم حذف المنتج من المجموعة', 'ok')
            return redirect(url_for('move_to_shipping'))

        if request.form.get('clear_products'):
            session['shipping_products'] = []
            flash('تم تفريغ قائمة المنتجات', 'ok')
            return redirect(url_for('move_to_shipping'))

        # 1) حفظ اختيار اسم المنتج والبيج للمجموعة
        if request.form.get('apply_name'):
            name = (request.form.get('product_name') or '').strip()
            pg = (request.form.get('page_name') or '').strip()
            session['product_name'] = name
            session['page_name'] = pg

            # حفظ قائمة المنتجات للمجموعة
            session['shipping_products'] = session.get('shipping_products', [])

            count = 0
            if session['shipping_items']:
                for txn in session['shipping_items']:
                    if not store.exists(txn):
                        continue

                    row = store.get_row(txn)
                    old_name = row.get('Product Name')
                    status = row.get('Status')

                    # نحول NaN إلى نص عادي
                    old_name_val = ''
                    if old_name is not None and not pd.isna(old_name):
                        old_name_val = str(old_name).strip()

                    # ✅ قائمة المنتجات الجديدة (أكثر من منتج + كمية لكل منتج)
                    new_items = session.get('shipping_products', []) or []
                    # تأكد من شكل العناصر
                    norm_new = []
                    for it in new_items:
                        if not isinstance(it, dict):
                            continue
                        code = str(it.get('code','') or '').strip()
                        nm = str(it.get('name','') or '').strip()
                        try:
                            qty = int(float(it.get('qty', 1) or 1))
                        except Exception:
                            qty = 1
                        if qty <= 0:
                            qty = 1
                        if not nm and code:
                            try:
                                inv = inventory.get_by_code(code)
                                if inv:
                                    nm = str(inv.get('Product Name','') or '').strip()
                            except Exception:
                                pass
                        if nm:
                            norm_new.append({'code': code, 'name': nm, 'qty': qty})

                    # نستخدم Product Name كعرض مختصر (للتوافق مع الصفحات القديمة)
                    if len(norm_new) == 1:
                        new_name_val = norm_new[0]['name']
                    elif len(norm_new) > 1:
                        new_name_val = 'أكثر من منتج'
                    else:
                        new_name_val = ''

                    # ✅ تصحيح المخزون عند تطبيق المنتجات على شحنة بحالة "قيد التوصيل"
                    # الفكرة: نحسب فرق (القديم - الجديد) لكل منتج ونطبقه على المخزن
                    if status == STATUS_SHIPPING and norm_new:
                        try:
                            row_dict = row.fillna('').to_dict() if hasattr(row, 'fillna') else (row.to_dict() if hasattr(row, 'to_dict') else dict(row))
                            old_items = parse_items_from_row(row_dict)

                            def _map_items(items):
                                m = {}
                                for it in items or []:
                                    if not isinstance(it, dict):
                                        continue
                                    code = str(it.get('code','') or '').strip()
                                    nm = str(it.get('name','') or '').strip()
                                    key = ('code:' + code) if code else ('name:' + nm)
                                    if not key or key == 'name:':
                                        continue
                                    try:
                                        q = int(float(it.get('qty', 1) or 1))
                                    except Exception:
                                        q = 1
                                    if q <= 0:
                                        q = 1
                                    m[key] = m.get(key, 0) + q
                                return m

                            old_map = _map_items(old_items)
                            new_map = _map_items(norm_new)

                            all_keys = set(old_map.keys()) | set(new_map.keys())
                            for k in all_keys:
                                old_q = old_map.get(k, 0)
                                new_q = new_map.get(k, 0)
                                delta = old_q - new_q  # + يرجع للمخزن، - ينقص من المخزن
                                if delta == 0:
                                    continue

                                # resolve name
                                if k.startswith('code:'):
                                    code = k.split(':',1)[1]
                                    nm = ''
                                    try:
                                        inv = inventory.get_by_code(code)
                                        if inv:
                                            nm = str(inv.get('Product Name','') or '').strip()
                                    except Exception:
                                        nm = ''
                                    if not nm:
                                        # حاول من القوائم نفسها
                                        for it in (old_items or []) + (norm_new or []):
                                            if str(it.get('code','') or '').strip() == code:
                                                nm = str(it.get('name','') or '').strip()
                                                break
                                    if nm:
                                        inventory.adjust_quantity(nm, delta)
                                else:
                                    nm = k.split(':',1)[1]
                                    if nm:
                                        inventory.adjust_quantity(nm, delta)
                        except Exception:
                            pass

                    # تحديث المنتجات واسم البيج في جدول الطلبات
                    sp = session.get('shipping_products', [])
                    if sp:
                        try:
                            store.df.at[txn, 'Items'] = _json.dumps(sp, ensure_ascii=False)
                        except Exception:
                            store.df.at[txn, 'Items'] = ''
                        # Product Name للعرض فقط (مجمّع)
                        store.df.at[txn, 'Product Name'] = ' + '.join([str(x.get('name','')) for x in sp if x.get('name')])
                    elif new_name_val:
                        store.df.at[txn, 'Product Name'] = new_name_val
                    if pg:
                        store.df.at[txn, 'Page Name'] = pg

                    count += 1

                store.save()
                flash(f'تم تطبيق الاسم/البيج على {count} شحنة', 'ok')

            # ✅ بعد ما خلصنا هذه المجموعة نفرّغ الجدول في الأسفل
            session['shipping_items'] = []
            return redirect(url_for('move_to_shipping'))

        # 2) (زر تطبيق الكل – حاليًا فقط رسالة، التحديث يتم عند إضافة كل شحنة)
        if request.form.get('apply_all'):
            flash('تم تحديث الحالات الحالية إلى قيد التوصيل', 'ok')
            return redirect(url_for('move_to_shipping'))

        # 3) إضافة شحنة جديدة إلى قائمة قيد التوصيل
        txn = (request.form.get('txn') or '').strip()
        if not txn:
            flash('يرجى إدخال رقم الشحنة', 'err')
            return redirect(url_for('move_to_shipping'))

        # لو الشحنة موجودة في الجدول، نضبط اسم المنتج والبيج قبل تغيير الحالة
        if store.exists(txn):

            sp = session.get('shipping_products', [])
            if sp:
                try:
                    store.df.at[txn, 'Items'] = _json.dumps(sp, ensure_ascii=False)
                except Exception:
                    store.df.at[txn, 'Items'] = ''
                store.df.at[txn, 'Product Name'] = ' + '.join([str(x.get('name','')) for x in sp if x.get('name')])
            if product_name:
                store.df.at[txn, 'Product Name'] = product_name
            if page_name:
                store.df.at[txn, 'Page Name'] = page_name

            # ✅ إذا اخترت منتجات (أكثر من منتج) نحفظها في Items قبل تغيير الحالة
            sp = session.get('shipping_products', []) or []
            if sp:
                try:
                    store.df.at[txn, 'Items'] = _json.dumps(sp, ensure_ascii=False)
                except Exception:
                    store.df.at[txn, 'Items'] = ''
                # Product Name للعرض فقط
                store.df.at[txn, 'Product Name'] = ' + '.join([str(x.get('name','')) for x in sp if x.get('name')]) or store.df.at[txn, 'Product Name']
            # الآن نغيّر الحالة (الهُوك الخاص بالمخزن سيستخدم اسم المنتج الصحيح إن وجد)
            ok, info = store.update_status(txn, STATUS_SHIPPING)
            if ok:
                if txn not in session['shipping_items']:
                    session['shipping_items'].append(txn)
                store.save()
            else:
                flash(info, 'err')
        else:
            flash('الشحنة غير موجودة في النظام', 'err')

        return redirect(url_for('move_to_shipping'))

    # دالة تجهيز صف للجدول
    def row(txn):
        p = store.get_row(txn) if store.exists(txn) else None
        if p is not None:
            try:
                page_val = p.get('Page Name', '')
                prod_val = p.get('Product Name', '')
            except Exception:
                page_val = p['Page Name'] if isinstance(p, dict) and 'Page Name' in p else ''
                prod_val = p['Product Name'] if isinstance(p, dict) and 'Product Name' in p else ''
        else:
            page_val, prod_val = '', ''
        return {
            "Transaction ID": txn,
            "Page": page_val,
            "Product": prod_val,
            "Status": STATUS_SHIPPING
        }

        # إحصائيات اليوم لعرضها في البوكسات (عدد طلبات اليوم / في قيد التوصيل / قيد التجهيز)
    today_stats = None
    try:
        d_all = store.df.copy()

        # 📅 طلبات اليوم (حسب تاريخ الإنشاء)
        total_today = 0
        if 'Time and Date' in d_all.columns:
            d_all['Time and Date'] = pd.to_datetime(d_all['Time and Date'], errors='coerce')
            today = date.today()
            total_today = int((d_all['Time and Date'].dt.date == today).sum())

        # 🚚 اليوم في قيد التوصيل (حسب تاريخ تحويل الحالة إلى قيد التوصيل)
        shipping_today = 0
        if 'Shipping At' in d_all.columns:
            d_all['Shipping At'] = pd.to_datetime(d_all['Shipping At'], errors='coerce')
            today = date.today()
            shipping_today = int(((d_all['Status'] == STATUS_SHIPPING) & (d_all['Shipping At'].dt.date == today)).sum())

        # 🧵 اليوم قيد التجهيز (حسب تاريخ الإنشاء)
        ready_today = 0
        if 'Time and Date' in d_all.columns:
            today = date.today()
            ready_today = int(((d_all['Status'] == STATUS_READY) & (d_all['Time and Date'].dt.date == today)).sum())

        today_stats = {
            "total_today": int(total_today),
            "shipping_today": int(shipping_today),
            "ready_today": int(ready_today),
        }
    except Exception:
        today_stats = None

    items = [row(t) for t in session['shipping_items']]
    return render_template_string(
        BULK_HTML,
        title=title,
        headers=headers,
        items=items,
        action_label=f"تطبيق الكل -> {STATUS_SHIPPING}",
        product_name=product_name,
        PAGES=PAGES,
        INVENTORY_OPTIONS=INVENTORY_OPTIONS,
        shipping_products=session.get('shipping_products', []),
        page_name=page_name,
        today_stats=today_stats,
    )



@app.route('/returns-bulk', methods=['GET', 'POST'])
@login_required
def returns_bulk():
    session.setdefault('returns_items', [])
    session['returns_items'] = list(dict.fromkeys(session['returns_items']))
    headers = ['Transaction ID', 'Status', 'Reason']
    title = 'إدارة الطلبات الراجعة'
    if request.method == 'POST':
        if request.form.get('apply_all'):
            for txn in session['returns_items']:
                if store.exists(txn):
                    store.update_status(txn, STATUS_RETURNED)
            store.save(); flash('تم تحديث الحالات', 'ok')
            return redirect(url_for('returns_bulk'))
        txn = (request.form.get('txn') or '').strip()
        if txn and txn not in session['returns_items']:
            session['returns_items'].append(txn)
        return redirect(url_for('returns_bulk'))
    items = [{"Transaction ID": t, "Status": STATUS_RETURNED, "Reason": ""} for t in session['returns_items']]
    return render_template_string(BULK_HTML, title=title, headers=headers, items=items,
                                  action_label=f"تطبيق الكل -> {STATUS_RETURNED}", product_name=None)


@app.route('/delivered-bulk', methods=['GET', 'POST'])
@login_required
def delivered_bulk():
    session.setdefault('delivered_items', [])
    session['delivered_items'] = list(dict.fromkeys(session['delivered_items']))
    headers = ['Transaction ID', 'Order Price', 'Status']
    title = 'إدارة الطلبات التي تم تسليمها'
    if request.method == 'POST':
        if request.form.get('apply_all'):
            for txn in session['delivered_items']:
                if store.exists(txn):
                    store.update_status(txn, STATUS_DELIVERED)
            store.save(); flash('تم تحديث الحالات', 'ok')
            return redirect(url_for('delivered_bulk'))
        txn = (request.form.get('txn') or '').strip()
        if txn and txn not in session['delivered_items']:
            session['delivered_items'].append(txn)
        return redirect(url_for('delivered_bulk'))
    def row(txn):
        pr = ''
        if store.exists(txn):
            pr = store.get_row(txn).get('Order Price', '')
        return {"Transaction ID": txn, "Order Price": pr, "Status": STATUS_DELIVERED}
    items = [row(t) for t in session['delivered_items']]
    return render_template_string(BULK_HTML, title=title, headers=headers, items=items,
                                  action_label=f"تطبيق الكل -> {STATUS_DELIVERED}", product_name=None)


@app.route('/pending')
@login_required
def pending():
    dfrom = request.args.get('from')
    dto = request.args.get('to')
    d = store.df.copy()
    d = d[d['Status'] == STATUS_SHIPPING]
    d['Time and Date'] = pd.to_datetime(d['Time and Date'], errors='coerce')
    if dfrom:
        start = datetime.strptime(dfrom, '%Y-%m-%d')
        d = d[d['Time and Date'] >= start]
    if dto:
        end = datetime.strptime(dto, '%Y-%m-%d')
        d = d[d['Time and Date'] <= end]
    d = d.sort_values('Time and Date', ascending=False)
    out = []
    for _, r in d.iterrows():
        ts = r['Time and Date']
        ts = ts.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(ts) else ''
        out.append({'Transaction ID': r['Transaction ID'], 'Time and Date': ts,
                    'Order Price': r['Order Price'], 'Status': r['Status']})
    return render_template_string(PENDING_HTML, rows=out, dfrom=dfrom, dto=dto)


@app.route('/stats', methods=['GET', 'POST'])
@login_required
def stats():
    """
    لوحة الإحصائيات / الداش بورد
    - محمية برمز ثانوي 998144
    - تعتمد على:
        store.df
        store.stats_global(df)
        store.stats_by_product_price(df)
        store.daily_trend(df)
        inventory.df
        STATUS_READY / STATUS_SHIPPING / STATUS_DELIVERED / STATUS_RETURNED
    """

    # 🔐 بوابة رمز الإحصائيات
    if not session.get('stats_auth'):
        if request.method == 'POST' and (request.form.get('code') or '').strip() == '998144':
            session['stats_auth'] = True
        else:
            return render_template_string("""
            {% extends 'base.html' %}
            {% block content %}
            <div class="row justify-content-center">
              <div class="col-md-5">
                <div class="card p-4 mt-4">
                  <h6 class="mb-3">رمز دخول لوحة الإحصائيات</h6>
                  <form method="post">
                    <input name="code" type="password" class="form-control mb-3" placeholder="••••••" autofocus>
                    <button class="btn btn-primary w-100">دخول</button>
                  </form>
                </div>
              </div>
            </div>
            {% endblock %}
            """)

    # ─────────────────────────────
    # 1) قراءة فلاتر التاريخ والصفحة
    # ─────────────────────────────
    dfrom = (request.args.get('from') or '').strip()
    dto = (request.args.get('to') or '').strip()
    sel_page = (request.args.get('page') or '').strip()

    # ─────────────────────────────
    # 1b) تكاليف إضافية لحساب صافي الربح
    # ─────────────────────────────
    try:
        shipping_fee = int(float((request.args.get('shipping_fee') or session.get('shipping_fee', 4000) or 4000)))
    except Exception:
        shipping_fee = 4000
    try:
        ads_cost = float((request.args.get('ads_cost') or session.get('ads_cost', 0) or 0))
    except Exception:
        ads_cost = 0.0
    try:
        other_cost = float((request.args.get('other_cost') or session.get('other_cost', 0) or 0))
    except Exception:
        other_cost = 0.0

    session['shipping_fee'] = shipping_fee
    session['ads_cost'] = ads_cost
    session['other_cost'] = other_cost

    d = store.df.copy()

    # تحويل العمود لتاريخ (إن وجد)
    if 'Time and Date' in d.columns:
        d['Time and Date'] = pd.to_datetime(d['Time and Date'], errors='coerce')

        if dfrom:
            try:
                start = datetime.strptime(dfrom, '%Y-%m-%d')
                d = d[d['Time and Date'] >= start]
            except Exception:
                pass

        if dto:
            try:
                end = datetime.strptime(dto, '%Y-%m-%d')
                d = d[d['Time and Date'] <= end]
            except Exception:
                pass

    # فلتر حسب اسم البيج
    if sel_page and 'Page Name' in d.columns:
        d = d[d['Page Name'].astype(str) == sel_page]

    # ─────────────────────────────
    # 2) ملخّص عام (global summary)
    # ─────────────────────────────
    # ⚠️ ملاحظة مهمة:
    # d هنا مُفلتر حسب "Time and Date" (تاريخ إنشاء الطلب)، وهذا مناسب للمجموع المالي والتحليل العام.
    # لكن عدّ حالات (قيد التوصيل/تم التوصيل/راجع) لازم يعتمد على تاريخ تغيير الحالة، وليس تاريخ الإنشاء.
    base_all = store.df.copy()
    if sel_page and 'Page Name' in base_all.columns:
        base_all = base_all[base_all['Page Name'].astype(str) == sel_page]

    def _parse_range(_from, _to):
        start_dt = None
        end_dt = None
        if _from:
            try:
                start_dt = datetime.strptime(_from, "%Y-%m-%d")
            except Exception:
                start_dt = None
        if _to:
            try:
                # نخلي النهاية شاملة لليوم المحدد
                end_dt = datetime.strptime(_to, "%Y-%m-%d") + timedelta(days=1)
            except Exception:
                end_dt = None
        return start_dt, end_dt

    r_start, r_end = _parse_range(dfrom, dto)

    def _count_by_datecol(df, status_value, date_col):
        if df.empty or 'Status' not in df.columns or date_col not in df.columns:
            return 0
        tmp = df.copy()
        tmp[date_col] = pd.to_datetime(tmp[date_col], errors='coerce')
        mask = (tmp['Status'] == status_value) & tmp[date_col].notna()
        if r_start is not None:
            mask = mask & (tmp[date_col] >= r_start)
        if r_end is not None:
            mask = mask & (tmp[date_col] < r_end)
        return int(mask.sum())

    # عدّ الحالات حسب تاريخ تحديث الحالة
    cnt_shipping = _count_by_datecol(base_all, STATUS_SHIPPING, 'Shipping At')
    cnt_delivered = _count_by_datecol(base_all, STATUS_DELIVERED, 'Delivered At')
    cnt_returned = _count_by_datecol(base_all, STATUS_RETURNED, 'Returned At')

    # قيد التجهيز عادة يعتمد على تاريخ الإنشاء (لأنها الحالة الافتراضية)
    cnt_ready = 0
    try:
        if 'Status' in base_all.columns and 'Time and Date' in base_all.columns:
            tmp = base_all.copy()
            tmp['Time and Date'] = pd.to_datetime(tmp['Time and Date'], errors='coerce')
            mask = (tmp['Status'] == STATUS_READY) & tmp['Time and Date'].notna()
            if r_start is not None:
                mask = mask & (tmp['Time and Date'] >= r_start)
            if r_end is not None:
                mask = mask & (tmp['Time and Date'] < r_end)
            cnt_ready = int(mask.sum())
    except Exception:
        cnt_ready = 0

    summary = store.stats_global(d)

    # ✅ استبدال أعداد الحالات في summary حتى تظهر الإحصائيات حسب تاريخ تحديث الحالة
    try:
        status_labels = [STATUS_DELIVERED, STATUS_RETURNED, STATUS_SHIPPING, STATUS_READY]
        total_orders = int(summary.get("العدد الكلي للطلبات", 0) or 0)

        # overwrite counts
        summary["عدد " + STATUS_DELIVERED] = int(cnt_delivered)
        summary["عدد " + STATUS_RETURNED]  = int(cnt_returned)
        summary["عدد " + STATUS_SHIPPING]  = int(cnt_shipping)
        summary["عدد " + STATUS_READY]     = int(cnt_ready)

        # overwrite percents (نسبة ضمن إجمالي الطلبات في نفس الفلتر)
        def _pct(x):
            return (float(x) / float(total_orders) * 100.0) if total_orders else 0.0

        summary["نسبة " + STATUS_DELIVERED + " %"] = round(_pct(cnt_delivered), 2)
        summary["نسبة " + STATUS_RETURNED  + " %"] = round(_pct(cnt_returned), 2)
        summary["نسبة " + STATUS_SHIPPING  + " %"] = round(_pct(cnt_shipping), 2)
        summary["نسبة " + STATUS_READY     + " %"] = round(_pct(cnt_ready), 2)
    except Exception:
        pass

    # ─────────────────────────────
    # 3) بيانات حسب السعر (للجدول)
    # ─────────────────────────────
    by_price_df = store.stats_by_product_price(d)
    if not by_price_df.empty:
        by_price_df = by_price_df.fillna("")
        by_price = by_price_df.to_dict(orient="records")
    else:
        by_price = []

    # ─────────────────────────────
    # 4) الترند اليومي (للـ Line chart + جدول)
    # ─────────────────────────────
    daily = []
    daily_df = store.daily_trend(d)
    if daily_df is not None and not daily_df.empty:
        for _, r in daily_df.iterrows():
            date_val = r.get("Date")
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)

            daily.append({
                "Date": date_str,
                "Order Count": int(r.get("Order Count", 0) or 0),
                "Trend": r.get("Trend", ""),
            })

    # ─────────────────────────────
    # 5) الإيراد الكلي للطلبات الموصَّلة
    # ─────────────────────────────
    rev = 0.0
    if not d.empty and 'Status' in d.columns and 'Order Price' in d.columns:
        delivered_mask = (d['Status'] == STATUS_DELIVERED)
        rev = pd.to_numeric(d.loc[delivered_mask, 'Order Price'], errors='coerce').sum()

    # ─────────────────────────────
    # 6) تحليل حسب المنتج + الربح من المخزن
    # ─────────────────────────────
    product_rows = []
    if not d.empty and 'Product Name' in d.columns and 'Status' in d.columns:

        for prod_name, g in d.groupby('Product Name'):
            if not str(prod_name).strip():
                continue

            total = len(g)
            delivered_count = int((g['Status'] == STATUS_DELIVERED).sum())
            returned_count = int((g['Status'] == STATUS_RETURNED).sum())
            shipping_count = int((g['Status'] == STATUS_SHIPPING).sum())
            shipping_amount = float(pd.to_numeric(g.loc[g['Status'] == STATUS_SHIPPING, 'Order Price'], errors='coerce').sum() or 0)


            delivered_amount = pd.to_numeric(
                g.loc[g['Status'] == STATUS_DELIVERED, 'Order Price'],
                errors='coerce'
            ).sum()

            return_rate = (returned_count / total * 100) if total else 0.0
            deliver_rate = (delivered_count / total * 100) if total else 0.0

            # محاولة قراءة تكاليف المنتج من المخزن
            inv_row = None
            try:
                inv_match = inventory.df[inventory.df['Product Name'].astype(str) == str(prod_name)]
                if not inv_match.empty:
                    inv_row = inv_match.iloc[0].to_dict()
            except Exception:
                inv_row = None

            cost_per_unit = 0.0
            avg_price = 0.0
            profit_per_unit = 0.0
            net_profit = 0.0

            if inv_row is not None:
                mpu = float(pd.to_numeric(inv_row.get('Meters per Unit'), errors='coerce') or 0)
                fabric_price = float(pd.to_numeric(inv_row.get('Fabric Meter Price'), errors='coerce') or 0)
                sew_cost = float(pd.to_numeric(inv_row.get('Sewing Cost'), errors='coerce') or 0)
                access_cost = float(pd.to_numeric(inv_row.get('Accessories Cost'), errors='coerce') or 0)
                extra_cost = float(pd.to_numeric(inv_row.get('Extra Costs'), errors='coerce') or 0)

                cost_per_unit = mpu * fabric_price + sew_cost + access_cost + extra_cost

            if delivered_count:
                avg_price = float(delivered_amount) / delivered_count if delivered_amount else 0.0
                profit_per_unit = avg_price - cost_per_unit
                net_profit = profit_per_unit * delivered_count

            product_rows.append({
                "المنتج": prod_name,
                "عدد الطلبات": int(total),
                "عدد الواصل": int(delivered_count),
                "عدد الراجع": int(returned_count),
                "عدد قيد التوصيل": int(shipping_count),
                "مبلغ قيد التوصيل": float(shipping_amount or 0),
                "نسبة الراجع %": round(return_rate, 2),
                "نسبة الوصول %": round(deliver_rate, 2),
                "الإيراد (واصل)": float(delivered_amount or 0),
                "تكلفة القطعة": round(cost_per_unit, 2),
                "متوسط سعر البيع": round(avg_price, 2),
                "الربح/قطعة": round(profit_per_unit, 2),
                "صافي الربح": round(net_profit, 2),
            })

    # الأعمدة الظاهرة في جدول تحليل المنتج
    product_cols = [
        "المنتج", "عدد الطلبات", "عدد الواصل", "عدد الراجع", "عدد قيد التوصيل", "مبلغ قيد التوصيل",
        "نسبة الراجع %", "نسبة الوصول %",
        "الإيراد (واصل)", "تكلفة القطعة",
        "متوسط سعر البيع", "الربح/قطعة", "صافي الربح",
    ]

    # ─────────────────────────────
    # 7) بيانات الجارتات (توزيع الحالات)
    # ─────────────────────────────
    status_labels = [STATUS_DELIVERED, STATUS_RETURNED, STATUS_SHIPPING, STATUS_READY]
    status_counts = []
    if not d.empty and 'Status' in d.columns:
        for s in status_labels:
            status_counts.append(int((d['Status'] == s).sum()))
    else:
        status_counts = [0, 0, 0, 0]


    # ─────────────────────────────
    # 7b) مبالغ الحالات + صافي الربح (تفصيلي)
    # ─────────────────────────────
    status_amounts = {STATUS_DELIVERED: 0.0, STATUS_RETURNED: 0.0, STATUS_SHIPPING: 0.0, STATUS_READY: 0.0}
    if not d.empty and 'Status' in d.columns and 'Order Price' in d.columns:
        for s in status_amounts.keys():
            status_amounts[s] = float(pd.to_numeric(d.loc[d['Status'] == s, 'Order Price'], errors='coerce').sum() or 0)

    delivered_orders_count = int((d['Status'] == STATUS_DELIVERED).sum()) if (not d.empty and 'Status' in d.columns) else 0
    shipping_total = float(delivered_orders_count * (shipping_fee or 0))

    # حساب تكلفة الخام + الخياطة + الإكسسوارات + التكاليف الأخرى للطلبات الموصلة (اعتماداً على اسم المنتج/المخزن)
    cogs_total = 0.0
    cogs_breakdown = []  # rows per product
    prod_qty_total = {}
    prod_qty_delivered = {}
    prod_qty_returned = {}
    prod_qty_shipping = {}
    prod_amt_shipping = {}  # مبلغ قيد التوصيل موزّع على المنتجات

    def _cost_per_unit(prod_name: str) -> float:
        try:
            inv_match = inventory.df[inventory.df['Product Name'].astype(str) == str(prod_name)]
            if inv_match.empty:
                return 0.0
            r = inv_match.iloc[0].to_dict()
            mpu = float(pd.to_numeric(r.get('Meters per Unit'), errors='coerce') or 0)
            fabric_price = float(pd.to_numeric(r.get('Fabric Meter Price'), errors='coerce') or 0)
            sew_cost = float(pd.to_numeric(r.get('Sewing Cost'), errors='coerce') or 0)
            acc_cost = float(pd.to_numeric(r.get('Accessories Cost'), errors='coerce') or 0)
            extra = float(pd.to_numeric(r.get('Extra Costs'), errors='coerce') or 0)
            return (mpu * fabric_price) + sew_cost + acc_cost + extra
        except Exception:
            return 0.0

    # تجميع بالاعتماد على Items إذا موجود (يدعم أكثر من منتج في الطلب)
    if not d.empty:
        for _, rr in d.iterrows():
            rowd = rr.to_dict()
            items = parse_items_from_row(rowd)
            if not items:
                continue
            st = rowd.get('Status')
            order_price_f = float(pd.to_numeric(rowd.get('Order Price'), errors='coerce') or 0)
            total_items_qty = 0
            try:
                total_items_qty = sum(int(it.get('qty', 1) or 1) for it in items if it.get('name'))
            except Exception:
                total_items_qty = 0

            for it in items:
                nm = it.get('name')
                q = int(it.get('qty', 1) or 1)
                if not nm:
                    continue

                prod_qty_total[nm] = prod_qty_total.get(nm, 0) + q

                if st == STATUS_DELIVERED:
                    prod_qty_delivered[nm] = prod_qty_delivered.get(nm, 0) + q
                    cogs_total += _cost_per_unit(nm) * q
                elif st == STATUS_RETURNED:
                    prod_qty_returned[nm] = prod_qty_returned.get(nm, 0) + q
                elif st == STATUS_SHIPPING:
                    prod_qty_shipping[nm] = prod_qty_shipping.get(nm, 0) + q
                    # توزيع مبلغ الطلب على المنتجات حتى لا يتضاعف عند وجود أكثر من منتج
                    if total_items_qty:
                        share = (order_price_f / total_items_qty) * q
                    else:
                        share = order_price_f
                    prod_amt_shipping[nm] = prod_amt_shipping.get(nm, 0.0) + float(share or 0)


    # أفضل المنتجات (الأكثر طلباً + الأفضل تسليماً بأقل راجع)
    top_ordered = sorted(prod_qty_total.items(), key=lambda x: x[1], reverse=True)[:10]
    # الأفضل تسليماً = أعلى تسليم مع أقل نسبة راجع
    best_delivery_rows = []
    for nm, dq in prod_qty_delivered.items():
        rq = prod_qty_returned.get(nm, 0)
        totalq = prod_qty_total.get(nm, dq + rq)
        rate = (rq / totalq * 100) if totalq else 0.0
        best_delivery_rows.append({'name': nm, 'delivered_qty': dq, 'returned_qty': rq, 'return_rate': rate})
    best_delivery_rows.sort(key=lambda r: (r['return_rate'], -r['delivered_qty']))
    best_delivered = best_delivery_rows[:10]

    # صافي الربح = الإيراد الموصّل - (تكلفة الخام والخياطة...) - الشحن - الإعلان - مصاريف أخرى
    net_profit = float(rev - cogs_total - shipping_total - float(ads_cost or 0) - float(other_cost or 0))
# قائمة الصفحات لجميع الطلبات (بدون فلتر)
    try:
        pages = sorted(list({str(x) for x in store.df['Page Name'].dropna().unique()}))
    except Exception:
        pages = []

    # ─────────────────────────────
    # 8) إرسال البيانات للتمبلت STATS_HTML
    # ─────────────────────────────
    return render_template_string(
        STATS_HTML,
        summary=summary,
        by_price=by_price,
        # ملاحظة: الأعمدة هنا تعتمد على ما ترجعه stats_by_product_price
        price_cols=list(by_price_df.columns) if not by_price_df.empty else [],
        by_product=product_rows,
        product_cols=product_cols,
        daily=daily,
        dfrom=dfrom,
        dto=dto,
        sel_page=sel_page,
        pages=pages,
        revenue=rev,

        shipping_fee=shipping_fee,
        ads_cost=ads_cost,
        other_cost=other_cost,
        status_amounts=status_amounts,
        net_profit=net_profit,
        cogs_total=cogs_total,
        shipping_total=shipping_total,
        top_ordered=top_ordered,
        best_delivered=best_delivered,
status_labels=status_labels,
        status_counts=status_counts,
    )




@app.route('/daily_analysis', methods=['GET'])
@login_required
def daily_analysis():
    store.reload_if_changed()
    try:
        inventory.reload_if_changed()
    except Exception:
        pass

    dfrom = (request.args.get('dfrom') or '').strip()
    dto   = (request.args.get('dto') or '').strip()
    offset = request.args.get('offset', '1')
    ship_fee = request.args.get('ship_fee', '0')
    ads = request.args.get('ads', '0')
    other = request.args.get('other', '0')

    def _to_float(x):
        try:
            return float(str(x or '0').replace(',', '').strip())
        except Exception:
            return 0.0

    offset_days = 1
    try:
        offset_days = int(float(offset))
        if offset_days < 0:
            offset_days = 0
    except Exception:
        offset_days = 1

    ship_fee_f = _to_float(ship_fee)
    ads_f = _to_float(ads)
    other_f = _to_float(other)

    if not dfrom or not dto:
        return render_template_string(
            DAILY_ANALYSIS_HTML,
            title="تحليل يومي",
            dfrom=dfrom, dto=dto, offset=offset_days,
            ship_fee=ship_fee_f, ads=ads_f, other=other_f,
            ready=None
        )

    try:
        inv_start = datetime.strptime(dfrom, "%Y-%m-%d")
        inv_end = datetime.strptime(dto, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except Exception:
        flash("صيغة التاريخ غير صحيحة", "err")
        return redirect(url_for('daily_analysis'))

    base_start = inv_start - timedelta(days=offset_days)
    base_end = (datetime.strptime(dto, "%Y-%m-%d") - timedelta(days=offset_days)) + timedelta(days=1) - timedelta(seconds=1)

    d = store.df.copy()

    for col in ["Time and Date", "Delivered At", "Returned At", "Status Updated At"]:
        if col in d.columns:
            d[col] = pd.to_datetime(d[col], errors="coerce")

    if "Time and Date" in d.columns:
        base_df = d[(d["Time and Date"] >= base_start) & (d["Time and Date"] <= base_end)].copy()
    else:
        base_df = d.copy()

    base_total = int(len(base_df))

    delivered_df = base_df[(base_df.get("Status") == STATUS_DELIVERED)]
    returned_df = base_df[(base_df.get("Status") == STATUS_RETURNED)]

    if "Delivered At" in base_df.columns:
        delivered_in_invoice = delivered_df[(delivered_df["Delivered At"] >= inv_start) & (delivered_df["Delivered At"] <= inv_end)]
    else:
        delivered_in_invoice = delivered_df.iloc[0:0]

    if "Returned At" in base_df.columns:
        returned_in_invoice = returned_df[(returned_df["Returned At"] >= inv_start) & (returned_df["Returned At"] <= inv_end)]
    else:
        returned_in_invoice = returned_df.iloc[0:0]

    delivered = int(len(delivered_in_invoice))
    returned = int(len(returned_in_invoice))

    ready = int((base_df.get("Status") == STATUS_READY).sum()) if "Status" in base_df.columns else 0
    shipping = int((base_df.get("Status") == STATUS_SHIPPING).sum()) if "Status" in base_df.columns else 0
    pending_total = int(ready + shipping)

    revenue = float(pd.to_numeric(delivered_in_invoice.get("Order Price"), errors="coerce").sum() or 0)

    def _cost_per_unit(prod_name: str) -> float:
        try:
            inv_match = inventory.df[inventory.df['Product Name'].astype(str) == str(prod_name)]
            if inv_match.empty:
                return 0.0
            r = inv_match.iloc[0].to_dict()
            mpu = float(pd.to_numeric(r.get('Meters per Unit'), errors='coerce') or 0)
            fabric_price = float(pd.to_numeric(r.get('Fabric Meter Price'), errors='coerce') or 0)
            sew_cost = float(pd.to_numeric(r.get('Sewing Cost'), errors='coerce') or 0)
            acc_cost = float(pd.to_numeric(r.get('Accessories Cost'), errors='coerce') or 0)
            extra = float(pd.to_numeric(r.get('Extra Costs'), errors='coerce') or 0)
            return (mpu * fabric_price) + sew_cost + acc_cost + extra
        except Exception:
            return 0.0

    cogs_total = 0.0
    prod_qty_total = {}
    prod_qty_delivered = {}
    prod_qty_returned = {}
    prod_qty_shipping = {}
    prod_amt_shipping = {}  # مبلغ قيد التوصيل موزّع على المنتجات

    if not base_df.empty:
        delivered_txns = set(str(x).strip() for x in delivered_in_invoice.get("Transaction ID", pd.Series(dtype=str)).astype(str).tolist())
        returned_txns = set(str(x).strip() for x in returned_in_invoice.get("Transaction ID", pd.Series(dtype=str)).astype(str).tolist())
        shipping_txns = set()
        try:
            if "Status" in base_df.columns and "Transaction ID" in base_df.columns:
                shipping_txns = set(str(x).strip() for x in base_df.loc[base_df["Status"] == STATUS_SHIPPING, "Transaction ID"].astype(str).tolist())
        except Exception:
            shipping_txns = set()


        for _, rr in base_df.iterrows():
            rowd = rr.to_dict()
            txn = str(rowd.get("Transaction ID", "")).strip()

            items = parse_items_from_row(rowd)
            if not items:
                continue

            order_price_f = float(pd.to_numeric(rowd.get("Order Price"), errors="coerce") or 0)
            total_items_qty = 0
            try:
                total_items_qty = sum(int(it.get('qty', 1) or 1) for it in items if it.get('name'))
            except Exception:
                total_items_qty = 0

            for it in items:
                nm = it.get('name')
                q = int(it.get('qty', 1) or 1)
                if not nm:
                    continue

                prod_qty_total[nm] = prod_qty_total.get(nm, 0) + q

                if txn in delivered_txns:
                    prod_qty_delivered[nm] = prod_qty_delivered.get(nm, 0) + q
                    cogs_total += _cost_per_unit(nm) * q
                if txn in returned_txns:
                    prod_qty_returned[nm] = prod_qty_returned.get(nm, 0) + q
                if txn in shipping_txns:
                    prod_qty_shipping[nm] = prod_qty_shipping.get(nm, 0) + q
                    if total_items_qty:
                        share = (order_price_f / total_items_qty) * q
                    else:
                        share = order_price_f
                    prod_amt_shipping[nm] = prod_amt_shipping.get(nm, 0.0) + float(share or 0)


    shipping_total = float(delivered * (ship_fee_f or 0))
    net_profit = float(revenue - cogs_total - shipping_total - ads_f - other_f)

    denom = (delivered + returned)
    return_rate = round((returned / denom * 100) if denom else 0.0, 2)

    best_rows = []
    for nm, dq in prod_qty_delivered.items():
        rq = prod_qty_returned.get(nm, 0)
        totalq = prod_qty_total.get(nm, dq + rq)
        rate = (rq / totalq * 100) if totalq else 0.0
        best_rows.append({'name': nm, 'delivered_qty': int(dq), 'returned_qty': int(rq), 'return_rate': float(rate)})
    best_rows.sort(key=lambda r: (-r['delivered_qty'], r['return_rate']))
    top_delivered = best_rows[:10]

    ret_rows = []
    for nm, rq in prod_qty_returned.items():
        dq = prod_qty_delivered.get(nm, 0)
        totalq = prod_qty_total.get(nm, dq + rq)
        rate = (rq / totalq * 100) if totalq else 0.0
        ret_rows.append({'name': nm, 'returned_qty': int(rq), 'delivered_qty': int(dq), 'return_rate': float(rate)})
    ret_rows.sort(key=lambda r: (-r['returned_qty'], r['return_rate']))
    top_returned = ret_rows[:10]

    page_rows = []
    if "Page Name" in base_df.columns:
        total_rev_for_ads = revenue if revenue else 0.0

        for pg, g in base_df.groupby("Page Name", dropna=False):
            pg_name = str(pg) if str(pg).strip() and str(pg).lower() != "nan" else "غير محدد"

            if not delivered_in_invoice.empty:
                gd = delivered_in_invoice[delivered_in_invoice["Page Name"] == pg].copy()
            else:
                gd = base_df.iloc[0:0]
            if not returned_in_invoice.empty:
                gr = returned_in_invoice[returned_in_invoice["Page Name"] == pg].copy()
            else:
                gr = base_df.iloc[0:0]

            delivered_orders = int(len(gd))
            returned_orders = int(len(gr))

            # قيد التوصيل ضمن الفلتر
            gs = g[g.get('Status') == STATUS_SHIPPING].copy() if ('Status' in g.columns) else g.iloc[0:0]
            shipping_orders = int(len(gs))
            shipping_amount = float(pd.to_numeric(gs.get('Order Price'), errors='coerce').sum() or 0)

            pg_revenue = float(pd.to_numeric(gd.get("Order Price"), errors="coerce").sum() or 0)

            pg_cogs = 0.0
            if delivered_orders and not gd.empty:
                for _, rr in gd.iterrows():
                    rowd = rr.to_dict()
                    items = parse_items_from_row(rowd)
                    for it in items:
                        nm = it.get('name')
                        q = int(it.get('qty', 1) or 1)
                        if not nm:
                            continue
                        pg_cogs += _cost_per_unit(nm) * q

            pg_shipping = float(delivered_orders * (ship_fee_f or 0))

            pg_ads_share = 0.0
            if ads_f and total_rev_for_ads:
                pg_ads_share = float(ads_f * (pg_revenue / total_rev_for_ads))

            pg_net = float(pg_revenue - pg_cogs - pg_shipping - pg_ads_share)

            pg_denom = delivered_orders + returned_orders
            pg_return_rate = float((returned_orders / pg_denom * 100) if pg_denom else 0.0)

            page_rows.append({
                "page": pg_name,
                "delivered_orders": delivered_orders,
                "returned_orders": returned_orders,
                "shipping_orders": shipping_orders,
                "shipping_amount": shipping_amount,
                "delivery_rate": float((delivered_orders / (delivered_orders + returned_orders) * 100) if (delivered_orders + returned_orders) else 0.0),
                "revenue": pg_revenue,
                "cogs": pg_cogs,
                "shipping": pg_shipping,
                "ads_share": pg_ads_share,
                "net_profit": pg_net,
                "return_rate": pg_return_rate
            })

    page_rows.sort(key=lambda r: r["net_profit"], reverse=True)

    # أفضل متجر: أكثر واصل + أفضل نسبة وصول
    best_store_by_delivered = None
    best_store_by_delivery_rate = None
    try:
        if page_rows:
            best_store_by_delivered = max(page_rows, key=lambda r: (r.get("delivered_orders", 0), r.get("delivery_rate", 0)))
            best_store_by_delivery_rate = max(page_rows, key=lambda r: (r.get("delivery_rate", 0), r.get("delivered_orders", 0)))
    except Exception:
        best_store_by_delivered = None
        best_store_by_delivery_rate = None

    return render_template_string(
        DAILY_ANALYSIS_HTML,
        title="تحليل يومي",
        dfrom=dfrom, dto=dto, offset=offset_days,
        ship_fee=ship_fee_f, ads=ads_f, other=other_f,

        base_total=base_total,
        base_from=base_start.strftime("%Y-%m-%d"),
        base_to=base_end.strftime("%Y-%m-%d"),

        delivered=delivered,
        returned=returned,
        ready=ready,
        shipping=shipping,
        pending_total=pending_total,

        revenue=revenue,
        cogs_total=cogs_total,
        shipping_total=shipping_total,
        net_profit=net_profit,
        return_rate=return_rate,

        top_delivered=top_delivered,
        top_returned=top_returned,
        page_rows=page_rows,
        best_store_by_delivered=best_store_by_delivered,
        best_store_by_delivery_rate=best_store_by_delivery_rate
    )

@app.route('/download/excel')
@login_required
def download_excel():
    # make sure latest is saved, then send
    store.save()
    d = Path(EXCEL_FILE).parent
    return send_from_directory(str(d), Path(EXCEL_FILE).name, as_attachment=True)


# ------------------------------ REPORTS ---------------------------------

def _coerce_numeric_series(s):
    return pd.to_numeric(s, errors="coerce").fillna(0)

def _orders_df():
    try:
        store.reload_if_changed()
    except Exception:
        pass
    d = store.df.copy()
    if "Order Price" in d.columns:
        d["Order Price"] = pd.to_numeric(d["Order Price"], errors="coerce")
    return d

def _inventory_snapshot_df():
    try:
        inventory.reload()
    except Exception:
        pass
    try:
        return inventory.df.copy()
    except Exception:
        return pd.DataFrame()

@app.route('/reports/orders/<status_key>')
@login_required
def report_orders_status(status_key):
    """
    Export analysis for orders in a given status:
    - status_key: ready | shipping
    - optional filters: from=YYYY-MM-DD, to=YYYY-MM-DD
    """
    status_key = (status_key or '').strip().lower()
    status_map = {'ready': STATUS_READY, 'shipping': STATUS_SHIPPING}
    if status_key not in status_map:
        abort(404)

    status = status_map[status_key]
    d = _orders_df()

    # date filter on Time and Date
    d["Time and Date"] = pd.to_datetime(d.get("Time and Date"), errors="coerce")
    d = d.dropna(subset=["Time and Date"])
    d["Date"] = d["Time and Date"].dt.date

    dfrom = (request.args.get('from') or '').strip()
    dto = (request.args.get('to') or '').strip()
    if dfrom:
        try:
            d_from = datetime.strptime(dfrom, '%Y-%m-%d').date()
            d = d[d["Date"] >= d_from]
        except Exception:
            pass
    if dto:
        try:
            d_to = datetime.strptime(dto, '%Y-%m-%d').date()
            d = d[d["Date"] <= d_to]
        except Exception:
            pass

    df = d[d.get("Status") == status].copy()

    if df.empty:
        flash(f'لا توجد طلبات بحالة: {status}', 'err')
        return redirect(url_for('home', status=status))

    # Orders list
    cols = [c for c in [
        "Transaction ID", "Time and Date", "Page Name", "Contact Numbers",
        "Address", "Order Price", "Status", "Status Updated At", "Notes", "Items"
    ] if c in df.columns]
    orders_sheet = df[cols].copy().fillna('')

    # Product aggregation (Pieces + Orders Count)
    prod_rows = []
    for _, r in df.iterrows():
        row = r.to_dict()
        items = parse_items_from_row(row)
        if not items:
            # fallback to Product Name
            items = [{"code": "", "name": str(row.get("Product Name", "") or "").strip(), "qty": 1}]
        seen = set()
        for it in items:
            code = str(it.get("code", "") or "").strip()
            name = str(it.get("name", "") or "").strip()
            qty = int(it.get("qty", 1) or 1)
            key = (code, name)
            prod_rows.append({
                "Product Code": code,
                "Product Name": name,
                "Pieces": qty,
                "Order Price (full order)": float(row.get("Order Price") or 0),
                "Transaction ID": str(row.get("Transaction ID", "") or ""),
                "Page Name": str(row.get("Page Name", "") or ""),
            })
            seen.add(key)

    prod_df = pd.DataFrame(prod_rows)
    if prod_df.empty:
        prod_summary = pd.DataFrame(columns=["Product Code","Product Name","Orders Count","Pieces Total","Sum Order Price (may double count)"])
    else:
        prod_summary = (
            prod_df.groupby(["Product Code","Product Name"], dropna=False)
                  .agg(**{
                      "Orders Count": ("Transaction ID", "nunique"),
                      "Pieces Total": ("Pieces", "sum"),
                      "Sum Order Price (may double count)": ("Order Price (full order)", "sum"),
                  })
                  .reset_index()
                  .sort_values(by=["Orders Count","Pieces Total"], ascending=[False, False])
        )

    # Page summary
    page_summary = (
        df.groupby("Page Name", dropna=False)
          .agg(**{
              "Orders": ("Transaction ID", "count"),
              "Total Amount": ("Order Price", "sum"),
          })
          .reset_index()
          .sort_values(by=["Orders","Total Amount"], ascending=[False, False])
    )

    # Overall summary
    overall = pd.DataFrame([{
        "Status": status,
        "Orders Count": int(len(df)),
        "Total Amount": float(_coerce_numeric_series(df.get("Order Price")).sum()),
        "Date From": dfrom or "",
        "Date To": dto or "",
        "Generated At": now_str(),
    }])

    out_dir = Path(user_data_dir()) / 'reports'
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = out_dir / f"orders_{status_key}_report_{stamp}.xlsx"

    with pd.ExcelWriter(str(out_path), engine='openpyxl') as writer:
        overall.to_excel(writer, index=False, sheet_name='Summary')
        orders_sheet.to_excel(writer, index=False, sheet_name='Orders')
        prod_summary.to_excel(writer, index=False, sheet_name='By Product')
        page_summary.to_excel(writer, index=False, sheet_name='By Page')

    return send_from_directory(str(out_dir), out_path.name, as_attachment=True)


@app.route('/reports/inventory/withdrawn')
@login_required
def report_inventory_withdrawn():
    """
    Export full inventory-withdrawn report based on movement log:
    - Withdraw movements (Delta < 0)
    - Summary by product
    - Summary by order/ref
    - Current inventory snapshot
    Optional: from=YYYY-MM-DD, to=YYYY-MM-DD
    """
    try:
        inventory.movements.reload()
    except Exception:
        pass

    mv = getattr(inventory, 'movements', None)
    mvdf = mv.df.copy() if mv is not None and hasattr(mv, 'df') else pd.DataFrame(columns=InventoryMovementStore.COLS)

    mvdf['Date'] = mvdf.get('Date', '').astype(str)
    mvdf['Delta'] = pd.to_numeric(mvdf.get('Delta'), errors='coerce').fillna(0).astype(int)
    mvdf['Movement Type'] = mvdf.get('Movement Type', '').astype(str)

    dfrom = (request.args.get('from') or '').strip()
    dto = (request.args.get('to') or '').strip()
    if dfrom:
        mvdf = mvdf[mvdf['Date'] >= dfrom]
    if dto:
        mvdf = mvdf[mvdf['Date'] <= dto]

    wd = mvdf[(mvdf['Movement Type'].str.lower() == 'withdraw') & (mvdf['Delta'] < 0)].copy()

    if wd.empty:
        flash('لا توجد حركات سحب (Withdraw) ضمن الفلاتر الحالية', 'err')
        return redirect(url_for('inventory_home'))

    wd['Withdrawn Pieces'] = (-wd['Delta']).astype(int)

    by_product = (
        wd.groupby(['Product Code','Product Name'], dropna=False)
          .agg(**{
              'Withdraw Movements': ('MoveID','count'),
              'Withdrawn Pieces': ('Withdrawn Pieces','sum'),
              'First Date': ('Date','min'),
              'Last Date': ('Date','max'),
          })
          .reset_index()
          .sort_values(by=['Withdrawn Pieces','Withdraw Movements'], ascending=[False, False])
    )

    by_ref = (
        wd.groupby(['Ref'], dropna=False)
          .agg(**{
              'Withdraw Movements': ('MoveID','count'),
              'Withdrawn Pieces': ('Withdrawn Pieces','sum'),
              'First DateTime': ('DateTime','min'),
              'Last DateTime': ('DateTime','max'),
          })
          .reset_index()
          .sort_values(by=['Withdrawn Pieces','Withdraw Movements'], ascending=[False, False])
    )

    inv_df = _inventory_snapshot_df().fillna('')
    out_dir = Path(user_data_dir()) / 'reports'
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = out_dir / f"inventory_withdrawn_{stamp}.xlsx"

    with pd.ExcelWriter(str(out_path), engine='openpyxl') as writer:
        pd.DataFrame([{
            "Generated At": now_str(),
            "Date From": dfrom or "",
            "Date To": dto or "",
            "Total Withdraw Movements": int(len(wd)),
            "Total Withdrawn Pieces": int(wd["Withdrawn Pieces"].sum()),
        }]).to_excel(writer, index=False, sheet_name='Summary')

        wd.fillna('').to_excel(writer, index=False, sheet_name='Withdraw Movements')
        by_product.to_excel(writer, index=False, sheet_name='By Product')
        by_ref.to_excel(writer, index=False, sheet_name='By Ref')
        inv_df.to_excel(writer, index=False, sheet_name='Inventory Snapshot')

    return send_from_directory(str(out_dir), out_path.name, as_attachment=True)

def delete_by_code(self, code):
    self.df = self.df[self.df["code"] != code]
    self.save()

@app.route("/inventory/delete/<code>", methods=["POST"])
def inventory_delete(code):
    inventory = InventoryStore()

    item = inventory.get_by_code(code)
    if not item:
        flash("المنتج غير موجود", "danger")
        return redirect(url_for("inventory_home"))

    inventory.delete_by_code(code)

    flash("تم حذف المنتج بنجاح", "success")
    return redirect(url_for("inventory_home"))


@app.route('/reports/system/export')
@login_required
def system_export():
    """
    Create a FULL system export (ZIP) for AI analysis:
    - Excel workbook with key sheets (Orders + Inventory + Movements + Summaries)
    - JSON summary (easy for LLM ingestion)
    """
    d = _orders_df()
    inv_df = _inventory_snapshot_df()

    # Orders subsets
    d_all = d.copy()
    d_all["Order Price"] = pd.to_numeric(d_all.get("Order Price"), errors="coerce")
    sheets = {}
    sheets["Orders_All"] = d_all.fillna('')

    def _subset(status):
        if "Status" not in d_all.columns:
            return d_all.iloc[0:0].copy()
        return d_all[d_all["Status"] == status].copy().fillna('')

    sheets["Orders_Ready"] = _subset(STATUS_READY)
    sheets["Orders_Shipping"] = _subset(STATUS_SHIPPING)
    sheets["Orders_Delivered"] = _subset(STATUS_DELIVERED)
    sheets["Orders_Returned"] = _subset(STATUS_RETURNED)

    # Global stats
    try:
        global_stats = store.stats_global(d_all)
    except Exception:
        global_stats = {}
    sheets["Stats_Global"] = pd.DataFrame([{"Metric": k, "Value": v} for k, v in global_stats.items()])

    try:
        by_price = store.stats_by_product_price(d_all)
    except Exception:
        by_price = pd.DataFrame()
    sheets["Stats_By_Price"] = by_price.fillna('')

    try:
        trend = store.daily_trend(d_all)
    except Exception:
        trend = pd.DataFrame()
    sheets["Stats_Daily_Trend"] = trend.fillna('')

    # Inventory + movements
    sheets["Inventory_Current"] = inv_df.fillna('')

    mvdf = pd.DataFrame(columns=InventoryMovementStore.COLS)
    try:
        inventory.movements.reload()
        mvdf = inventory.movements.df.copy()
    except Exception:
        pass
    sheets["Inventory_Movements"] = mvdf.fillna('')

    # Movement summary pivot
    try:
        x = mvdf.copy()
        x['Delta'] = pd.to_numeric(x.get('Delta'), errors='coerce').fillna(0).astype(int)
        summary = (x.groupby(['Product Code','Product Name','Movement Type'])['Delta'].sum().reset_index())
        piv = summary.pivot_table(index=['Product Code','Product Name'],
                                  columns='Movement Type',
                                  values='Delta',
                                  aggfunc='sum',
                                  fill_value=0).reset_index()
    except Exception:
        piv = pd.DataFrame()
    sheets["Inventory_Movement_Summary"] = piv.fillna('')

    # JSON summary for AI
    try:
        json_summary = {
            "generated_at": now_str(),
            "orders": {
                "total": int(len(d_all)),
                "by_status": {
                    STATUS_READY: int((d_all.get("Status") == STATUS_READY).sum()) if "Status" in d_all.columns else 0,
                    STATUS_SHIPPING: int((d_all.get("Status") == STATUS_SHIPPING).sum()) if "Status" in d_all.columns else 0,
                    STATUS_DELIVERED: int((d_all.get("Status") == STATUS_DELIVERED).sum()) if "Status" in d_all.columns else 0,
                    STATUS_RETURNED: int((d_all.get("Status") == STATUS_RETURNED).sum()) if "Status" in d_all.columns else 0,
                },
                "amounts": {
                    "total": float(_coerce_numeric_series(d_all.get("Order Price")).sum()),
                    "delivered": float(_coerce_numeric_series(sheets["Orders_Delivered"].get("Order Price")).sum()) if not sheets["Orders_Delivered"].empty else 0.0,
                    "returned": float(_coerce_numeric_series(sheets["Orders_Returned"].get("Order Price")).sum()) if not sheets["Orders_Returned"].empty else 0.0,
                    "shipping": float(_coerce_numeric_series(sheets["Orders_Shipping"].get("Order Price")).sum()) if not sheets["Orders_Shipping"].empty else 0.0,
                    "ready": float(_coerce_numeric_series(sheets["Orders_Ready"].get("Order Price")).sum()) if not sheets["Orders_Ready"].empty else 0.0,
                }
            },
            "inventory": {
                "products": int(len(inv_df)) if inv_df is not None else 0,
            },
            "stats_global": global_stats,
        }
    except Exception:
        json_summary = {"generated_at": now_str()}

    out_dir = Path(user_data_dir()) / 'reports'
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    xlsx_path = out_dir / f"system_export_{stamp}.xlsx"
    json_path = out_dir / f"system_export_{stamp}.json"
    zip_path = out_dir / f"system_export_{stamp}.zip"

    # Write excel
    with pd.ExcelWriter(str(xlsx_path), engine='openpyxl') as writer:
        for name, df_sheet in sheets.items():
            # Excel sheet name limit 31
            sheet_name = name[:31]
            try:
                df_sheet.to_excel(writer, index=False, sheet_name=sheet_name)
            except Exception:
                pd.DataFrame(df_sheet).to_excel(writer, index=False, sheet_name=sheet_name)

    # Write json
    import json as _json
    json_path.write_text(_json.dumps(json_summary, ensure_ascii=False, indent=2), encoding="utf-8")

    # Write zip (xlsx + json)
    import zipfile as _zipfile
    with _zipfile.ZipFile(str(zip_path), 'w', compression=_zipfile.ZIP_DEFLATED) as z:
        z.write(str(xlsx_path), arcname=xlsx_path.name)
        z.write(str(json_path), arcname=json_path.name)

    return send_from_directory(str(out_dir), zip_path.name, as_attachment=True)





@app.route('/telegram/send_inventory_daily')
@login_required
def telegram_send_inventory_daily():
    """Send today's (or selected date) inventory daily report + withdrawn summary to Telegram."""
    d = (request.args.get('date') or date.today().isoformat()).strip()

    try:
        send_telegram(build_withdrawn_daily_summary(d))
    except Exception:
        pass

    try:
        b, fname = build_inventory_daily_excel_bytes(d)
        if b and fname:
            send_telegram_document(b, fname, caption=f"📄 تقرير المخزن اليومي - {d}")
    except Exception:
        pass

    flash('تم إرسال تقرير المخزن اليومي + ملخص السحب إلى التلكرام (إذا كانت الإعدادات مفعلة).', 'ok')
    return redirect(url_for('inventory_home'))


@app.route('/telegram/send_withdrawn_daily')
@login_required
def telegram_send_withdrawn_daily():
    """Send withdrawn-only daily summary to Telegram."""
    d = (request.args.get('date') or date.today().isoformat()).strip()
    try:
        send_telegram(build_withdrawn_daily_summary(d))
    except Exception:
        pass
    flash('تم إرسال ملخص السحب اليومي إلى التلكرام.', 'ok')
    return redirect(url_for('inventory_home'))

# ----------------------- Inventory dashboard cache -----------------------
_INV_DASH_CACHE = {
    "ts": 0.0,
    "inv_mtime": None,
    "mv_mtime": None,
    "orders_mtime": None,
    "result": None,
}

def _safe_mtime(p: str):
    try:
        return os.path.getmtime(p)
    except Exception:
        return None

def build_inventory_dashboard(inv_rows: list[dict]):
    """
    Build dashboard numbers fast.
    Cached for 60s or until any related file changes (orders/inventory/movements).
    """
    global _INV_DASH_CACHE

    inv_m = _safe_mtime(inventory.path)
    mv_m = _safe_mtime(inventory.movements.path)
    ord_m = _safe_mtime(store.path)

    now_ts = time.time()
    cached = _INV_DASH_CACHE.get("result")
    if (
        cached is not None
        and (now_ts - float(_INV_DASH_CACHE.get("ts") or 0)) < 60
        and _INV_DASH_CACHE.get("inv_mtime") == inv_m
        and _INV_DASH_CACHE.get("mv_mtime") == mv_m
        and _INV_DASH_CACHE.get("orders_mtime") == ord_m
    ):
        return cached

    # Ensure freshest in-memory data without heavy reload each request
    try:
        inventory.reload_if_changed()
    except Exception:
        pass
    try:
        inventory.movements.reload_if_changed()
    except Exception:
        pass
    try:
        store.reload_if_changed()
    except Exception:
        pass

    # Build name->code mapping for legacy items without code
    name_to_code = {}
    name_counts = {}
    for r in inv_rows:
        n = str(r.get("Product Name") or "").strip()
        c = str(r.get("Product Code") or "").strip()
        if not n or not c:
            continue
        name_counts[n] = name_counts.get(n, 0) + 1
        if name_counts[n] == 1:
            name_to_code[n] = c
        else:
            # duplicated name => don't resolve by name
            name_to_code.pop(n, None)

    delivered_by_code = {}
    returned_by_code = {}

    try:
        d = store.df
        if d is not None and not d.empty and 'Status' in d.columns:
            dd = d[d['Status'].isin([STATUS_DELIVERED, STATUS_RETURNED])].copy()
            # iterate rows efficiently
            for _, row in dd.iterrows():
                status = row.get('Status')
                items = parse_items_from_row(row.to_dict())
                if not items:
                    continue
                for it in items:
                    it_code = str(it.get('code', '') or '').strip()
                    it_name = str(it.get('name', '') or '').strip()
                    code = it_code or (name_to_code.get(it_name) if it_name else "")
                    if not code:
                        continue
                    try:
                        qty = int(it.get('qty', 1) or 1)
                    except Exception:
                        qty = 1
                    if qty <= 0:
                        qty = 1
                    if status == STATUS_DELIVERED:
                        delivered_by_code[code] = delivered_by_code.get(code, 0) + qty
                    elif status == STATUS_RETURNED:
                        returned_by_code[code] = returned_by_code.get(code, 0) + qty
    except Exception:
        pass

    # Top selling list
    stats_rows = []
    for r in inv_rows:
        code = str(r.get("Product Code") or "").strip()
        if not code:
            continue
        delivered = int(delivered_by_code.get(code, 0) or 0)
        returned = int(returned_by_code.get(code, 0) or 0)
        total = delivered + returned
        stats_rows.append({
            "Product Code": code,
            "Product Name": r.get("Product Name", ""),
            "Delivered Pieces": delivered,
            "Returned Pieces": returned,
            "Delivered %": round((delivered / total * 100) if total else 0.0, 2),
            "Returned %": round((returned / total * 100) if total else 0.0, 2),
        })

    stats_rows_sorted = sorted(
        stats_rows,
        key=lambda x: (int(x.get('Delivered Pieces', 0) or 0), int(x.get('Returned Pieces', 0) or 0)),
        reverse=True
    )

    delivered_total = sum(int(x.get("Delivered Pieces", 0) or 0) for x in stats_rows)
    returned_total = sum(int(x.get("Returned Pieces", 0) or 0) for x in stats_rows)
    overall_total = delivered_total + returned_total

    result = {
        "top_selling": stats_rows_sorted[:10],
        "overall_rates": {
            "delivered_pieces": delivered_total,
            "returned_pieces": returned_total,
            "delivered_pct": round((delivered_total / overall_total * 100) if overall_total else 0.0, 2),
            "returned_pct": round((returned_total / overall_total * 100) if overall_total else 0.0, 2),
        }
    }

    _INV_DASH_CACHE = {
        "ts": now_ts,
        "inv_mtime": inv_m,
        "mv_mtime": mv_m,
        "orders_mtime": ord_m,
        "result": result,
    }
    return result


@app.route('/inventory')
@login_required
def inventory_home():
    # Avoid heavy reload on every request
    try:
        inventory.reload_if_changed()
    except Exception:
        pass

    q = (request.args.get('q') or '').strip()
    all_rows = inventory.df.fillna("").to_dict(orient='records')

    if q:
        ql = q.lower()
        def _hit(r):
            return (ql in str(r.get('Product Name','')).lower()
                    or ql in str(r.get('Product Code','')).lower()
                    or ql in str(r.get('Type','')).lower())
        rows = [r for r in all_rows if _hit(r)]
    else:
        rows = all_rows

    added = request.args.get('added')
    taken = request.args.get('taken')

    dash = build_inventory_dashboard(all_rows)
    top_selling = dash.get("top_selling", [])
    overall_rates = dash.get("overall_rates", {"delivered_pieces":0,"returned_pieces":0,"delivered_pct":0,"returned_pct":0})

    return render_template_string(
        INVENTORY_HTML,
        rows=rows,
        all_rows=all_rows,
        q=q,
        added=added,
        taken=taken,
        name=(request.args.get('name') or '').strip(),
        top_selling=top_selling,
        overall_rates=overall_rates,
    )

@app.route('/inventory/product/<code>')
@login_required
def inventory_product(code):
    try:
        inventory.reload()
        inventory.movements.reload()
    except Exception:
        pass
    item = inventory.get_by_code(code)
    if not item:
        flash('المنتج غير موجود', 'err')
        return redirect(url_for('inventory_home'))

    stats = inventory_product_stats(code) or {}
    moves = []
    try:
        mv = inventory.movements.filter_by_product_code(code).copy()
        if not mv.empty:
            mv = mv.tail(50)
            moves = mv.fillna('').to_dict(orient='records')
    except Exception:
        pass
    return render_template_string(INVENTORY_PRODUCT_HTML, item=item, stats=stats, moves=moves)

@app.route('/inventory/report/daily')
@login_required
def inventory_daily_report():
    d = (request.args.get('date') or date.today().isoformat()).strip()
    try:
        inventory.movements.reload()
        mv = inventory.movements.filter_by_date(d).copy()
    except Exception:
        mv = pd.DataFrame(columns=InventoryMovementStore.COLS)

    if mv.empty:
        flash('لا توجد حركات لهذا اليوم', 'err')
        return redirect(url_for('inventory_home'))

    mv['Delta'] = pd.to_numeric(mv.get('Delta'), errors='coerce').fillna(0).astype(int)

    summary = (mv.groupby(['Product Code','Product Name','Movement Type'])['Delta']
                 .sum().reset_index())
    piv = summary.pivot_table(index=['Product Code','Product Name'],
                              columns='Movement Type',
                              values='Delta',
                              aggfunc='sum',
                              fill_value=0).reset_index()

    out_dir = Path(user_data_dir()) / 'reports'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"inventory_report_{d}.xlsx"

    with pd.ExcelWriter(str(out_path), engine='openpyxl') as writer:
        mv.to_excel(writer, index=False, sheet_name='Movements')
        piv.to_excel(writer, index=False, sheet_name='Summary')

    return send_from_directory(str(out_dir), out_path.name, as_attachment=True)

@app.route('/inventory/stagnant')
@login_required
def inventory_stagnant():
    try:
        inventory.reload()
        inventory.movements.reload()
    except Exception:
        pass

    try:
        days = int(request.args.get('days') or 30)
    except Exception:
        days = 30

    today = date.today()
    rows = inventory.df.fillna('').to_dict(orient='records')
    mv = inventory.movements.df.copy() if hasattr(inventory, 'movements') else pd.DataFrame(columns=InventoryMovementStore.COLS)

    mv['Delta'] = pd.to_numeric(mv.get('Delta'), errors='coerce').fillna(0).astype(int)
    mv['Movement Type'] = mv.get('Movement Type','').astype(str)
    mv['Date'] = mv.get('Date','').astype(str)
    mv['Product Code'] = mv.get('Product Code','').astype(str)

    stale = []
    for r in rows:
        code = str(r.get('Product Code','')).strip()
        qty = int(pd.to_numeric(r.get('Quantity',0), errors='coerce') or 0)
        if qty <= 0 or not code:
            continue

        wd = mv[(mv['Product Code'] == code) & (mv['Movement Type'].str.lower() == 'withdraw') & (mv['Delta'] < 0)]
        last_date = ''
        if not wd.empty:
            try:
                last_date = wd['Date'].astype(str).max()
            except Exception:
                last_date = ''

        if last_date:
            try:
                dt = datetime.strptime(last_date, '%Y-%m-%d').date()
                age = (today - dt).days
            except Exception:
                age = 9999
        else:
            age = 9999

        if age >= days:
            stale.append({
                'Product Code': code,
                'Product Name': r.get('Product Name',''),
                'Quantity': qty,
                'Days Since Last Withdraw': age,
            })

    stale = sorted(stale, key=lambda x: (-x['Days Since Last Withdraw'], -x['Quantity']))[:200]

    STAGNANT_HTML = r'''
    {% extends 'base.html' %}
    {% block content %}
    <div class="d-flex justify-content-between align-items-center mb-3">
      <div>
        <h5 class="mb-0">البضاعة الراكدة</h5>
        <div class="text-muted small">منتجات لم يتم سحبها من المخزن لمدة {{ days }} يوم أو أكثر (أو لم تُسحب أبداً).</div>
      </div>
      <a class="btn btn-outline-secondary" href="{{ url_for('inventory_home') }}">رجوع</a>
    </div>

    <form class="row g-2 mb-2" method="get">
      <div class="col-auto">
        <input name="days" type="number" class="form-control" value="{{ days }}" min="1">
      </div>
      <div class="col-auto">
        <button class="btn btn-dark">تحديث</button>
      </div>
    </form>

    <div class="card p-3">
      <div class="table-responsive">
        <table class="table table-striped align-middle mb-0">
          <thead>
            <tr>
              <th>الكود</th>
              <th>المنتج</th>
              <th>الكمية</th>
              <th>عدد الأيام بدون سحب</th>
              <th></th>
            </tr>
          </thead>
          <tbody>
            {% for r in rows %}
            <tr>
              <td class="fw-bold">{{ r['Product Code'] }}</td>
              <td>{{ r['Product Name'] }}</td>
              <td>{{ r['Quantity'] }}</td>
              <td>{{ r['Days Since Last Withdraw'] }}</td>
              <td class="text-end">
                <a class="btn btn-sm btn-outline-primary" href="{{ url_for('inventory_product', code=r['Product Code']) }}">تفاصيل</a>
              </td>
            </tr>
            {% endfor %}
            {% if not rows %}
            <tr><td colspan="5" class="text-center text-muted py-3">لا توجد بضاعة راكدة ضمن هذا الشرط.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
    {% endblock %}
    '''
    return render_template_string(STAGNANT_HTML, rows=stale, days=days)


@app.route('/products')
@login_required
def products_home():
    try:
        inventory.reload()
    except Exception:
        pass
    rows = inventory.df.fillna("").to_dict(orient='records')
    return render_template_string(PRODUCTS_HTML, rows=rows)

@app.route('/products/delete/<code>')
@login_required
def products_delete(code):
    deleted = inventory.delete_item(code)
    if deleted:
        flash('تم حذف المنتج', 'ok')
    else:
        flash('المنتج غير موجود', 'err')
    return redirect(url_for('products_home'))


@app.route('/products/add', methods=['POST'])
@login_required
def products_add():
    # إعادة استخدام نفس منطق inventory_add لضمان توحيد البيانات
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('يرجى إدخال اسم المنتج', 'err'); return redirect(url_for('products_home'))

    row = {
        'Product Code': inventory.next_code(),
        'Product Name': name,
        'Type': (request.form.get('type') or '').strip(),
        'Quantity': int(request.form.get('qty') or 0),
        'Fabric Meters': float(request.form.get('fabric') or 0),
        'Meters per Unit': float(request.form.get('mpu') or 0),
        'Fabric Meter Price': float(request.form.get('fabric_price') or 0),
        'Sewing Cost': float(request.form.get('sew') or 0),
        'Accessories Cost': float(request.form.get('access') or 0),
        'Extra Costs': float(request.form.get('extra') or 0),
        'Sale Price': float(request.form.get('price') or 0),
    }
    inventory.add_item(row)
    flash('تم حفظ المنتج وإضافته للمخزن', 'ok')
    return redirect(url_for('products_home'))


@app.route('/inventory/add', methods=['POST'])
@login_required
def inventory_add():
    """
    Add qty to existing product (by code) without creating duplicates.
    If code not found, create minimal item then add.
    """
    code = (request.form.get('code') or request.form.get('name') or '').strip()
    pname = (request.form.get('pname') or '').strip()
    ptype = (request.form.get('type') or '').strip()

    try:
        qty = int(request.form.get('qty') or 0)
    except Exception:
        qty = 0

    if not code:
        flash('يرجى اختيار المنتج', 'err')
        return redirect(url_for('inventory_home'))

    item = inventory.get_by_code(code)

    # If not exists, create it (minimal fields)
    if not item:
        if not pname:
            pname = code
        row = {
            'Product Code': code if code else inventory.next_code(),
            'Product Name': pname,
            'Type': ptype,
            'Quantity': 0,
            'Fabric Meters': 0.0,
            'Meters per Unit': 0.0,
            'Fabric Meter Price': 0.0,
            'Sewing Cost': 0.0,
            'Accessories Cost': 0.0,
            'Extra Costs': 0.0,
            'Sale Price': 0.0,
        }
        inventory.add_item(row)

    if qty != 0:
        ok, info = inventory.adjust_quantity(code, qty, movement_type='Production', ref='MANUAL_ADD', notes='inventory_add')
        if not ok:
            flash(str(info), 'err')
            return redirect(url_for('inventory_home'))
        applied = int(info.get("applied", qty))
        flash('تم تحديث المخزن', 'ok')
        return redirect(url_for('inventory_home', added=str(max(applied,0)), taken=str(abs(min(applied,0))), name=str(info.get("name") or pname or code)))
    else:
        flash('تم حفظ المنتج (بدون تغيير كمية)', 'ok')
        return redirect(url_for('inventory_home'))

@app.route('/inventory/edit/<code>', methods=['GET', 'POST'])
@login_required
def inventory_edit(code):
    code = str(code).strip()
    item = inventory.get_by_code(code)
    if not item:
        flash('المنتج غير موجود في المخزن', 'err')
        return redirect(url_for('inventory_home'))

    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        typ = (request.form.get('type') or '').strip()
        qty = request.form.get('qty')
        buy = request.form.get('buy')
        extra = request.form.get('extra')
        sell = request.form.get('sell')

        def _num(x, is_int=False):
            try:
                if x is None or str(x).strip() == '':
                    return 0 if is_int else 0.0
                return int(float(x)) if is_int else float(x)
            except Exception:
                return 0 if is_int else 0.0

        updates = {
            "Product Name": name if name else item.get("Product Name", ""),
            "Type": typ if typ else item.get("Type", ""),
            "Quantity": _num(qty, is_int=True),
            "Buying Price": _num(buy, is_int=False),
            "Extra Costs": _num(extra, is_int=False),
            "Selling Price": _num(sell, is_int=False),
        }

        ok = inventory.update_item(code, **updates)
        if ok:
            flash('تم تعديل المنتج', 'ok')
        else:
            flash('فشل التعديل', 'err')
        return redirect(url_for('inventory_home'))

    EDIT_INV_HTML = r'''
    {% extends 'base.html' %}
    {% block content %}
    <div class="row justify-content-center">
      <div class="col-md-7">
        <div class="card p-4 mt-3">
          <h5 class="mb-3">تعديل منتج المخزن</h5>
          <form method="post" class="row g-2">
            <div class="col-md-4">
              <label class="form-label">Product Code</label>
              <input class="form-control" value="{{ item['Product Code'] }}" disabled>
            </div>
            <div class="col-md-8">
              <label class="form-label">اسم المنتج</label>
              <input name="name" class="form-control" value="{{ item['Product Name'] }}">
            </div>

            <div class="col-md-4">
              <label class="form-label">النوع</label>
              <input name="type" class="form-control" value="{{ item.get('Type','') }}">
            </div>
            <div class="col-md-4">
              <label class="form-label">الكمية</label>
              <input name="qty" type="number" class="form-control" value="{{ item.get('Quantity',0) }}">
            </div>
            <div class="col-md-4">
              <label class="form-label">سعر البيع</label>
              <input name="sell" type="number" class="form-control" value="{{ item.get('Selling Price',0) }}">
            </div>

            <div class="col-md-6">
              <label class="form-label">سعر الشراء</label>
              <input name="buy" type="number" class="form-control" value="{{ item.get('Buying Price',0) }}">
            </div>
            <div class="col-md-6">
              <label class="form-label">تكاليف إضافية</label>
              <input name="extra" type="number" class="form-control" value="{{ item.get('Extra Costs',0) }}">
            </div>

            <div class="col-12 d-flex gap-2 mt-2">
              <button class="btn btn-primary">حفظ</button>
              <a class="btn btn-outline-secondary" href="{{ url_for('inventory_home') }}">رجوع</a>
            </div>
          </form>
        </div>
      </div>
    </div>
    {% endblock %}
    '''
    return render_template_string(EDIT_INV_HTML, item=item)
@app.route('/inventory/adjust-bulk', methods=['POST'])
@login_required
def inventory_adjust_bulk():
    code = (request.form.get('code') or request.form.get('name') or '').strip()
    try:
        qty_in = int(request.form.get('qty') or 0)
    except Exception:
        qty_in = 0

    if not code or qty_in == 0:
        flash('يرجى إدخال كود المنتج والكمية', 'err')
        return redirect(url_for('inventory_home'))

    ok, info = inventory.adjust_quantity(code, qty_in, movement_type='Manual', ref='MANUAL', notes='inventory_adjust_bulk')
    if not ok:
        flash(str(info), 'err')
        return redirect(url_for('inventory_home'))

    applied = int(info.get("applied", qty_in))
    pname = str(info.get("name", "") or "").strip() or code

    flash('تم تعديل الكمية', 'ok')
    if applied > 0:
        return redirect(url_for('inventory_home', added=str(applied), name=pname))
    else:
        return redirect(url_for('inventory_home', taken=str(abs(applied)), name=pname))

@app.route('/seamstresses')
@login_required
def seam_home():
    # الخياطات
    seamstresses_df = seams.mast.fillna('')
    seamstresses = seamstresses_df.to_dict(orient='records')
    seam_name_map = {r['ID']: r['Name'] for _, r in seamstresses_df.iterrows()}

    # قيم الفلتر من الـ query string
    dfrom = request.args.get('from') or ''
    dto = request.args.get('to') or ''
    sel_sid = request.args.get('sid') or ''
    sel_paid = request.args.get('paid') or ''

    logs = []
    if hasattr(seams, 'log') and isinstance(seams.log, pd.DataFrame) and not seams.log.empty:
        logs_df = seams.log.copy().fillna('')

        # تحويل التاريخ لنوع datetime حتى نفلتر صح
        logs_df['Date'] = pd.to_datetime(logs_df['Date'], errors='coerce')

        if dfrom:
            start = datetime.strptime(dfrom, '%Y-%m-%d')
            logs_df = logs_df[logs_df['Date'] >= start]
        if dto:
            end = datetime.strptime(dto, '%Y-%m-%d')
            logs_df = logs_df[logs_df['Date'] <= end]

        if sel_sid:
            try:
                sid_int = int(sel_sid)
                logs_df = logs_df[logs_df['SeamstressID'] == sid_int]
            except Exception:
                pass

        if sel_paid in ('paid', 'unpaid'):
            if sel_paid == 'paid':
                logs_df = logs_df[logs_df['Paid'] == True]
            else:
                logs_df = logs_df[logs_df['Paid'] == False]

        logs_df = logs_df.sort_values(by='Date', ascending=False)
        # تنسيق التاريخ للعرض
        logs_df['Date'] = logs_df['Date'].dt.strftime('%Y-%m-%d')
        logs = logs_df.to_dict(orient='records')

    products_df = inventory.df.fillna('')
    products = products_df.to_dict(orient='records')

    return render_template_string(
        SEAMSTRESS_HTML,
        seamstresses=seamstresses,
        logs=logs,
        seam_name_map=seam_name_map,
        dfrom=dfrom,
        dto=dto,
        sel_sid=sel_sid,
        sel_paid=sel_paid,
        products=products,
    )

@app.route('/seam/add', methods=['POST'])
@login_required
def seam_add():
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('يرجى إدخال الاسم', 'err'); return redirect(url_for('home'))
    seams.add_seamstress(name, (request.form.get('phone') or '').strip(), (request.form.get('notes') or '').strip())
    flash('تمت الإضافة', 'ok'); return redirect(url_for('home'))

@app.route('/seam/edit', methods=['POST'])
@login_required
def seam_edit():
    try:
        sid = int(request.form.get('id') or 0)
    except Exception:
        sid = 0
    if not sid:
        flash('معرّف غير صالح', 'err'); return redirect(url_for('home'))
    seams.update_seamstress(sid, Name=request.form.get('name', ''), Phone=request.form.get('phone', ''), Notes=request.form.get('notes', ''), Active=bool(request.form.get('active')))
    flash('تم الحفظ', 'ok'); return redirect(url_for('home'))

@app.route('/seam/delete/<int:sid>')
@login_required
def seam_delete(sid):
    seams.delete_seamstress(sid)
    flash('تم الحذف', 'ok'); return redirect(url_for('home'))

@app.route('/sew/add', methods=['POST'])
@login_required
def sew_add_log():
    try:
        sid = int(request.form.get('sid') or 0)
        pieces = int(request.form.get('pieces') or 0)
        unit = float(request.form.get('unit') or 0)
    except Exception:
        flash('بيانات غير صالحة', 'err'); return redirect(url_for('home'))
    model = (request.form.get('model') or '').strip()
    if not sid or not model or pieces<=0:
        flash('الرجاء إدخال الخياطة، الموديل، وعدد صحيح', 'err'); return redirect(url_for('home'))
    seams.add_log(sid, model, pieces, unit)
    flash('تم تسجيل الإنجاز وزيادة المخزون', 'ok')
    return redirect(url_for('home'))

@app.route('/sew/paid/<int:log_id>')
@login_required
def sew_mark_paid(log_id):
    seams.set_paid(log_id, True); flash('تمت التصفية', 'ok'); return redirect(url_for('home'))

@app.route('/sew/unpaid/<int:log_id>')
@login_required
def sew_mark_unpaid(log_id):
    seams.set_paid(log_id, False); flash('تم الإلغاء', 'ok'); return redirect(url_for('home'))

# ------------------------------ ISSUES ROUTES ---------------------------
@app.route('/issues')
@login_required
def issues_home():
    rows = issues.df.fillna('').sort_values(by='CreatedAt', ascending=False).to_dict(orient='records') if not issues.df.empty else []
    return render_template_string(ISSUES_HTML, rows=rows)

@app.route('/issues/add', methods=['POST'])
@login_required
@limiter.limit('20/minute')
def issues_add():
    title = (request.form.get('title') or '').strip()
    if not title:
        flash('العنوان مطلوب', 'err'); return redirect(url_for('issues_home'))
    img = request.files.get('image')
    img_path = _save_image(img)
    desc = (request.form.get('desc') or '').strip()

    issues.add_issue(title, desc, img_path)

    # 🔔 إشعار تلغرام عند إضافة مشكلة جديدة
    try:
        msg = (
            "⚠️ تم تسجيل مشكلة جديدة\n"
            f"العنوان: {title}\n"
            f"الوصف: {desc or 'لا يوجد'}\n"
            f"الوقت: {now_str()}"
        )
        send_telegram(msg)
    except Exception:
        pass

    flash('تمت إضافة المشكلة', 'ok')
    return redirect(url_for('issues_home'))


@app.route('/issues/solve', methods=['POST'])
@login_required
def issues_solve():
    try:
        iid = int(request.form.get('id') or 0)
    except Exception:
        iid = 0
    solver = (request.form.get('solver') or '').strip()
    if not iid or not solver:
        flash('بيانات غير مكتملة', 'err'); return redirect(url_for('issues_home'))
    issues.solve(iid, solver)
    flash('تم الحل', 'ok'); return redirect(url_for('issues_home'))

@app.route('/issues/delete/<int:iid>')
@login_required
def issues_delete(iid):
    issues.delete(iid); flash('تم الحذف', 'ok'); return redirect(url_for('issues_home'))

@app.route('/static-proxy')
@login_required
def static_proxy():
    # لعرض الصور المخزّنة خارج static
    from flask import send_file, request as _rq
    f = _rq.args.get('f')
    if not f or not Path(f).exists():
        return ('', 404)
    return send_file(f)

# ------------------------------ CUTTINGS STORE --------------------------
class CuttingsStore:
    COLS = ['ID', 'Model', 'ImagePath', 'DueDate', 'RequiredQty',
            'Status', 'Notes', 'RejectionReason', 'CreatedAt']

    def __init__(self, root_dir: Path):
        self.path = root_dir / 'cuttings.xlsx'
        self.df = self._load()

    def _load(self):
        if not self.path.exists():
            df = pd.DataFrame(columns=self.COLS)
            try:
                df.to_excel(self.path, index=False)
            except Exception:
                # لو ما قدرنا نكتب إكسل (مكتبة/صلاحيات) نخليها بالذاكرة فقط
                pass
            return df

        # لو الملف تالف/مقفول/ما ينقري، نعيد إنشاءه حتى ما يوقع البرنامج
        try:
            df = pd.read_excel(self.path)
        except Exception:
            df = pd.DataFrame(columns=self.COLS)
            try:
                df.to_excel(self.path, index=False)
            except Exception:
                pass
            return df

        for c in self.COLS:
            if c not in df.columns:
                df[c] = pd.NA
        return df[self.COLS]

    def _save(self):
        # حفظ آمن: نكتب لملف مؤقت ثم نستبدل (لتقليل احتمال تلف الملف)
        tmp_path = self.path.with_suffix('.tmp.xlsx')
        try:
            self.df.to_excel(tmp_path, index=False)
            try:
                os.replace(tmp_path, self.path)
            except PermissionError as e:
                # غالبًا ملف الإكسل مفتوح
                try:
                    tmp_path.unlink(missing_ok=True)
                except Exception:
                    pass
                raise PermissionError("ملف cuttings.xlsx مفتوح. اغلقه ثم حاول مرة أخرى.") from e
        finally:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass

    def _next_id(self):
        if self.df.empty:
            return 1
        vals = pd.to_numeric(self.df['ID'], errors='coerce').dropna()
        return int(vals.max() + 1) if len(vals) else 1

    def add(self, model, due, qty, notes='', img_path=''):
        new_id = self._next_id()
        row = {
            'ID': new_id,
            'Model': model,
            'ImagePath': img_path,
            'DueDate': due,
            'RequiredQty': qty,
            'Status': 'قيد الانتظار',
            'Notes': notes,
            'RejectionReason': '',
            'CreatedAt': now_str(),
        }
        self.df = pd.concat([self.df, pd.DataFrame([row])], ignore_index=True)
        self._save()

    def update_status(self, cid, status, reason=None):
        idx = self.df[self.df['ID'] == cid].index
        if not len(idx):
            return
        i = idx[0]
        self.df.at[i, 'Status'] = status
        if reason is not None:
            self.df.at[i, 'RejectionReason'] = reason
        self._save()

    def delete(self, cid):
        self.df = self.df[self.df['ID'] != cid]
        self._save()


cuttings = CuttingsStore(_data_root)
# ------------------------------ CUTTING ROUTES --------------------------
@app.route('/cutting')
@login_required
def cutting_home():
    # إعادة تحميل بيانات الفصال من ملف الإكسل في كل مرة
    df = cuttings._load()
    cuttings.df = df  # نحدّث النسخة الموجودة في الذاكرة أيضًا

    if not df.empty:
        df = df.fillna('')
        # ترتيب حسب تاريخ الإنشاء من الأحدث إلى الأقدم
        try:
            rows = df.sort_values(by='CreatedAt', ascending=False).to_dict(orient='records')
        except Exception:
            # لو صار أي خطأ في CreatedAt نعرضها بدون ترتيب
            rows = df.to_dict(orient='records')
    else:
        rows = []

    return render_template_string(CUTTING_HTML, rows=rows)

@app.route('/cutting/add', methods=['POST'])
@login_required
def cutting_add():
    model = (request.form.get('model') or '').strip()
    due = (request.form.get('due') or '').strip()
    try:
        qty = int(request.form.get('qty') or 0)
    except Exception:
        qty = 0
    if not model or not due or qty<=0:
        flash('بيانات غير مكتملة', 'err'); return redirect(url_for('cutting_home'))
    img = request.files.get('image')
    img_path = _save_image(img)
    notes = (request.form.get('notes') or '').strip()

    # إضافة الفصال في الإكسل
    try:
        cuttings.add(model, due, qty, notes, img_path)
    except PermissionError as e:
        flash(str(e), 'err')
        return redirect(url_for('cutting_home'))
    except Exception as e:
        _fatal_box('cutting_add', e)
        flash('صار خطأ أثناء حفظ طلب الفصال. راجع error_log.txt', 'err')
        return redirect(url_for('cutting_home'))

    # 🔔 إشعار تلغرام
    try:
        msg = (
            "🧵 تم إنشاء طلب فصال جديد\n"
            f"الموديل: {model}\n"
            f"الكمية المطلوبة: {qty}\n"
            f"موعد الفصال: {due}\n"
            f"ملاحظات: {notes or 'لا يوجد'}\n"
            f"الوقت: {now_str()}"
        )
        send_telegram(msg)
    except Exception:
        pass

    flash('تم إنشاء طلب الفصال', 'ok')
    return redirect(url_for('cutting_home'))

@app.route('/cutting/status/<int:cid>')
@login_required
def cutting_status(cid):
    s = (request.args.get('s') or '').strip()
    if s not in ['قيد الانتظار','قيد العمل','مكتمل','مرفوض']:
        flash('حالة غير صالحة', 'err'); return redirect(url_for('cutting_home'))
    cuttings.update_status(cid, s)
    flash('تم التحديث', 'ok'); return redirect(url_for('cutting_home'))

@app.route('/cutting/reject', methods=['POST'])
@login_required
def cutting_reject():
    try:
        cid = int(request.form.get('id') or 0)
    except Exception:
        cid = 0
    reason = (request.form.get('reason') or '').strip()
    if not cid or not reason:
        flash('بيانات غير مكتملة', 'err'); return redirect(url_for('cutting_home'))
    cuttings.update_status(cid, 'مرفوض', reason)
    flash('تم الرفض', 'ok'); return redirect(url_for('cutting_home'))

@app.route('/cutting/delete/<int:cid>')
@login_required
def cutting_delete(cid):
    cuttings.delete(cid); flash('تم الحذف', 'ok'); return redirect(url_for('cutting_home'))

# --------------------------- ERROR HANDLING ----------------------------

def _fatal_box(title, exc):
    try:
        with open(ERROR_LOG, 'a', encoding='utf-8') as f:
            f.write(f"[{now_str()}] {title}: {type(exc).__name__}: {exc}\n")
            f.write(traceback.format_exc() + "\n")
    except Exception:
        pass


# ------------------------------ RUN -----------------------------------


# ------------------------------ POS (Alias) ------------------------------
# بعض القوالب قد تشير إلى endpoint اسمه 'pos'. لو لم تكن صفحة POS كاملة مفعّلة هنا،
# نعطي alias بسيط يوجّه لصفحة إضافة/إدارة المنتجات حتى لا يحدث BuildError.
if 'pos' not in app.view_functions:
    @app.route('/pos')
    @login_required
    def pos():
        return redirect(url_for('products_add'))


if __name__ == '__main__':
    app.run(debug=True)
